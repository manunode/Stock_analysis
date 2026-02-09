#!/usr/bin/env python3
"""
Trendlyne Stock Analyzer
Parses JSON dumps from Trendlyne and extracts key metrics for stock selection.
Outputs a CSV with computed scores and red flags for pruning 150 stocks to 35.

Usage:
    python Create_stock_analysis.py                                    # JSONs in current directory
    python Create_stock_analysis.py --data-dir /path/to/financials     # Specify financial JSON directory
    python Create_stock_analysis.py --shareholding-dir /path/to/sh     # Separate shareholding directory
    python Create_stock_analysis.py --data-dir ./financials --shareholding-dir ./shareholding
"""

import json
import os
import csv
import sys
import re
import argparse
import time
from datetime import datetime
from pathlib import Path
from statistics import mean
from concurrent.futures import ProcessPoolExecutor, as_completed
from typing import Optional, Dict, List, Any

# Configuration (defaults - can be overridden via command line)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "Financials")  # Directory containing financial JSONs
SHAREHOLDING_DIR = os.path.join(BASE_DIR, "Shareholding")  # Directory containing shareholding JSONs (None = same as DATA_DIR)
CORPORATE_ACTIONS_DIR = os.path.join(BASE_DIR, "CorporateActions")  # Directory containing dividend/bonus/split JSONs
OVERVIEW_DIR = os.path.join(BASE_DIR, "Overview")  # Directory containing overview JSONs (TTM metrics, Piotroski)
INSIDER_TRADING_DIR = os.path.join(BASE_DIR, "InsiderTrading")  # Directory containing insider trading JSONs
OUTPUT_FILE = os.path.join(BASE_DIR, "stock_analysis.xlsx")
RED_FLAGS_FILE = os.path.join(BASE_DIR, "red_flags_detail.csv")

# Years we care about (most recent first)
YEARS_ORDER = ["Mar 2025", "Mar 2024", "Mar 2023", "Mar 2022", "Mar 2021"]

# Month abbreviation to number for sorting non-March fiscal years
_MONTH_NUM = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}

def extract_available_years(yearly_data: dict, max_years: int = 5) -> List[str]:
    """
    Dynamically extract and sort fiscal year keys from yearly_data.
    
    Handles any fiscal year end month (Mar, Jun, Dec, etc.).
    Returns keys sorted newest-first, excluding 'TTM'.
    
    Most Indian companies use March FY (matched by YEARS_ORDER),
    but ~5-8% use June/December (e.g., Gillette, Colgate, HUL).
    Without this, those companies are silently dropped with
    'Insufficient data' despite having 10+ years of data.
    """
    year_keys = [k for k in yearly_data.keys() if k != 'TTM']
    
    def sort_key(yk: str):
        """Parse 'Mon YYYY' → (year, month) for chronological sorting."""
        parts = yk.split()
        if len(parts) == 2:
            month_str, year_str = parts
            month = _MONTH_NUM.get(month_str, 0)
            try:
                year = int(year_str)
            except ValueError:
                year = 0
            return (year, month)
        return (0, 0)
    
    # Sort newest first
    year_keys.sort(key=sort_key, reverse=True)
    
    # Limit to max_years to keep analysis window consistent
    return year_keys[:max_years]

# Red Flag Definitions - what each flag means and why it matters
# Split into STRUCTURAL (business quality issues) vs PRICING (valuation concerns)
# SEVERITY TIERS: CRITICAL (2.0), MAJOR (1.0), MINOR (0.5)
# - CRITICAL: Survival risk, cash burn, unsustainable
# - MAJOR: Significant quality concern, likely structural
# - MINOR: May be cyclical or sector-normal, needs context

STRUCTURAL_RED_FLAGS = {
    "LOW_ROE": {
        "name": "Low Return on Equity",
        "trigger": "ROE < 10%",
        "meaning": "Company generates weak returns on shareholder capital",
        "concern": "Poor capital efficiency; equity is not being deployed productively",
        "severity": "MAJOR",
        "weight": 1.0,
    },
    "DECLINING_ROE": {
        "name": "Declining ROE to Mediocre",
        "trigger": "ROE trend slope < -2 AND Latest ROE < 15%",
        "meaning": "Return on equity falling AND now at mediocre level",
        "concern": "Business quality eroding (ignores mean reversion from exceptional levels)",
        "severity": "MAJOR",
        "weight": 1.0,
    },
    "LOW_ROCE": {
        "name": "Low Return on Capital Employed",
        "trigger": "ROCE < 10%",
        "meaning": "Company earns weak returns on total capital (debt + equity)",
        "concern": "Business doesn't earn enough to justify its capital base",
        "severity": "MAJOR",
        "weight": 1.0,
    },
    "DECLINING_ROCE": {
        "name": "Declining ROCE to Mediocre",
        "trigger": "ROCE trend slope < -2 AND Latest ROCE < 15%",
        "meaning": "Return on capital employed falling AND now at mediocre level",
        "concern": "Capital efficiency worsening (ignores mean reversion from exceptional levels)",
        "severity": "MAJOR",
        "weight": 1.0,
    },
    "POOR_CASH_CONVERSION": {
        "name": "Poor Cash Conversion",
        "trigger": "3-year avg CFO/PAT < 0.7",
        "meaning": "Reported profits are not converting to actual cash",
        "concern": "Earnings quality suspect; possible aggressive accounting or working capital issues",
        "severity": "CRITICAL",
        "weight": 2.0,
    },
    "NEGATIVE_CFO": {
        "name": "Negative Operating Cash Flow",
        "trigger": "Latest CFO < 0",
        "meaning": "Core operations are burning cash, not generating it",
        "concern": "Business may need external funding to survive; unsustainable",
        "severity": "CRITICAL",
        "weight": 2.0,
    },
    "INCONSISTENT_CFO": {
        "name": "Inconsistent Cash Flow",
        "trigger": "Multiple years with negative CFO",
        "meaning": "Operating cash flow is not reliably positive",
        "concern": "Lumpy or unreliable business; hard to value or trust reported earnings",
        "severity": "CRITICAL",
        "weight": 2.0,
    },
    "FREQUENT_EXCEPTIONALS": {
        "name": "Frequent Exceptional Items",
        "trigger": "Exceptional items in 2+ years",
        "meaning": "Company repeatedly reports one-time gains/losses",
        "concern": "May be using exceptionals to manage earnings; core profitability unclear",
        "severity": "MAJOR",
        "weight": 1.0,
    },
    "HIGH_OTHER_INCOME": {
        "name": "High Other Income Dependency",
        "trigger": "Other income > 30% of PAT for 2+ years, OR > 50% in latest year",
        "meaning": "Significant portion of profit consistently comes from non-core activities",
        "concern": "Not a real operating business; may be sector-normal for holding companies",
        "severity": "MAJOR",
        "weight": 1.0,
        "sector_sensitive": True,  # May be normal for holding companies
    },
    "RISING_RECEIVABLES": {
        "name": "Rising Receivable Days",
        "trigger": "Receivable days trend > 10 days/year increase",
        "meaning": "Company taking longer to collect payments from customers",
        "concern": "Cash conversion cycle worsening; possible channel stuffing or weak customers",
        "severity": "MINOR",
        "weight": 0.5,
        "sector_sensitive": True,  # IT services normally have 60-90 day cycles
    },
    "RISING_INVENTORY": {
        "name": "Rising Inventory Days",
        "trigger": "Inventory days trend > 10 days/year increase",
        "meaning": "Company holding more inventory relative to sales",
        "concern": "Potential obsolescence risk, demand slowdown, or channel stuffing shifted to inventory",
        "severity": "MINOR",
        "weight": 0.5,
    },
    "MARGIN_COMPRESSION": {
        "name": "Margin Compression",
        "trigger": "Operating margin trend < -1% per year",
        "meaning": "Profit margins are shrinking over time",
        "concern": "Pricing power weakening; cost pressures; may be cyclical trough",
        "severity": "MINOR",
        "weight": 0.5,
        "sector_sensitive": True,  # May be cyclical, not structural
    },
    "RISING_DEBT": {
        "name": "Rising Debt Levels",
        "trigger": "Debt/Equity trend > 0.05 per year",
        "meaning": "Company is taking on more leverage over time",
        "concern": "Financial risk increasing; may indicate cash flow problems or aggressive expansion",
        "severity": "CRITICAL",
        "weight": 2.0,
    },
    "WC_DIVERGENCE": {
        "name": "Working Capital Divergence",
        "trigger": "WC Growth > 1.5x Revenue Growth",
        "meaning": "Working capital growing much faster than sales",
        "concern": "Company may be 'buying' sales via channel stuffing or unpaid bills; cash trap",
        "severity": "MINOR",
        "weight": 0.5,
    },
    "NPM_OPM_DIVERGENCE": {
        "name": "Net vs Operating Margin Divergence",
        "trigger": "NPM growth exceeds OPM growth by >5%",
        "meaning": "Bottom line improving faster than operating performance",
        "concern": "Earnings quality suspect; growth from tax maneuvers, interest savings, or other income",
        "severity": "MAJOR",
        "weight": 1.0,
    },
}

PRICING_RED_FLAGS = {
    "HIGH_PE": {
        "name": "High P/E Ratio",
        "trigger": "P/E > 50",
        "meaning": "Stock is priced at very high multiple of earnings",
        "concern": "Expensive valuation; high expectations baked in; limited margin of safety",
        "severity": "MINOR",
        "weight": 0.5,
    },
    "NEGATIVE_PE": {
        "name": "Negative P/E (Loss-Making)",
        "trigger": "P/E < 0",
        "meaning": "Company is currently unprofitable",
        "concern": "No earnings to value; speculative investment",
        "severity": "MAJOR",
        "weight": 1.0,
    },
    "HIGH_EV_EBITDA": {
        "name": "High EV/EBITDA",
        "trigger": "EV/EBITDA > 25",
        "meaning": "Enterprise value is very high relative to operating profit",
        "concern": "Expensive on cash flow basis; high expectations or low profitability",
        "severity": "MINOR",
        "weight": 0.5,
    },
    "NEGATIVE_EBITDA": {
        "name": "Negative EBITDA",
        "trigger": "EV/EBITDA < 0 (implies EBITDA < 0)",
        "meaning": "Company loses money at operating level before interest, taxes, depreciation",
        "concern": "Cannot value via standard cash-flow metrics; burning cash; survival risk",
        "severity": "CRITICAL",
        "weight": 2.0,
    },
    "HIGH_PBV_ROE": {
        "name": "High PBV Relative to ROE",
        "trigger": "PBV > (ROE / 2)",
        "meaning": "Stock price implies returns the company cannot deliver",
        "concern": "Overvalued relative to actual return generation; PBV of 3 with ROE of 8% is expensive",
        "severity": "MINOR",
        "weight": 0.5,
    },
}

# Combined for backward compatibility
RED_FLAG_DEFINITIONS = {**STRUCTURAL_RED_FLAGS, **PRICING_RED_FLAGS}

# Pre-computed financial sector set for O(1) lookups instead of O(n) per stock
FINANCIAL_SECTORS_LOWER = frozenset(s.lower() for s in [
    'Banks', 'Private Banks', 'Public Banks', 'Foreign Banks',
    'Finance', 'Financial Services', 'NBFC', 'Non Banking Financial Company',
    'Insurance', 'Life Insurance', 'General Insurance',
    'Housing Finance', 'Consumer Finance', 'Asset Management',
    'Diversified Financials', 'Financial Institution',
])


def build_companion_file_index(data_dir: str, shareholding_dir: Optional[str] = None,
                                corporate_actions_dir: Optional[str] = None,
                                overview_dir: Optional[str] = None,
                                insider_trading_dir: Optional[str] = None) -> Dict[str, Dict[str, Path]]:
    """
    Pre-index all companion files (shareholding, dividend, overview, insider_trading) ONCE at startup.
    Returns dict mapping stock_code -> {'shareholding': Path, 'dividend': Path, 'overview': Path, 'insider_trading': Path}
    
    This eliminates ~4 filesystem stat() calls per stock per companion type.
    At 4,500 stocks that's ~18,000 stat calls replaced by 4 directory scans.
    """
    index = {}  # stock_code -> {type: filepath}
    
    def scan_dir(directory, suffix, file_type):
        """Scan a directory and index files by stock code."""
        if not directory or not os.path.exists(directory):
            return
        for f in Path(directory).glob(f"*{suffix}.json"):
            # Extract stock code: "RELIANCE_shareholding.json" -> "RELIANCE"
            code = f.stem.replace(suffix, '')
            # Also handle alternate patterns: "RELIANCE_new_shareholding" -> "RELIANCE_new"
            if code not in index:
                index[code] = {}
            index[code][file_type] = f
            # Also index without _new/_consolidated/_standalone suffixes
            base_code = code.replace('_new', '').replace('_consolidated', '').replace('_standalone', '')
            if base_code != code:
                if base_code not in index:
                    index[base_code] = {}
                # Don't overwrite if exact match already exists
                if file_type not in index[base_code]:
                    index[base_code][file_type] = f
    
    # Scan dedicated directories
    scan_dir(shareholding_dir, '_shareholding', 'shareholding')
    scan_dir(corporate_actions_dir, '_dividend', 'dividend')
    scan_dir(overview_dir, '_overview', 'overview')
    scan_dir(insider_trading_dir, '_insider_trading', 'insider_trading')
    
    # Also scan data_dir for companion files co-located with financial JSONs
    scan_dir(data_dir, '_shareholding', 'shareholding')
    scan_dir(data_dir, '_dividend', 'dividend')
    scan_dir(data_dir, '_overview', 'overview')
    scan_dir(data_dir, '_insider_trading', 'insider_trading')
    
    return index


def safe_get(data: dict, key: str, default=None):
    """Safely get a value, returning default if None or missing."""
    val = data.get(key)
    return val if val is not None else default


def safe_float(val, default=None) -> Optional[float]:
    """Convert to float safely."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def round_dict_values(d: dict, decimals: int = 2) -> dict:
    """Round all float values in a dictionary to specified decimal places."""
    result = {}
    for key, val in d.items():
        if isinstance(val, float):
            result[key] = round(val, decimals)
        else:
            result[key] = val
    return result


def compute_trend(values: List[Optional[float]]) -> Optional[float]:
    """
    Compute trend using Theil-Sen median-of-slopes estimator.
    Positive = improving, Negative = deteriorating.
    values should be in order [newest, ..., oldest]
    
    Theil-Sen is robust to outliers: instead of fitting a single OLS line
    (where one exceptional year can dominate), it computes the slope between
    every pair of points and takes the median. A single outlier year (tax
    refund, asset sale, one-off write-down) cannot swing the result.
    
    Requires minimum 3 valid data points for reliability.
    """
    # Filter out None values and reverse so oldest is first
    valid = [(i, v) for i, v in enumerate(reversed(values)) if v is not None]
    
    # Require at least 3 data points
    if len(valid) < 3:
        return None
    
    x_vals = [i for i, _ in valid]
    y_vals = [v for _, v in valid]
    
    # Compute pairwise slopes between all distinct point pairs
    pairwise_slopes = []
    n = len(x_vals)
    for i in range(n):
        for j in range(i + 1, n):
            dx = x_vals[j] - x_vals[i]
            if dx != 0:
                pairwise_slopes.append((y_vals[j] - y_vals[i]) / dx)
    
    if not pairwise_slopes:
        return 0.0
    
    # Median of all pairwise slopes = Theil-Sen estimator
    pairwise_slopes.sort()
    mid = len(pairwise_slopes) // 2
    if len(pairwise_slopes) % 2 == 0:
        slope = (pairwise_slopes[mid - 1] + pairwise_slopes[mid]) / 2
    else:
        slope = pairwise_slopes[mid]
    
    return round(slope, 2)


def compute_cagr(start_val: Optional[float], end_val: Optional[float], years: int) -> Optional[float]:
    """Compute CAGR given start, end values and number of years."""
    if start_val is None or end_val is None or years <= 0:
        return None
    if start_val <= 0 or end_val <= 0:
        return None
    try:
        return ((end_val / start_val) ** (1 / years) - 1) * 100
    except (ValueError, TypeError, ZeroDivisionError, OverflowError):
        return None


def load_shareholding_data(financial_filepath: str, shareholding_dir: Optional[str] = None,
                           resolved_path: Optional[Path] = None) -> Optional[Dict[str, Any]]:
    """
    Load shareholding data from companion JSON file.
    If resolved_path is provided (from pre-built file index), skip filesystem discovery.
    Otherwise looks for {stock_code}_shareholding.json in:
      1. shareholding_dir (if specified)
      2. Same directory as the financial JSON (if shareholding_dir is None)
    """
    if resolved_path is not None:
        shareholding_file = resolved_path
    else:
        filepath = Path(financial_filepath)
        stock_code = filepath.stem
        
        # Determine where to look for shareholding file
        if shareholding_dir:
            search_dir = Path(shareholding_dir)
        else:
            search_dir = filepath.parent
        
        # Try to find shareholding file
        shareholding_file = search_dir / f"{stock_code}_shareholding.json"
        
        if not shareholding_file.exists():
            # Try alternate patterns
            base_code = stock_code.replace('_new', '').replace('_consolidated', '').replace('_standalone', '')
            if base_code != stock_code:
                shareholding_file = search_dir / f"{base_code}_shareholding.json"
            
            if not shareholding_file.exists():
                return None
    
    try:
        with open(shareholding_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        shareholding = data.get('shareholding', {})
        current = shareholding.get('current_holdings', {})
        
        result = {
            'Promoter_Holding': current.get('Promoter'),
            'FII_Holding': current.get('FII'),
            'MF_Holding': current.get('MF'),
            'Public_Holding': current.get('Public'),
            'Others_Holding': current.get('Others'),
            'Promoter_Pledge': shareholding.get('pledge_percent'),
        }
        
        # Helper function to find holding from ~1 year ago
        # Handles irregular quarter names like "Jul 14, 2025", "Sep 2024", etc.
        def find_year_ago_holding(history):
            """
            Find the holding value from approximately 1 year ago.
            Strategy: Find entry with same quarter name from previous year,
            or fall back to index-based approach for entries ~4 quarters back.
            """
            if not history or len(history) < 2:
                return None
            
            latest_quarter = history[0].get('quarter', '')
            latest_holding = history[0].get('holding')
            
            # Try to parse the latest quarter to extract month/year
            # Format examples: "Dec 2025", "Sep 2024", "Jul 14, 2025"
            
            # Extract year from latest quarter
            year_match = re.search(r'20\d{2}', latest_quarter)
            if not year_match:
                # Fall back to index 4 if we can't parse
                if len(history) >= 5:
                    return history[4].get('holding')
                return None
            
            latest_year = int(year_match.group())
            target_year = latest_year - 1
            
            # Extract month from latest quarter
            month_map = {
                'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
            }
            latest_month = None
            for month_name, month_num in month_map.items():
                if month_name in latest_quarter:
                    latest_month = month_num
                    break
            
            # Search for matching quarter from 1 year ago
            for entry in history[1:]:  # Skip the first (current) entry
                quarter = entry.get('quarter', '')
                year_match = re.search(r'20\d{2}', quarter)
                if not year_match:
                    continue
                    
                entry_year = int(year_match.group())
                
                # Check if this is from ~1 year ago
                if entry_year == target_year:
                    # If we know the month, try to match it
                    if latest_month:
                        for month_name, month_num in month_map.items():
                            if month_name in quarter:
                                # Same month, previous year = perfect match
                                if month_num == latest_month:
                                    return entry.get('holding')
                                # Close enough (within 1 quarter)
                                if abs(month_num - latest_month) <= 3:
                                    return entry.get('holding')
                                break
                    else:
                        # No month info, just use first entry from target year
                        return entry.get('holding')
            
            # Fall back to index-based approach if date matching fails
            # Try to find an entry that's roughly 4 entries back (4 quarters)
            if len(history) >= 5:
                return history[4].get('holding')
            elif len(history) >= 4:
                return history[-1].get('holding')
            
            return None
        
        # Compute promoter holding change (latest vs ~1 year ago)
        promoter_history = shareholding.get('promoter_history', [])
        if promoter_history:
            latest = promoter_history[0].get('holding')
            year_ago = find_year_ago_holding(promoter_history)
            if latest is not None and year_ago is not None:
                result['Promoter_Change_1Yr'] = round(latest - year_ago, 2)
        
        # Compute FII change
        fii_history = shareholding.get('fii_history', [])
        if fii_history:
            latest = fii_history[0].get('holding')
            year_ago = find_year_ago_holding(fii_history)
            if latest is not None and year_ago is not None:
                result['FII_Change_1Yr'] = round(latest - year_ago, 2)
        
        # Compute MF change
        mf_history = shareholding.get('mf_history', [])
        if mf_history:
            latest = mf_history[0].get('holding')
            year_ago = find_year_ago_holding(mf_history)
            if latest is not None and year_ago is not None:
                result['MF_Change_1Yr'] = round(latest - year_ago, 2)
        
        return result
        
    except Exception as e:
        print(f"[WARN] Could not load shareholding data from {shareholding_file}: {e}")
        return None


def load_insider_trading_data(financial_filepath: str, insider_trading_dir: Optional[str] = None,
                              resolved_path: Optional[Path] = None,
                              market_cap_cr: Optional[float] = None) -> Optional[Dict[str, Any]]:
    """
    Load insider trading data from companion JSON file.
    Analyzes transaction patterns to distinguish genuine promoter selling from
    inter-se transfers, family restructuring, and other non-conviction events.
    
    This is CRITICAL context for the Conviction Override (VALUE_TRAP / CONTRARIAN_BET)
    logic. Without it, a promoter family consolidation (inter-se transfer) looks
    identical to insider exit selling in the shareholding data.
    
    Key classifications:
      - INTER_SE: Off-market transfers between promoter entities (family restructuring)
      - MARKET_SALE: Open-market sales (genuine liquidity/exit selling)
      - MARKET_BUY: Open-market purchases (conviction buying)
      - ESOP/GIFT: Off-market gifts or ESOP exercises (not conviction signal)
    
    Returns dict with insider trading metrics, or None if file not found.
    """
    if resolved_path is not None:
        it_file = resolved_path
    else:
        filepath = Path(financial_filepath)
        stock_code = filepath.stem
        
        if insider_trading_dir:
            search_dir = Path(insider_trading_dir)
        else:
            search_dir = filepath.parent
        
        it_file = search_dir / f"{stock_code}_insider_trading.json"
        
        if not it_file.exists():
            base_code = stock_code.replace('_new', '').replace('_consolidated', '').replace('_standalone', '')
            if base_code != stock_code:
                it_file = search_dir / f"{base_code}_insider_trading.json"
            if not it_file.exists():
                return None
    
    try:
        with open(it_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        insider_data = data.get('insider_trading', {})
        summary = insider_data.get('summary', {})
        transactions = insider_data.get('transactions', [])
        
        if not transactions:
            return None
        
        result = {
            'Insider_Trading_Available': True,
            'Insider_Sentiment': summary.get('insider_sentiment', 'UNKNOWN'),
            'Total_Insider_Transactions': summary.get('total_transactions', 0),
        }
        
        # ============================================================
        # CLASSIFY TRANSACTIONS: Inter-se vs Market vs Gift/ESOP
        # ============================================================
        # Inter-se transfers are identified by:
        #   - mode containing "Inter-se", "Off Market", "Gift", or "-"
        #   - value = 0 or null AND avg_price = 0 or null
        #   - Often SAST regulations
        # Market transactions have:
        #   - mode containing "Market Sale" or "Market Purchase" or "Market"
        #   - value > 0 and avg_price > 0
        
        from datetime import datetime, timedelta
        
        now = datetime.now()
        one_year_ago = now - timedelta(days=365)
        six_months_ago = now - timedelta(days=180)
        
        # Counters for promoter/promoter group transactions in last 1 year
        market_buy_qty_1y = 0
        market_buy_value_1y = 0.0
        market_sell_qty_1y = 0
        market_sell_value_1y = 0.0
        inter_se_disposal_qty_1y = 0
        inter_se_acquisition_qty_1y = 0
        total_disposal_qty_1y = 0
        
        # Also track 6-month market buys (stronger recency signal)
        market_buy_qty_6m = 0
        market_buy_value_6m = 0.0
        
        # Track unique market buyers and sellers
        market_buyers_1y = set()
        market_sellers_1y = set()
        
        promoter_categories = {'Promoter', 'Promoter Group', 'Promoter & Director'}
        
        for txn in transactions:
            # Parse transaction date
            reported_date = txn.get('reported_date', '')
            try:
                txn_date = datetime.strptime(reported_date, '%Y-%m-%d')
            except (ValueError, TypeError):
                continue  # Skip unparseable dates
            
            # Only analyze last 1 year of transactions
            if txn_date < one_year_ago:
                continue
            
            category = txn.get('category', '')
            if category not in promoter_categories:
                continue  # Only analyze promoter/promoter group transactions
            
            action = txn.get('action', '')
            mode = str(txn.get('mode', '') or '').lower()
            value = txn.get('value') or 0
            avg_price = txn.get('avg_price') or 0
            quantity = txn.get('quantity', 0) or 0
            client_name = txn.get('client_name', '')
            
            # Classify the transaction
            is_market_txn = False
            is_inter_se = False
            
            if 'market sale' in mode or 'market purchase' in mode:
                is_market_txn = True
            elif 'inter-se' in mode or 'gift' in mode:
                is_inter_se = True
            elif 'off market' in mode:
                # Off-market with zero value = inter-se/gift
                # Off-market with real value = block deal (treat as market signal)
                if value == 0 and avg_price == 0:
                    is_inter_se = True
                else:
                    is_market_txn = True
            elif mode in ['market', 'none', '-', '']:
                # Ambiguous mode — use value/price as discriminator
                if value > 0 and avg_price > 0:
                    is_market_txn = True
                elif value == 0 and avg_price == 0:
                    is_inter_se = True
                # else: genuinely ambiguous, skip
            
            if action == 'Disposal':
                total_disposal_qty_1y += quantity
                
                if is_market_txn:
                    market_sell_qty_1y += quantity
                    market_sell_value_1y += float(value)
                    market_sellers_1y.add(client_name)
                elif is_inter_se:
                    inter_se_disposal_qty_1y += quantity
            
            elif action == 'Acquisition':
                if is_market_txn:
                    market_buy_qty_1y += quantity
                    market_buy_value_1y += float(value)
                    market_buyers_1y.add(client_name)
                    
                    if txn_date >= six_months_ago:
                        market_buy_qty_6m += quantity
                        market_buy_value_6m += float(value)
                elif is_inter_se:
                    inter_se_acquisition_qty_1y += quantity
        
        # ============================================================
        # COMPUTE DERIVED METRICS
        # ============================================================
        
        # What fraction of total disposals were inter-se (not real selling)?
        inter_se_pct_of_disposals = 0.0
        if total_disposal_qty_1y > 0:
            inter_se_pct_of_disposals = round(
                (inter_se_disposal_qty_1y / total_disposal_qty_1y) * 100, 1
            )
        
        # Net market action (excluding inter-se)
        net_market_qty = market_buy_qty_1y - market_sell_qty_1y
        net_market_value = market_buy_value_1y - market_sell_value_1y
        
        # Market sell as % of market cap
        market_sell_pct_of_mcap = None
        if market_cap_cr and market_cap_cr > 0 and market_sell_value_1y > 0:
            market_sell_pct_of_mcap = round(
                (market_sell_value_1y / (market_cap_cr * 1e7)) * 100, 2
            )
        
        # Determine insider action classification
        if market_buy_qty_1y > 0 and market_sell_qty_1y == 0:
            insider_action = "ONLY_BUYING"
        elif market_buy_qty_1y == 0 and market_sell_qty_1y == 0:
            if inter_se_disposal_qty_1y > 0:
                insider_action = "INTER_SE_ONLY"
            else:
                insider_action = "NO_ACTIVITY"
        elif market_buy_qty_1y > 0 and market_sell_qty_1y > 0:
            if net_market_value > 0:
                insider_action = "NET_BUYER"
            else:
                insider_action = "NET_SELLER_WITH_BUYS"
        else:
            insider_action = "NET_SELLER"
        
        # Build context string explaining what happened
        context_parts = []
        if inter_se_pct_of_disposals > 50:
            context_parts.append(
                f"{inter_se_pct_of_disposals:.0f}% of disposals were inter-se transfers (family restructuring)"
            )
        if market_buy_value_1y > 0:
            context_parts.append(
                f"Promoter bought ₹{market_buy_value_1y/1e5:.1f}L in open market"
            )
        if market_sell_value_1y > 0:
            sell_cr = market_sell_value_1y / 1e7
            context_parts.append(
                f"Promoter sold ₹{sell_cr:.1f}Cr in open market"
            )
            if market_sell_pct_of_mcap is not None:
                context_parts.append(f"({market_sell_pct_of_mcap:.1f}% of MCap)")
        if market_buy_qty_6m > 0:
            context_parts.append("Recent market buying (last 6m)")
        
        result.update({
            'Insider_Action': insider_action,
            'Market_Buy_Qty_1Yr': market_buy_qty_1y,
            'Market_Buy_Value_1Yr': round(market_buy_value_1y, 0),
            'Market_Sell_Qty_1Yr': market_sell_qty_1y,
            'Market_Sell_Value_1Yr': round(market_sell_value_1y, 0),
            'Inter_Se_Disposal_Qty_1Yr': inter_se_disposal_qty_1y,
            'Inter_Se_Pct_Of_Disposals': inter_se_pct_of_disposals,
            'Net_Market_Value_1Yr': round(net_market_value, 0),
            'Market_Sell_Pct_MCap': market_sell_pct_of_mcap,
            'Recent_Market_Buy_6m': market_buy_qty_6m > 0,
            'Unique_Market_Buyers_1Yr': len(market_buyers_1y),
            'Unique_Market_Sellers_1Yr': len(market_sellers_1y),
            'Insider_Context': " | ".join(context_parts) if context_parts else "No significant activity",
        })
        
        return result
        
    except Exception as e:
        print(f"[WARN] Could not load insider trading data from {it_file}: {e}")
        return None


def load_corporate_actions_data(financial_filepath: str, corporate_actions_dir: Optional[str] = None,
                                resolved_path: Optional[Path] = None) -> Optional[Dict[str, Any]]:
    """
    Load corporate actions data (dividends, bonus, splits, rights, board meetings) from companion JSON file.
    If resolved_path is provided (from pre-built file index), skip filesystem discovery.
    Otherwise looks for {stock_code}_dividend.json in the specified directory.
    
    Returns summary metrics useful for investment analysis.
    """
    if resolved_path is not None:
        corp_file = resolved_path
    else:
        filepath = Path(financial_filepath)
        stock_code = filepath.stem
        
        # Determine where to look for corporate actions file
        if corporate_actions_dir:
            search_dir = Path(corporate_actions_dir)
        else:
            search_dir = filepath.parent
        
        # Try to find corporate actions file (named *_dividend.json based on the sample)
        corp_file = search_dir / f"{stock_code}_dividend.json"
        
        if not corp_file.exists():
            # Try alternate patterns
            base_code = stock_code.replace('_new', '').replace('_consolidated', '').replace('_standalone', '')
            if base_code != stock_code:
                corp_file = search_dir / f"{base_code}_dividend.json"
            
            if not corp_file.exists():
                return None
    
    try:
        with open(corp_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if data.get('head', {}).get('status') != '0':
            return None
        
        body = data.get('body', {})
        events = body.get('eventsData', {})
        
        result = {}
        current_year = datetime.now().year
        
        # ========================================
        # DIVIDEND ANALYSIS
        # ========================================
        dividends = events.get('dividendTableData', [])
        result['Dividend_Count'] = len(dividends)
        
        if dividends:
            # Sort by date (most recent first)
            dividends_sorted = sorted(dividends, key=lambda x: x.get('exDate', ''), reverse=True)
            
            # Latest dividend
            latest_div = dividends_sorted[0]
            result['Latest_Dividend_Date'] = latest_div.get('exDate')
            result['Latest_Dividend_Amount'] = latest_div.get('dividendAmount')
            result['Latest_Dividend_Type'] = latest_div.get('dividendType')
            
            # Years since last dividend
            if result['Latest_Dividend_Date']:
                try:
                    last_div_year = int(result['Latest_Dividend_Date'][:4])
                    result['Years_Since_Dividend'] = current_year - last_div_year
                except (ValueError, TypeError, IndexError):
                    result['Years_Since_Dividend'] = None
            
            # Total dividends paid (sum)
            total_div = sum(d.get('dividendAmount', 0) for d in dividends if d.get('dividendAmount'))
            result['Total_Dividends_Paid'] = round(total_div, 2)
            
            # Dividend years (unique years with dividends)
            div_years = set()
            for d in dividends:
                if d.get('exDate'):
                    try:
                        div_years.add(int(d['exDate'][:4]))
                    except (ValueError, TypeError, IndexError):
                        pass
            result['Dividend_Years_Count'] = len(div_years)
            
            # Recent 5-year dividend consistency (how many of last 5 years had dividends)
            recent_5yr = [y for y in div_years if y >= current_year - 5]
            result['Dividend_5Yr_Consistency'] = len(recent_5yr)
            
            # Dividend trend (compare recent 3 vs prior 3 if available)
            if len(dividends_sorted) >= 6:
                recent_3 = sum(d.get('dividendAmount', 0) for d in dividends_sorted[:3])
                prior_3 = sum(d.get('dividendAmount', 0) for d in dividends_sorted[3:6])
                if prior_3 > 0:
                    result['Dividend_Trend'] = "INCREASING" if recent_3 > prior_3 else ("DECREASING" if recent_3 < prior_3 else "STABLE")
                else:
                    result['Dividend_Trend'] = "NEW"
            else:
                result['Dividend_Trend'] = None
        else:
            result['Latest_Dividend_Date'] = None
            result['Latest_Dividend_Amount'] = None
            result['Latest_Dividend_Type'] = None
            result['Years_Since_Dividend'] = None
            result['Total_Dividends_Paid'] = 0
            result['Dividend_Years_Count'] = 0
            result['Dividend_5Yr_Consistency'] = 0
            result['Dividend_Trend'] = "NEVER_PAID"
        
        # ========================================
        # BONUS SHARES ANALYSIS
        # ========================================
        bonus_list = events.get('bonusTableData', [])
        result['Bonus_Count'] = len(bonus_list)
        
        if bonus_list:
            bonus_sorted = sorted(bonus_list, key=lambda x: x.get('exDate', ''), reverse=True)
            latest_bonus = bonus_sorted[0]
            result['Latest_Bonus_Date'] = latest_bonus.get('exDate')
            result['Latest_Bonus_Ratio'] = latest_bonus.get('bonusRatio')
            
            # Years since last bonus
            if result['Latest_Bonus_Date']:
                try:
                    last_bonus_year = int(result['Latest_Bonus_Date'][:4])
                    result['Years_Since_Bonus'] = current_year - last_bonus_year
                except (ValueError, TypeError, IndexError):
                    result['Years_Since_Bonus'] = None
        else:
            result['Latest_Bonus_Date'] = None
            result['Latest_Bonus_Ratio'] = None
            result['Years_Since_Bonus'] = None
        
        # ========================================
        # STOCK SPLIT ANALYSIS
        # ========================================
        splits = events.get('splitTableData', [])
        result['Split_Count'] = len(splits)
        
        if splits:
            splits_sorted = sorted(splits, key=lambda x: x.get('exDate', ''), reverse=True)
            latest_split = splits_sorted[0]
            result['Latest_Split_Date'] = latest_split.get('exDate')
            result['Latest_Split_Ratio'] = latest_split.get('splitRatio', latest_split.get('ratioText'))
        else:
            result['Latest_Split_Date'] = None
            result['Latest_Split_Ratio'] = None
        
        # ========================================
        # RIGHTS ISSUE ANALYSIS
        # ========================================
        rights = events.get('rightTableData', [])
        result['Rights_Count'] = len(rights)
        
        if rights:
            rights_sorted = sorted(rights, key=lambda x: x.get('xrDate', ''), reverse=True)
            latest_rights = rights_sorted[0]
            result['Latest_Rights_Date'] = latest_rights.get('xrDate')
            result['Latest_Rights_Ratio'] = latest_rights.get('ratioText')
            result['Latest_Rights_Premium'] = latest_rights.get('rightsPremium')
            
            # Recent rights issue (within 2 years) could indicate capital need
            if result['Latest_Rights_Date']:
                try:
                    rights_year = int(result['Latest_Rights_Date'][:4])
                    result['Years_Since_Rights'] = current_year - rights_year
                except (ValueError, TypeError, IndexError):
                    result['Years_Since_Rights'] = None
        else:
            result['Latest_Rights_Date'] = None
            result['Latest_Rights_Ratio'] = None
            result['Latest_Rights_Premium'] = None
            result['Years_Since_Rights'] = None
        
        # ========================================
        # BOARD MEETINGS
        # ========================================
        meetings = events.get('boardMeetingTableData', [])
        result['Board_Meetings_Count'] = len(meetings)
        
        # Find next upcoming meeting
        today = datetime.now().strftime('%Y-%m-%d')
        upcoming = [m for m in meetings if m.get('boardMeetDate', '') >= today]
        if upcoming:
            upcoming_sorted = sorted(upcoming, key=lambda x: x.get('boardMeetDate', ''))
            next_meeting = upcoming_sorted[0]
            result['Next_Board_Meeting'] = next_meeting.get('boardMeetDate')
            result['Next_Meeting_Purpose'] = next_meeting.get('purpose')
        else:
            result['Next_Board_Meeting'] = None
            result['Next_Meeting_Purpose'] = None
        
        # ========================================
        # RESEARCH COVERAGE (Analyst Reports)
        # ========================================
        research = body.get('researchReports', {}).get('tableData', [])
        result['Analyst_Reports_Count'] = len(research)
        
        if research:
            # Get latest recommendation
            research_sorted = sorted(research, key=lambda x: x.get('recoDate', ''), reverse=True)
            latest_reco = research_sorted[0]
            result['Latest_Target_Price'] = latest_reco.get('targetPrice')
            result['Latest_Reco_Type'] = latest_reco.get('recoType')
            result['Latest_Upside'] = latest_reco.get('upside')
            result['Latest_Reco_Date'] = latest_reco.get('recoDate')
        else:
            result['Latest_Target_Price'] = None
            result['Latest_Reco_Type'] = None
            result['Latest_Upside'] = None
            result['Latest_Reco_Date'] = None
        
        # ========================================
        # COMPANY PROFILE SUMMARY
        # ========================================
        profile = body.get('companyProfileData', {})
        
        executives = profile.get('topExecutivesTableData', [])
        result['Executive_Count'] = len(executives)
        
        directors = profile.get('directorsTableData', [])
        result['Director_Count'] = len(directors)
        
        result['Website'] = profile.get('websiteUrl')
        
        # ========================================
        # CAPITAL ALLOCATION ASSESSMENT
        # ========================================
        # Shareholder-friendly score based on dividends, bonus, and dilutive actions
        shareholder_score = 0
        allocation_notes = []
        
        # Consistent dividend payer gets points
        if result['Dividend_5Yr_Consistency'] >= 4:
            shareholder_score += 2
            allocation_notes.append("Consistent dividend payer")
        elif result['Dividend_5Yr_Consistency'] >= 2:
            shareholder_score += 1
            allocation_notes.append("Occasional dividends")
        
        # Recent bonus is shareholder-friendly
        if result['Bonus_Count'] > 0 and result.get('Years_Since_Bonus') is not None and result['Years_Since_Bonus'] <= 5:
            shareholder_score += 1
            allocation_notes.append("Recent bonus")
        
        # Recent rights issue could indicate capital need (neutral to negative)
        if result['Rights_Count'] > 0 and result.get('Years_Since_Rights') is not None and result['Years_Since_Rights'] <= 2:
            allocation_notes.append("Recent rights issue (capital raising)")
        
        if result['Dividend_Trend'] == "INCREASING":
            shareholder_score += 1
            allocation_notes.append("Growing dividends")
        elif result['Dividend_Trend'] == "NEVER_PAID":
            allocation_notes.append("Never paid dividends")
        
        result['Capital_Allocation_Score'] = shareholder_score
        result['Capital_Allocation_Notes'] = "; ".join(allocation_notes) if allocation_notes else "No notable actions"
        
        return result
        
    except Exception as e:
        print(f"[WARN] Could not load corporate actions data from {corp_file}: {e}")
        return None


def load_overview_data(financial_filepath: str, overview_dir: Optional[str] = None,
                       resolved_path: Optional[Path] = None) -> Optional[Dict[str, Any]]:
    """
    Load overview data (TTM metrics, Piotroski score, etc.) from companion JSON file.
    If resolved_path is provided (from pre-built file index), skip filesystem discovery.
    Otherwise looks for {stock_code}_overview.json in the specified directory.
    
    Returns TTM metrics which are more current than annual data.
    """
    if resolved_path is not None:
        overview_file = resolved_path
    else:
        filepath = Path(financial_filepath)
        stock_code = filepath.stem
        
        # Determine where to look for overview file
        if overview_dir:
            search_dir = Path(overview_dir)
        else:
            search_dir = filepath.parent
        
        # Try to find overview file
        overview_file = search_dir / f"{stock_code}_overview.json"
        
        if not overview_file.exists():
            # Try alternate patterns
            base_code = stock_code.replace('_new', '').replace('_consolidated', '').replace('_standalone', '')
            if base_code != stock_code:
                overview_file = search_dir / f"{base_code}_overview.json"
            
            if not overview_file.exists():
                return None
    
    try:
        with open(overview_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if data.get('head', {}).get('status') != '0':
            return None
        
        body = data.get('body', {})
        result = {}
        
        # Helper to safely extract value from nested structure
        def get_metric(key):
            metric = body.get(key, {})
            if isinstance(metric, dict):
                return metric.get('value')
            return None
        
        def get_metric_color(key):
            """Get the color/sentiment indicator"""
            metric = body.get(key, {})
            if isinstance(metric, dict):
                return metric.get('color1')
            return None
        
        def get_metric_text(key):
            """Get the short text description"""
            metric = body.get(key, {})
            if isinstance(metric, dict):
                return metric.get('st1')
            return None
        
        # ========================================
        # VALUATION METRICS (TTM - more current)
        # ========================================
        result['PE_TTM'] = get_metric('PE_TTM')
        result['PEG_TTM'] = get_metric('PEG_TTM')
        result['PBV_Overview'] = get_metric('PBV_A')
        result['Market_Cap_Overview'] = get_metric('MCAP_Q')
        
        # ========================================
        # PIOTROSKI SCORE (Key quality metric)
        # ========================================
        result['Piotroski_Score'] = get_metric('PITROSKI_F')
        result['Piotroski_Assessment'] = get_metric_text('PITROSKI_F')
        
        # ========================================
        # TTM GROWTH METRICS
        # ========================================
        result['Revenue_Growth_TTM'] = get_metric('SR_TTM_GROWTH')
        result['Revenue_Growth_Qtr_YoY'] = get_metric('REV4Q_Q')
        result['NP_Growth_TTM'] = get_metric('NP_TTM_GROWTH')
        result['NP_Growth_Qtr_YoY'] = get_metric('NP_Q_GROWTH')
        
        # ========================================
        # TTM PROFITABILITY
        # ========================================
        result['OPM_TTM'] = get_metric('OPMPCT_TTM')
        result['OPM_Qtr'] = get_metric('OPMPCT_Q')
        result['ROE_Overview'] = get_metric('ROE_A')
        result['ROA_Overview'] = get_metric('ROA_A')
        
        # ========================================
        # INSTITUTIONAL HOLDING (from overview)
        # ========================================
        result['Institutional_Holding_Overview'] = get_metric('INSTIHOLD')
        result['Institutional_Holding_Change'] = get_metric_text('INSTIHOLD')
        
        # ========================================
        # RELATIVE PERFORMANCE
        # ========================================
        result['Returns_vs_Nifty50_Qtr'] = get_metric('relp_nifty50_qtrChangeP')
        result['Returns_vs_Sector_Qtr'] = get_metric('relp_sector_qtrChangeP')
        
        # ========================================
        # 52 WEEK HIGH/LOW (from prepend_params)
        # ========================================
        prepend_params = data.get('prepend_params', [])
        for param in prepend_params:
            param_name = param.get('param_name', '')
            unique_name = param.get('unique_name', '')
            
            if param_name == '52_Week_High_Low':
                result['Week52_High'] = param.get('high')
                result['Week52_Low'] = param.get('low')
                result['LTP'] = param.get('ltp')
                result['Return_1Yr'] = param.get('change_val')
                
                # Calculate distance from 52-week high/low
                if result['LTP'] and result['Week52_High'] and result['Week52_Low']:
                    high = result['Week52_High']
                    low = result['Week52_Low']
                    ltp = result['LTP']
                    if high > low:
                        # Where is current price in the 52-week range (0% = at low, 100% = at high)
                        result['Price_Position_52W'] = round(((ltp - low) / (high - low)) * 100, 1)
                        # Distance from 52-week high (negative = below high)
                        result['Distance_From_52W_High'] = round(((ltp - high) / high) * 100, 1)
            
            if unique_name == '1yr_vs_nifty50':
                # Value is like "-48.1%" - extract numeric part
                value = param.get('value', '')
                if isinstance(value, str):
                    try:
                        result['Return_vs_Nifty_1Yr'] = float(value.replace('%', ''))
                    except (ValueError, TypeError, IndexError):
                        result['Return_vs_Nifty_1Yr'] = None
                else:
                    result['Return_vs_Nifty_1Yr'] = value
        
        return result
        
    except Exception as e:
        print(f"[WARN] Could not load overview data from {overview_file}: {e}")
        return None


def analyze_stock(filepath: str, shareholding_dir: Optional[str] = None, corporate_actions_dir: Optional[str] = None, overview_dir: Optional[str] = None, insider_trading_dir: Optional[str] = None, file_index: Optional[Dict] = None) -> Optional[Dict[str, Any]]:
    """
    Analyze a single stock JSON file.
    Returns a dict of metrics, trends, and red flags.
    
    Args:
        filepath: Path to the financial JSON file
        shareholding_dir: Optional directory to look for shareholding JSONs
                         (if None, looks in same directory as financial JSON)
        corporate_actions_dir: Optional directory to look for corporate actions JSONs
                              (dividends, bonus, splits, rights issues)
        overview_dir: Optional directory to look for overview JSONs
                     (TTM metrics, Piotroski score, 52-week high/low)
        insider_trading_dir: Optional directory to look for insider trading JSONs
                            (promoter buys/sells, inter-se transfers, SAST data)
        file_index: Optional pre-built dict mapping stock_code -> {type: Path}
                   from build_companion_file_index(). Eliminates per-stock filesystem lookups.
    """
    nse_code = Path(filepath).stem
    
    # Resolve companion file paths from pre-built index (avoids filesystem stat calls)
    _stock_files = file_index.get(nse_code, {}) if file_index else {}
    _resolved_sh = _stock_files.get('shareholding')
    _resolved_ca = _stock_files.get('dividend')
    _resolved_ov = _stock_files.get('overview')
    _resolved_it = _stock_files.get('insider_trading')
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"[ERROR] Failed to parse {filepath}: {e}")
        return None
    
    # Check for successful response
    if data.get('head', {}).get('status') != '0':
        print(f"[WARN] {nse_code}: Non-success status in JSON")
        return None
    
    body = data.get('body', {})
    annual_dump = body.get('annualDataDump', {})
    #print(json.dumps(annual_dump,indent=2)); sys.exit()
    
    # 1. Try to get consolidated data first
    yearly_data = annual_dump.get('consolidated', {})
    available_years = extract_available_years(yearly_data)
    data_type = 'consolidated'
    consolidated_rejected = False # Initialize the flag

    # 2. Fallback to standalone if consolidated is missing OR has insufficient years
    if len(available_years) < 2:
        if yearly_data: # If it existed but was too sparse
            consolidated_rejected = True
        yearly_data = annual_dump.get('standalone', {})
        available_years = extract_available_years(yearly_data)
        data_type = 'standalone'

    # 3. Final check: if even standalone is insufficient, then we bail
    if len(available_years) < 2:
        print(f"[WARN] {nse_code}: Insufficient data in both consolidated and standalone")
        return None
    
    # Prefer consolidated, fallback to standalone
    '''
    if annual_dump.get('consolidated'):
        yearly_data = annual_dump['consolidated']
        data_type = 'consolidated'
    else:
        yearly_data = annual_dump.get('standalone', {})
        data_type = 'standalone'
    
    if not yearly_data:
        print(f"[WARN] {nse_code}: No annual data found")
        return None
    
    # Get available years
    available_years = extract_available_years(yearly_data)
    if len(available_years) < 2:
        print(f"[WARN] {nse_code}: Insufficient year data ({len(available_years)} years)")
        return None
    '''
    latest_year = available_years[0]
    latest = yearly_data[latest_year]
    
    # ========================================
    # 0. IDENTITY & CONTEXT (from stockData)
    # ========================================
    stock_data = body.get('stockData', {})
    
    result = {
        'NSE_Code': nse_code,
        'Stock_Name': stock_data.get('stockName', ''),
        'Sector': stock_data.get('sectorName', ''),
        'Industry': stock_data.get('industryName', ''),
        'BSE_Code': stock_data.get('BSEcode', ''),
        'ISIN': stock_data.get('ISIN', ''),
        'Data_Type': data_type,
        'Years_Available': len(available_years),
        'Latest_Year': latest_year,
        'Fiscal_Year_End': latest_year.split()[0] if latest_year else 'Unknown',  # "Mar", "Jun", "Dec" etc.
    }
    
    # Detect true fiscal year end from the most common month in available years
    # (handles companies that recently switched FY, e.g., Jun→Mar)
    fy_months = [y.split()[0] for y in available_years if ' ' in y]
    if fy_months:
        from collections import Counter
        fy_mode = Counter(fy_months).most_common(1)[0][0]
        result['Fiscal_Year_End'] = fy_mode
    
    # ========================================
    # 0b. LOAD OVERVIEW DATA EARLY (TTM metrics are more current)
    # ========================================
    # Load overview data first so we can use TTM values as primary where available
    overview = load_overview_data(filepath, overview_dir, resolved_path=_resolved_ov)
    overview_available = overview is not None
    
    # Pre-extract key overview values for use in calculations
    pe_ttm = overview.get('PE_TTM') if overview else None
    peg_ttm = overview.get('PEG_TTM') if overview else None
    market_cap_overview = overview.get('Market_Cap_Overview') if overview else None
    opm_ttm = overview.get('OPM_TTM') if overview else None
    piotroski_score = overview.get('Piotroski_Score') if overview else None
    
    # ========================================
    # 1. VALUATION METRICS (Latest Year, with TTM overrides)
    # ========================================
    # Use TTM values as PRIMARY when available (more current than annual)
    result['PE'] = pe_ttm if pe_ttm is not None else safe_float(safe_get(latest, 'PE_A'))
    result['PE_Annual'] = safe_float(safe_get(latest, 'PE_A'))  # Keep annual for reference
    result['PE_TTM'] = pe_ttm
    
    result['PBV'] = safe_float(safe_get(latest, 'PBV_A'))
    result['EV_EBITDA'] = safe_float(safe_get(latest, 'EVPerEBITDA_A'))
    result['Price_To_Sales'] = safe_float(safe_get(latest, 'PriceToSales_A'))
    result['Earnings_Yield'] = safe_float(safe_get(latest, 'EYield_A'))
    result['Market_Cap_Sales'] = safe_float(safe_get(latest, 'MarketCapPerSales_A'))
    
    # ========================================
    # 1b. NEW EXTRACTIONS (Gap Analysis Items)
    # ========================================
    # EBITDA and EBITDA Margin
    result['EBITDA'] = safe_float(safe_get(latest, 'EBIDT_A'))
    result['EBITDA_Margin'] = safe_float(safe_get(latest, 'EBIDTPCT_A'))
    
    # Depreciation (for Capex/Depreciation ratio analysis)
    result['Depreciation'] = safe_float(safe_get(latest, 'DEP_A'))
    
    # Cash EPS (earnings quality metric)
    result['CEPS'] = safe_float(safe_get(latest, 'CEPS_A'))
    
    # Cash from Investing (proxy for capex direction - negative = investing/capex)
    result['CFI'] = safe_float(safe_get(latest, 'CFI_A'))
    
    # Trade Payables (for payables days calculation)
    result['Trade_Payables'] = safe_float(safe_get(latest, 'TradePayables_A'))
    
    # Retention Ratio (already computed by Trendlyne)
    result['Retention_Ratio'] = safe_float(safe_get(latest, 'RetentionRatios_A'))
    
    # PEG Ratio - compute after we have PE and EPS growth
    # (Will be computed later after EPS_Growth_3Yr is available)
    
    # ========================================
    # 2. PROFITABILITY METRICS (Multi-Year)
    # ========================================
    roe_series = [safe_float(safe_get(yearly_data.get(y, {}), 'ROE_A')) for y in available_years]
    roce_series = [safe_float(safe_get(yearly_data.get(y, {}), 'ROCE_A')) for y in available_years]
    opm_series = [safe_float(safe_get(yearly_data.get(y, {}), 'OPMPCT_A')) for y in available_years]
    npm_series = [safe_float(safe_get(yearly_data.get(y, {}), 'NETPCT_A')) for y in available_years]
    roa_series = [safe_float(safe_get(yearly_data.get(y, {}), 'ROA_A')) for y in available_years]
    
    result['ROE_Latest'] = roe_series[0] if roe_series else None
    result['ROCE_Latest'] = roce_series[0] if roce_series else None
    result['OPM_Latest'] = opm_series[0] if opm_series else None
    result['NPM_Latest'] = npm_series[0] if npm_series else None
    result['ROA_Latest'] = roa_series[0] if roa_series else None
    
    # ROE - ROA Gap: High gap means returns driven by leverage, not operations
    if result['ROE_Latest'] is not None and result['ROA_Latest'] is not None:
        result['ROE_ROA_Gap'] = result['ROE_Latest'] - result['ROA_Latest']
        # Flag if leverage is driving returns (gap > 15 is suspicious)
        result['Leverage_Driven'] = "Yes" if result['ROE_ROA_Gap'] > 15 else "No"
    else:
        result['ROE_ROA_Gap'] = None
        result['Leverage_Driven'] = "Unknown"
    
    # Averages - 3 Year
    result['ROE_3Yr_Avg'] = mean([v for v in roe_series[:3] if v is not None]) if len([v for v in roe_series[:3] if v is not None]) >= 2 else None
    result['ROCE_3Yr_Avg'] = mean([v for v in roce_series[:3] if v is not None]) if len([v for v in roce_series[:3] if v is not None]) >= 2 else None
    
    # Averages - 5 Year (NEW: from Gap Analysis)
    result['ROE_5Yr_Avg'] = mean([v for v in roe_series[:5] if v is not None]) if len([v for v in roe_series[:5] if v is not None]) >= 3 else None
    result['ROCE_5Yr_Avg'] = mean([v for v in roce_series[:5] if v is not None]) if len([v for v in roce_series[:5] if v is not None]) >= 3 else None
    
    # Trends (positive = improving)
    result['ROE_Trend'] = compute_trend(roe_series)
    result['ROCE_Trend'] = compute_trend(roce_series)
    result['OPM_Trend'] = compute_trend(opm_series)
    result['NPM_Trend'] = compute_trend(npm_series)
    
    # ========================================
    # 3. CASH FLOW QUALITY
    # ========================================
    cfo_series = [safe_float(safe_get(yearly_data.get(y, {}), 'CFO_A')) for y in available_years]
    pat_series = [safe_float(safe_get(yearly_data.get(y, {}), 'PAT_A')) for y in available_years]
    np_series = [safe_float(safe_get(yearly_data.get(y, {}), 'NP_A')) for y in available_years]
    
    result['CFO_Latest'] = cfo_series[0] if cfo_series else None
    result['PAT_Latest'] = pat_series[0] if pat_series else None
    result['NP_Latest'] = np_series[0] if np_series else None
    
    # CFO to PAT ratio - Latest year only
    if result['CFO_Latest'] is not None and result['PAT_Latest'] is not None and result['PAT_Latest'] > 0:
        result['CFO_PAT_Latest'] = round(result['CFO_Latest'] / result['PAT_Latest'], 2)
    else:
        result['CFO_PAT_Latest'] = None
    
    # CFO/PAT 3-Year: Use SUM method, not average of ratios
    # CRITIQUE FIX: Averaging ratios is mathematically flawed
    # Example: [0.5, 0.5, 200] averages to 67, which is misleading
    # Sum(CFO) / Sum(PAT) is correct: if CFO = [1, 1, 2] and PAT = [2, 2, 0.01]
    # Sum method: 4 / 4.01 = 0.99 (accurate), Average method: 67 (wrong)
    cfo_3yr = [c for c in cfo_series[:3] if c is not None]
    pat_3yr = [p for p in pat_series[:3] if p is not None]
    
    if len(cfo_3yr) >= 2 and len(pat_3yr) >= 2:
        sum_cfo = sum(cfo_3yr)
        sum_pat = sum(pat_3yr)
        if sum_pat > 0:  # Only if total PAT is positive
            result['CFO_PAT_3Yr_Avg'] = round(sum_cfo / sum_pat, 2)
        else:
            result['CFO_PAT_3Yr_Avg'] = None
    else:
        result['CFO_PAT_3Yr_Avg'] = None
    
    # Count years with positive CFO
    positive_cfo_years = sum(1 for c in cfo_series if c is not None and c > 0)
    result['Positive_CFO_Years'] = positive_cfo_years
    
    # Accruals (lower is better - high accruals = low quality earnings)
    result['Accruals'] = safe_float(safe_get(latest, 'ACCRUALS_A'))
    result['CFROA'] = safe_float(safe_get(latest, 'CFROA_A'))
    
    # CFO trend
    result['CFO_Trend'] = compute_trend(cfo_series)
    
    # ========================================
    # 4. WORKING CAPITAL EFFICIENCY
    # ========================================
    result['Working_Capital'] = safe_float(safe_get(latest, 'WC_A'))
    result['WC_Turnover'] = safe_float(safe_get(latest, 'WCTO_A'))
    result['Asset_Turnover'] = safe_float(safe_get(latest, 'ASETTO_A'))
    result['Current_Ratio'] = safe_float(safe_get(latest, 'CRATIO_A'))
    result['Quick_Ratio'] = safe_float(safe_get(latest, 'QuickRatio_A'))
    
    # WC Growth vs Revenue Growth (NEW - catches channel stuffing)
    # If WC grows much faster than Revenue, company is "buying" sales
    wc_series = [safe_float(safe_get(yearly_data.get(y, {}), 'WC_A')) for y in available_years]
    rev_series_wc = [safe_float(safe_get(yearly_data.get(y, {}), 'SR_A')) for y in available_years]
    
    valid_wc = [(i, w) for i, w in enumerate(wc_series) if w is not None and w > 0]
    valid_rev_wc = [(i, r) for i, r in enumerate(rev_series_wc) if r is not None and r > 0]
    
    if len(valid_wc) >= 2 and len(valid_rev_wc) >= 2:
        # Calculate growth from oldest to newest
        wc_newest, wc_oldest = valid_wc[0][1], valid_wc[-1][1]
        rev_newest, rev_oldest = valid_rev_wc[0][1], valid_rev_wc[-1][1]
        
        if wc_oldest > 0 and rev_oldest > 0:
            wc_growth = ((wc_newest / wc_oldest) - 1) * 100
            rev_growth = ((rev_newest / rev_oldest) - 1) * 100
            result['WC_Growth_Pct'] = round(wc_growth, 1)
            result['Rev_Growth_Pct'] = round(rev_growth, 1)
            
            # Ratio: How much faster is WC growing vs Revenue?
            if rev_growth > 0:
                result['WC_Rev_Growth_Ratio'] = round(wc_growth / rev_growth, 2)
            elif wc_growth > 0:
                result['WC_Rev_Growth_Ratio'] = 99.0  # WC growing, Rev flat/negative
            else:
                result['WC_Rev_Growth_Ratio'] = 0.0
        else:
            result['WC_Growth_Pct'] = None
            result['Rev_Growth_Pct'] = None
            result['WC_Rev_Growth_Ratio'] = None
    else:
        result['WC_Growth_Pct'] = None
        result['Rev_Growth_Pct'] = None
        result['WC_Rev_Growth_Ratio'] = None
    
    # Receivable Days = (TradeReceivables / Revenue) * 365
    trade_recv = safe_float(safe_get(latest, 'TradeReceivables_A'))
    revenue = safe_float(safe_get(latest, 'SR_A'))
    inventory = safe_float(safe_get(latest, 'Inventories_A'))
    
    if trade_recv is not None and revenue is not None and revenue > 0:
        result['Receivable_Days'] = (trade_recv / revenue) * 365
    else:
        result['Receivable_Days'] = None
    
    # Inventory Days = (Inventory / Revenue) * 365
    if inventory is not None and revenue is not None and revenue > 0:
        result['Inventory_Days'] = (inventory / revenue) * 365
    else:
        result['Inventory_Days'] = None
    
    # Payables Days = (Trade Payables / Revenue) * 365 (NEW: from Gap Analysis)
    # Higher payables days can indicate strong supplier bargaining power OR payment stress
    trade_payables = safe_float(safe_get(latest, 'TradePayables_A'))
    if trade_payables is not None and revenue is not None and revenue > 0:
        result['Payables_Days'] = (trade_payables / revenue) * 365
    else:
        result['Payables_Days'] = None
    
    # Cash Conversion Cycle = Receivable Days + Inventory Days - Payables Days
    # Lower (or negative) is better - means company gets paid before it has to pay suppliers
    if result['Receivable_Days'] is not None and result['Inventory_Days'] is not None and result['Payables_Days'] is not None:
        result['Cash_Conversion_Cycle'] = round(result['Receivable_Days'] + result['Inventory_Days'] - result['Payables_Days'], 1)
    else:
        result['Cash_Conversion_Cycle'] = None
    
    # Receivable days trend (increasing = bad)
    recv_days_series = []
    for y in available_years:
        yr_data = yearly_data.get(y, {})
        tr = safe_float(safe_get(yr_data, 'TradeReceivables_A'))
        sr = safe_float(safe_get(yr_data, 'SR_A'))
        if tr is not None and sr is not None and sr > 0:
            recv_days_series.append((tr / sr) * 365)
        else:
            recv_days_series.append(None)
    
    result['Recv_Days_Trend'] = compute_trend(recv_days_series)
    
    # Inventory days trend (increasing = equally dangerous as receivables)
    # CRITIQUE FIX: Decreasing receivables but spiking inventory is just as bad
    inv_days_series = []
    for y in available_years:
        yr_data = yearly_data.get(y, {})
        inv = safe_float(safe_get(yr_data, 'Inventories_A'))
        sr = safe_float(safe_get(yr_data, 'SR_A'))
        if inv is not None and sr is not None and sr > 0:
            inv_days_series.append((inv / sr) * 365)
        else:
            inv_days_series.append(None)
    
    result['Inv_Days_Trend'] = compute_trend(inv_days_series)
    
    # ========================================
    # 5. EARNINGS QUALITY FLAGS
    # ========================================
    # Exceptional items - should be 0 or minimal
    exceptional_items = [safe_float(safe_get(yearly_data.get(y, {}), 'ExceptionalItems_A'), 0) for y in available_years]
    non_zero_exceptional = sum(1 for e in exceptional_items if e != 0)
    result['Exceptional_Items_Count'] = non_zero_exceptional
    result['Exceptional_Items_Latest'] = exceptional_items[0] if exceptional_items else 0
    
    # Other income as % of PAT (high = suspect)
    # ENHANCED: Check multiple years, not just latest
    other_income = safe_float(safe_get(latest, 'OI_A'), 0)
    pat = safe_float(safe_get(latest, 'PAT_A'))
    if other_income and pat and pat > 0:
        result['Other_Income_Pct_PAT'] = (other_income / pat) * 100
    else:
        result['Other_Income_Pct_PAT'] = None
    
    # Count years where Other Income > 30% of PAT
    # If company relies on Other Income for 2+ consecutive years, it's not a real business
    high_other_income_years = 0
    for y in available_years[:3]:  # Check last 3 years
        yr_data = yearly_data.get(y, {})
        oi = safe_float(safe_get(yr_data, 'OI_A'), 0)
        yr_pat = safe_float(safe_get(yr_data, 'PAT_A'))
        if oi and yr_pat and yr_pat > 0:
            if (oi / yr_pat) * 100 > 30:
                high_other_income_years += 1
    result['High_Other_Income_Years'] = high_other_income_years
    
    # Deferred Tax Assets growing faster than profits (potential flag)
    result['Deferred_Tax_Assets'] = safe_float(safe_get(latest, 'DeferredTaxAssets_A'))
    
    # ========================================
    # 6. CAPITAL ALLOCATION
    # ========================================
    result['Dividend_Per_Share'] = safe_float(safe_get(latest, 'DividendPerShare_A'))
    result['Dividend_Payout_NP'] = safe_float(safe_get(latest, 'DividendPayoutNP_A'))
    result['ROIC'] = safe_float(safe_get(latest, 'ROIC_A'))
    
    # Dividend consistency
    div_series = [safe_float(safe_get(yearly_data.get(y, {}), 'DividendPerShare_A'), 0) for y in available_years]
    dividend_years = sum(1 for d in div_series if d > 0)
    result['Dividend_Paying_Years'] = dividend_years
    
    # ========================================
    # 7. DEBT & COVERAGE
    # ========================================
    result['Debt_Equity'] = safe_float(safe_get(latest, 'DEBT_CE_A'))
    result['LT_Debt_Equity'] = safe_float(safe_get(latest, 'LTDE_A'))
    result['Interest_Coverage'] = safe_float(safe_get(latest, 'IC_A'))
    
    # Debt trend
    debt_series = [safe_float(safe_get(yearly_data.get(y, {}), 'DEBT_CE_A')) for y in available_years]
    result['Debt_Trend'] = compute_trend(debt_series)
    
    # Net Debt / EBITDA (NEW: from Gap Analysis)
    # Net Debt = Total Debt - Cash (approximation using Debt/Equity * Equity - Cash)
    # Lower is better - < 2 is comfortable, > 4 is concerning
    cash = safe_float(safe_get(latest, 'CashAndCashEquivalents_A'))
    total_debt = safe_float(safe_get(latest, 'Borrowings_A'))  # Total borrowings
    if total_debt is None:
        # Fallback: estimate from D/E ratio and shareholders funds
        shareholders_funds = safe_float(safe_get(latest, 'TotalShareHoldersFunds_A'))
        if result['Debt_Equity'] is not None and shareholders_funds is not None:
            total_debt = result['Debt_Equity'] * shareholders_funds
    
    if total_debt is not None and result['EBITDA'] is not None and result['EBITDA'] > 0:
        net_debt = total_debt - (cash if cash else 0)
        result['Net_Debt_EBITDA'] = round(net_debt / result['EBITDA'], 2)
    else:
        result['Net_Debt_EBITDA'] = None
    
    # ========================================
    # 8. GROWTH METRICS (from JSON)
    # ========================================
    # REV1_A = 1Y, REV2_A = 2Y, REV3_A = 3Y (Trendlyne naming convention)
    result['Revenue_Growth_1Yr'] = safe_float(safe_get(latest, 'REV1_A'))
    result['Revenue_Growth_2Yr'] = safe_float(safe_get(latest, 'REV2_A'))
    result['Revenue_Growth_3Yr'] = safe_float(safe_get(latest, 'REV3_A'))
    result['NP_Growth'] = safe_float(safe_get(latest, 'NP_A_GROWTH'))
    result['EPS_Growth_3Yr'] = safe_float(safe_get(latest, 'CEPS3_A'))
    
    # PEG Ratio: Use PEG_TTM from overview as PRIMARY (more current)
    # Fallback to computed PEG = PE / EPS Growth Rate
    peg_computed = None
    if result['PE'] is not None and result['PE'] > 0 and result['EPS_Growth_3Yr'] is not None and result['EPS_Growth_3Yr'] > 0:
        peg_computed = round(result['PE'] / result['EPS_Growth_3Yr'], 2)
    
    # Prefer PEG_TTM from overview
    if peg_ttm is not None:
        result['PEG'] = peg_ttm
    elif peg_computed is not None:
        result['PEG'] = peg_computed
    else:
        result['PEG'] = None
    
    result['PEG_TTM'] = peg_ttm  # Store TTM value separately
    result['PEG_Computed'] = peg_computed  # Keep computed for reference
    
    # Revenue CAGR (calculated)
    rev_series = [safe_float(safe_get(yearly_data.get(y, {}), 'SR_A')) for y in available_years]
    if len(rev_series) >= 4:
        result['Revenue_CAGR_3Yr'] = compute_cagr(rev_series[3], rev_series[0], 3)
    else:
        result['Revenue_CAGR_3Yr'] = None
    
    # ========================================
    # 8b. GROWTH DURABILITY METRICS
    # ========================================
    # Revenue Volatility: Coefficient of Variation (std dev / mean)
    # Lower is better - indicates stable, predictable growth
    valid_revenues = [r for r in rev_series if r is not None and r > 0]
    if len(valid_revenues) >= 3:
        rev_mean = mean(valid_revenues)
        rev_std = (sum((r - rev_mean) ** 2 for r in valid_revenues) / len(valid_revenues)) ** 0.5
        result['Revenue_Volatility'] = (rev_std / rev_mean) * 100 if rev_mean > 0 else None
    else:
        result['Revenue_Volatility'] = None
    
    # Profit Consistency: % of years with positive NP
    np_series_full = [safe_float(safe_get(yearly_data.get(y, {}), 'NP_A')) for y in available_years]
    valid_np = [n for n in np_series_full if n is not None]
    if valid_np:
        positive_np_count = sum(1 for n in valid_np if n > 0)
        result['Profit_Consistency'] = (positive_np_count / len(valid_np)) * 100
    else:
        result['Profit_Consistency'] = None
    
    # Profit Growth Consistency: % of years with positive NP growth
    np_growth_series = [safe_float(safe_get(yearly_data.get(y, {}), 'NP_A_GROWTH')) for y in available_years]
    valid_np_growth = [g for g in np_growth_series if g is not None]
    if valid_np_growth:
        positive_growth_count = sum(1 for g in valid_np_growth if g > 0)
        result['Profit_Growth_Consistency'] = (positive_growth_count / len(valid_np_growth)) * 100
    else:
        result['Profit_Growth_Consistency'] = None
    
    # Share Dilution Detection
    # CRITIQUE FIX: EQCAP change includes bonus shares (not real dilution)
    # Better approach: Compare EPS growth vs NP growth
    # If NP grows 50% but EPS only grows 30%, there's 20% dilution
    eps_series = [safe_float(safe_get(yearly_data.get(y, {}), 'EPS_A')) for y in available_years]
    
    # Method 1: EPS vs NP growth comparison (more accurate)
    valid_eps = [(i, e) for i, e in enumerate(eps_series) if e is not None and e > 0]
    valid_np = [(i, n) for i, n in enumerate(np_series_full) if n is not None and n > 0]
    
    if len(valid_eps) >= 2 and len(valid_np) >= 2:
        # Calculate growth rates
        newest_eps, oldest_eps = valid_eps[0][1], valid_eps[-1][1]
        newest_np, oldest_np = valid_np[0][1], valid_np[-1][1]
        
        if oldest_eps > 0 and oldest_np > 0:
            eps_growth = (newest_eps / oldest_eps - 1) * 100
            np_growth = (newest_np / oldest_np - 1) * 100
            
            # Dilution = NP growth - EPS growth (if positive, shares were issued)
            result['Share_Dilution_Pct'] = round(np_growth - eps_growth, 1)
        else:
            result['Share_Dilution_Pct'] = None
    else:
        # Fallback to EQCAP method if EPS/NP data insufficient
        eqcap_series = [safe_float(safe_get(yearly_data.get(y, {}), 'EQCAP_A')) for y in available_years]
        valid_eqcap = [(i, e) for i, e in enumerate(eqcap_series) if e is not None and e > 0]
        if len(valid_eqcap) >= 2:
            newest_eqcap, oldest_eqcap = valid_eqcap[0][1], valid_eqcap[-1][1]
            if oldest_eqcap > 0:
                result['Share_Dilution_Pct'] = round(((newest_eqcap / oldest_eqcap) - 1) * 100, 1)
            else:
                result['Share_Dilution_Pct'] = 0
        else:
            result['Share_Dilution_Pct'] = None
    
    # ========================================
    # 9. SIZE METRICS
    # ========================================
    result['Total_Revenue'] = revenue
    result['Total_Assets'] = safe_float(safe_get(latest, 'TA_A'))
    result['Total_Equity'] = safe_float(safe_get(latest, 'TotalShareHoldersFunds_A'))
    result['Equity_Capital'] = safe_float(safe_get(latest, 'EQCAP_A'))
    result['Book_Value_Per_Share'] = safe_float(safe_get(latest, 'BVSH_A'))
    result['EPS'] = safe_float(safe_get(latest, 'EPS_A'))
    
    # Market Cap: Use overview MCAP_Q (current) as PRIMARY, fallback to computed from P/S ratio
    price_to_sales = safe_float(safe_get(latest, 'PriceToSales_A'))
    market_cap_computed = price_to_sales * revenue if (price_to_sales is not None and revenue is not None) else None
    
    # Prefer overview Market Cap (more current) over computed annual
    if market_cap_overview is not None:
        result['Market_Cap'] = market_cap_overview
    elif market_cap_computed is not None:
        result['Market_Cap'] = market_cap_computed
    else:
        result['Market_Cap'] = None
    
    result['Market_Cap_Annual'] = market_cap_computed  # Keep computed for reference
    
    # Enterprise Value (directly from JSON)
    result['Enterprise_Value'] = safe_float(safe_get(latest, 'EnterpriseValue_A'))
    
    # ========================================
    # 10. RED FLAGS - Split into Structural vs Pricing
    # ========================================
    quality_flags = []
    pricing_flags = []
    
    # --- QUALITY FLAGS (Business Quality Issues) ---
    
    # Negative or declining ROE
    if result['ROE_Latest'] is not None and result['ROE_Latest'] < 10:
        quality_flags.append("LOW_ROE")
    # Declining ROE - but only flag if ending ROE is also mediocre
    # CRITIQUE FIX: A company dropping from 60% to 30% ROE is still excellent
    # Only flag if: slope < -2 AND Latest_ROE < 15%
    if result['ROE_Trend'] is not None and result['ROE_Trend'] < -2:
        if result['ROE_Latest'] is not None and result['ROE_Latest'] < 15:
            quality_flags.append("DECLINING_ROE")
    
    # Negative or declining ROCE
    if result['ROCE_Latest'] is not None and result['ROCE_Latest'] < 10:
        quality_flags.append("LOW_ROCE")
    # Declining ROCE - same logic: only flag if ending ROCE is mediocre
    if result['ROCE_Trend'] is not None and result['ROCE_Trend'] < -2:
        if result['ROCE_Latest'] is not None and result['ROCE_Latest'] < 15:
            quality_flags.append("DECLINING_ROCE")
    
    # Poor cash conversion
    if result['CFO_PAT_3Yr_Avg'] is not None and result['CFO_PAT_3Yr_Avg'] < 0.7:
        quality_flags.append("POOR_CASH_CONVERSION")
    
    # Negative CFO
    if result['CFO_Latest'] is not None and result['CFO_Latest'] < 0:
        quality_flags.append("NEGATIVE_CFO")
    
    # Multiple years of negative CFO
    if result['Positive_CFO_Years'] < len(available_years) - 1:
        quality_flags.append("INCONSISTENT_CFO")
    
    # Exceptional items in multiple years
    if result['Exceptional_Items_Count'] >= 2:
        quality_flags.append("FREQUENT_EXCEPTIONALS")
    
    # High other income dependency - check for PERSISTENCE, not just latest year
    # CRITIQUE FIX: If Other Income > 30% for 2+ years, it's a "hedge fund masquerading as business"
    if result['High_Other_Income_Years'] >= 2:
        quality_flags.append("HIGH_OTHER_INCOME")
    elif result['Other_Income_Pct_PAT'] is not None and result['Other_Income_Pct_PAT'] > 50:
        # Single year but extremely high (>50%) is still a flag
        quality_flags.append("HIGH_OTHER_INCOME")
    
    # Increasing receivable days (cash collection issues)
    if result['Recv_Days_Trend'] is not None and result['Recv_Days_Trend'] > 10:
        quality_flags.append("RISING_RECEIVABLES")
    
    # Increasing inventory days (equally dangerous)
    # CRITIQUE FIX: Inventory spike is as dangerous as receivables spike
    if result['Inv_Days_Trend'] is not None and result['Inv_Days_Trend'] > 10:
        quality_flags.append("RISING_INVENTORY")
    
    # Margin compression
    if result['OPM_Trend'] is not None and result['OPM_Trend'] < -1:
        quality_flags.append("MARGIN_COMPRESSION")
    
    # Increasing debt
    if result['Debt_Trend'] is not None and result['Debt_Trend'] > 0.05:
        quality_flags.append("RISING_DEBT")
    
    # Working Capital Divergence - WC growing faster than revenue
    # CRITIQUE FIX: Catches channel stuffing / "buying" sales
    if result['WC_Rev_Growth_Ratio'] is not None and result['WC_Rev_Growth_Ratio'] > 1.5:
        quality_flags.append("WC_DIVERGENCE")
    
    # NPM vs OPM Divergence - bottom line growing without operating improvement
    # CRITIQUE FIX: Catches tax maneuvers and other income tricks
    if result['NPM_Trend'] is not None and result['OPM_Trend'] is not None:
        npm_opm_gap = result['NPM_Trend'] - result['OPM_Trend']
        if npm_opm_gap > 5:  # NPM improving 5%+ faster than OPM
            quality_flags.append("NPM_OPM_DIVERGENCE")
    
    # --- PRICING FLAGS (Valuation Concerns) ---
    
    if result['PE'] is not None:
        if result['PE'] > 50:
            pricing_flags.append("HIGH_PE")
        elif result['PE'] < 0:
            pricing_flags.append("NEGATIVE_PE")
    
    if result['EV_EBITDA'] is not None:
        if result['EV_EBITDA'] > 25:
            pricing_flags.append("HIGH_EV_EBITDA")
        elif result['EV_EBITDA'] < 0:
            # CRITIQUE FIX: Negative EV/EBITDA means negative EBITDA
            # Company loses money at operating level - cannot value via cash-flow metrics
            pricing_flags.append("NEGATIVE_EBITDA")
    
    # PBV vs ROE ratio check - CRITIQUE FIX: PBV of 3 with ROE of 8% is expensive
    # Formula: PBV should not exceed ROE/2 (e.g., ROE 20% → max PBV 10)
    if result['PBV'] is not None and result['ROE_Latest'] is not None and result['ROE_Latest'] > 0:
        max_fair_pbv = result['ROE_Latest'] / 2
        if result['PBV'] > max_fair_pbv:
            pricing_flags.append("HIGH_PBV_ROE")
    
    # Combine all flags
    all_flags = quality_flags + pricing_flags
    
    # ========================================
    # 10a. FINANCIAL SECTOR BYPASS (FIX 3.1)
    # ========================================
    # For Banks, NBFCs, Insurance, and Financial Services:
    # - EBITDA is meaningless (interest IS operating income)
    # - D/E is meaningless (leverage IS the business model)
    # - CFO is misleading (loan disbursements = negative CFO)
    # 
    # These flags should NOT be applied to financials - mark for manual review instead.
    
    # Use module-level pre-computed frozenset for O(1) lookups instead of O(n) .lower() per stock
    sector = result.get('Sector', '')
    industry = result.get('Industry', '')
    sector_lower = sector.lower()
    industry_lower = industry.lower()
    is_financial = any(fs in sector_lower or fs in industry_lower 
                       for fs in FINANCIAL_SECTORS_LOWER)
    
    result['Is_Financial_Sector'] = "Yes" if is_financial else "No"
    
    if is_financial:
        # Remove inapplicable flags for financials
        flags_to_remove = ['POOR_CASH_CONVERSION', 'NEGATIVE_CFO', 'INCONSISTENT_CFO', 
                          'NEGATIVE_EBITDA', 'HIGH_EV_EBITDA', 'RISING_DEBT']
        
        original_quality_flags = quality_flags.copy()
        original_pricing_flags = pricing_flags.copy()
        
        quality_flags = [f for f in quality_flags if f not in flags_to_remove]
        pricing_flags = [f for f in pricing_flags if f not in flags_to_remove]
        
        # Track what was bypassed
        bypassed = [f for f in original_quality_flags + original_pricing_flags if f in flags_to_remove]
        if bypassed:
            result['Financial_Sector_Bypass'] = f"Bypassed: {', '.join(bypassed)} (inapplicable to financials)"
            result['Manual_Review_Reason'] = "Financial sector - requires bank-specific metrics (NIM, GNPA, CAR)"
        else:
            result['Financial_Sector_Bypass'] = "None needed"
            result['Manual_Review_Reason'] = None
        
        # Update all_flags
        all_flags = quality_flags + pricing_flags
    else:
        result['Financial_Sector_Bypass'] = "N/A (non-financial)"
        result['Manual_Review_Reason'] = None
    
    result['Quality_Flag_Count'] = len(quality_flags)
    result['Pricing_Flag_Count'] = len(pricing_flags)
    result['Red_Flag_Count'] = len(all_flags)
    result['Quality_Flags'] = "|".join(quality_flags) if quality_flags else "NONE"
    result['Pricing_Flags'] = "|".join(pricing_flags) if pricing_flags else "NONE"
    result['Red_Flags'] = "|".join(all_flags) if all_flags else "NONE"
    
    # ========================================
    # 10b. SECTOR-AWARE SEVERITY ADJUSTMENT (NEW - FIX #2)
    # ========================================
    # BEFORE calculating severity, check if flags should be downgraded
    # because they may be sector-normal behavior, not quality issues.
    #
    # This prevents mechanical rejection of valid business models:
    # - NBFCs/Banks: Negative CFO during growth is normal (loan disbursements)
    # - Infrastructure: Long receivable cycles are normal (govt contracts)
    # - Capital-intensive: Poor cash conversion may be capex timing
    #
    # Downgraded flags still appear in output, but with reduced severity weight.
    
    sector_adjusted_flags = {}  # Track which flags were downgraded and why
    
    # Check POOR_CASH_CONVERSION - may be capital-intensive timing issue
    if "POOR_CASH_CONVERSION" in quality_flags:
        if result['Asset_Turnover'] is not None and result['Asset_Turnover'] < 0.8:
            # Capital-intensive business - CFO timing may be normal
            sector_adjusted_flags["POOR_CASH_CONVERSION"] = {
                "original_severity": "CRITICAL",
                "adjusted_severity": "MINOR",
                "adjusted_weight": 0.5,
                "reason": "Capital-intensive (Asset_Turnover < 0.8)—CFO timing may be normal"
            }
    
    # Check NEGATIVE_CFO - may be growth-phase NBFC/lender or capex cycle
    if "NEGATIVE_CFO" in quality_flags:
        if result['Asset_Turnover'] is not None and result['Asset_Turnover'] < 0.5:
            # Very capital-intensive (likely NBFC, infra, heavy industry)
            sector_adjusted_flags["NEGATIVE_CFO"] = {
                "original_severity": "CRITICAL",
                "adjusted_severity": "MAJOR",
                "adjusted_weight": 1.0,
                "reason": "Very capital-intensive (Asset_Turnover < 0.5)—may be growth capex or lending"
            }
    
    # Check INCONSISTENT_CFO - may be project-based or cyclical business
    if "INCONSISTENT_CFO" in quality_flags:
        if result['Asset_Turnover'] is not None and result['Asset_Turnover'] < 0.8:
            sector_adjusted_flags["INCONSISTENT_CFO"] = {
                "original_severity": "CRITICAL",
                "adjusted_severity": "MAJOR",
                "adjusted_weight": 1.0,
                "reason": "Capital-intensive—CFO may be lumpy due to project timing"
            }
    
    # Check RISING_RECEIVABLES - may be IT services or infrastructure
    if "RISING_RECEIVABLES" in quality_flags:
        if result['Receivable_Days'] is not None and result['Receivable_Days'] > 60:
            # Long receivable cycles may be sector-normal
            sector_adjusted_flags["RISING_RECEIVABLES"] = {
                "original_severity": "MINOR",
                "adjusted_severity": "MINOR",  # Already MINOR, just note it
                "adjusted_weight": 0.5,
                "reason": "High receivable days (>60)—may be IT services or govt contractor"
            }
    
    # ========================================
    # 10c. SEVERITY SCORE (Weighted flag score)
    # ========================================
    # CRITICAL flags = 2.0, MAJOR = 1.0, MINOR = 0.5
    # BUT: Use adjusted weights if sector-normal conditions detected
    quality_severity = 0.0
    pricing_severity = 0.0
    critical_flags = []
    major_flags = []
    minor_flags = []
    
    for flag in quality_flags:
        defn = STRUCTURAL_RED_FLAGS.get(flag, {})
        
        # Check if this flag was sector-adjusted
        if flag in sector_adjusted_flags:
            weight = sector_adjusted_flags[flag]["adjusted_weight"]
            severity = sector_adjusted_flags[flag]["adjusted_severity"]
        else:
            weight = defn.get('weight', 1.0)
            severity = defn.get('severity', 'MAJOR')
        
        quality_severity += weight
        
        # Categorize by EFFECTIVE severity (after adjustment)
        if severity == 'CRITICAL':
            critical_flags.append(flag)
        elif severity == 'MAJOR':
            major_flags.append(flag)
        else:
            minor_flags.append(flag)
    
    for flag in pricing_flags:
        defn = PRICING_RED_FLAGS.get(flag, {})
        weight = defn.get('weight', 0.5)
        severity = defn.get('severity', 'MINOR')
        pricing_severity += weight
        if severity == 'CRITICAL':
            critical_flags.append(flag)
        elif severity == 'MAJOR':
            major_flags.append(flag)
        else:
            minor_flags.append(flag)
    
    result['Quality_Severity'] = round(quality_severity, 1)
    result['Pricing_Severity'] = round(pricing_severity, 1)
    result['Total_Severity'] = round(quality_severity + pricing_severity, 1)
    result['Critical_Flags'] = "|".join(critical_flags) if critical_flags else "NONE"
    result['Major_Flags'] = "|".join(major_flags) if major_flags else "NONE"
    result['Minor_Flags'] = "|".join(minor_flags) if minor_flags else "NONE"
    
    # Track sector adjustments made
    if sector_adjusted_flags:
        adjustments = [f"{flag}: {info['original_severity']}→{info['adjusted_severity']} ({info['reason']})" 
                      for flag, info in sector_adjusted_flags.items()]
        result['Sector_Adjustments_Made'] = " | ".join(adjustments)
    else:
        result['Sector_Adjustments_Made'] = "None"
    
    # ========================================
    # 10d. SECTOR ADJUSTMENT NEEDED FLAG (Post-hoc warnings)
    # ========================================
    # Flags that may STILL need manual sector verification
    # (even after automatic adjustments above)
    sector_warnings = []
    
    # High Other Income may be normal for holding companies
    # Can't auto-adjust without sector classification, so just warn
    if "HIGH_OTHER_INCOME" in quality_flags:
        sector_warnings.append("Other_Income_May_Be_Business_Model")
    
    # Rising receivables may be normal for IT services (60-90 day cycles)
    if "RISING_RECEIVABLES" in quality_flags:
        if result['Receivable_Days'] is not None and result['Receivable_Days'] > 45:
            sector_warnings.append("Receivables_May_Be_Sector_Normal")
    
    # Poor cash conversion may be normal for capital-intensive businesses
    if "POOR_CASH_CONVERSION" in quality_flags:
        if result['Asset_Turnover'] is not None and result['Asset_Turnover'] < 0.8:
            sector_warnings.append("CFO_May_Be_Capex_Timing")
    
    # Margin compression may be cyclical, not structural
    if "MARGIN_COMPRESSION" in quality_flags:
        sector_warnings.append("Margins_May_Be_Cyclical")
    
    if sector_warnings:
        result['Sector_Adjustment_Needed'] = "Yes"
        result['Sector_Warnings'] = "|".join(sector_warnings)
    else:
        result['Sector_Adjustment_Needed'] = "No"
        result['Sector_Warnings'] = "None"
    
    # Build detailed explanation string
    if all_flags:
        explanations = []
        for f in all_flags:
            defn = RED_FLAG_DEFINITIONS.get(f, {})
            severity = defn.get('severity', 'MAJOR')
            explanations.append(f"[{severity}] {defn.get('name', f)}: {defn.get('concern', 'N/A')}")
        result['Red_Flags_Explained'] = " || ".join(explanations)
    else:
        result['Red_Flags_Explained'] = "No red flags detected"
    
    # Add individual boolean columns for each red flag (1 = flagged, 0 = clean)
    for flag_code in RED_FLAG_DEFINITIONS.keys():
        result[f'FLAG_{flag_code}'] = 1 if flag_code in all_flags else 0
    
    # ========================================
    # 11. EARNINGS QUALITY SUMMARY (Clean/Mixed/Aggressive)
    # ========================================
    earnings_issues = 0
    if result['CFO_PAT_3Yr_Avg'] is not None and result['CFO_PAT_3Yr_Avg'] < 0.8:
        earnings_issues += 1
    if result['Accruals'] is not None and result['Accruals'] > 0.5:
        earnings_issues += 1
    if "FREQUENT_EXCEPTIONALS" in quality_flags:
        earnings_issues += 1
    if "HIGH_OTHER_INCOME" in quality_flags:
        earnings_issues += 1
    if result['Positive_CFO_Years'] < len(available_years) - 1:
        earnings_issues += 1
    
    if earnings_issues == 0:
        result['Earnings_Quality'] = "Clean"
    elif earnings_issues <= 2:
        result['Earnings_Quality'] = "Mixed"
    else:
        result['Earnings_Quality'] = "Aggressive"
    
    # ========================================
    # 12. CAPITAL INTENSITY CLASSIFIER
    # ========================================
    # Based on Asset Turnover: High turnover = asset-light, Low turnover = capital-intensive
    asset_turnover = result['Asset_Turnover']
    if asset_turnover is not None:
        if asset_turnover > 2.0:
            result['Capital_Intensity'] = "Asset-Light"
        elif asset_turnover > 0.8:
            result['Capital_Intensity'] = "Moderate"
        else:
            result['Capital_Intensity'] = "Capital-Intensive"
    else:
        result['Capital_Intensity'] = "Unknown"
    
    # ========================================
    # 13. DATA CONFIDENCE SCORE
    # ========================================
    # Based on: years of data, missing critical fields, data completeness
    confidence = 100
    
    # Penalize for fewer years of data
    if result['Years_Available'] < 5:
        confidence -= (5 - result['Years_Available']) * 10
    
    # Penalize for missing critical fields (latest year)
    critical_fields = ['ROE_Latest', 'ROCE_Latest', 'CFO_Latest', 'PE', 'OPM_Latest']
    missing_critical = sum(1 for f in critical_fields if result.get(f) is None)
    confidence -= missing_critical * 10
    
    # Penalize for missing trend data
    trend_fields = ['ROE_Trend', 'ROCE_Trend', 'CFO_Trend', 'OPM_Trend']
    missing_trends = sum(1 for f in trend_fields if result.get(f) is None)
    confidence -= missing_trends * 5
    
    # CRITIQUE FIX: Penalize for sparse data WITHIN available years
    # Check how many years have complete core data (ROE, ROCE, CFO, Revenue)
    complete_years = 0
    for y in available_years:
        yr_data = yearly_data.get(y, {})
        has_roe = safe_float(safe_get(yr_data, 'ROE_A')) is not None
        has_roce = safe_float(safe_get(yr_data, 'ROCE_A')) is not None
        has_cfo = safe_float(safe_get(yr_data, 'CFO_A')) is not None
        has_rev = safe_float(safe_get(yr_data, 'SR_A')) is not None
        if has_roe and has_roce and has_cfo and has_rev:
            complete_years += 1
    
    # Penalize if less than 60% of years have complete data
    if result['Years_Available'] > 0:
        completeness_ratio = complete_years / result['Years_Available']
        if completeness_ratio < 0.6:
            confidence -= 15
        elif completeness_ratio < 0.8:
            confidence -= 10
    
    result['Data_Confidence'] = max(0, min(100, confidence))
    
    # ========================================
    # 14. VALUATION BAND (Cheap/Fair/Expensive)
    # ========================================
    pe = result['PE']
    ev_ebitda = result['EV_EBITDA']
    
    valuation_score = 0
    valuation_factors = 0
    
    if pe is not None and pe > 0:
        if pe < 15:
            valuation_score += 2  # Cheap
        elif pe < 25:
            valuation_score += 1  # Fair
        elif pe < 40:
            valuation_score += 0  # Slightly expensive
        else:
            valuation_score -= 1  # Expensive
        valuation_factors += 1
    
    if ev_ebitda is not None and ev_ebitda > 0:
        if ev_ebitda < 10:
            valuation_score += 2  # Cheap
        elif ev_ebitda < 15:
            valuation_score += 1  # Fair
        elif ev_ebitda < 20:
            valuation_score += 0  # Slightly expensive
        else:
            valuation_score -= 1  # Expensive
        valuation_factors += 1
    
    if valuation_factors > 0:
        avg_val_score = valuation_score / valuation_factors
        if avg_val_score >= 1.5:
            result['Valuation_Band'] = "Cheap"
        elif avg_val_score >= 0.5:
            result['Valuation_Band'] = "Fair"
        elif avg_val_score >= -0.5:
            result['Valuation_Band'] = "Expensive"
        else:
            result['Valuation_Band'] = "Very_Expensive"
    else:
        result['Valuation_Band'] = "Unknown"
    
    # ========================================
    # 15. FOUR SUB-SCORES (Replaces single composite)
    # ========================================
    
    # --- A) BUSINESS QUALITY SCORE (0-100) ---
    # Based on: ROE, ROCE, margins, trends, leverage-driven check
    bq_score = 50  # Start at midpoint
    
    if result['ROE_Latest'] is not None:
        if result['ROE_Latest'] >= 20: bq_score += 15
        elif result['ROE_Latest'] >= 15: bq_score += 10
        elif result['ROE_Latest'] >= 10: bq_score += 5
        elif result['ROE_Latest'] < 5: bq_score -= 15
    
    if result['ROCE_Latest'] is not None:
        if result['ROCE_Latest'] >= 20: bq_score += 15
        elif result['ROCE_Latest'] >= 15: bq_score += 10
        elif result['ROCE_Latest'] >= 10: bq_score += 5
        elif result['ROCE_Latest'] < 5: bq_score -= 15
    
    if result['OPM_Latest'] is not None:
        if result['OPM_Latest'] >= 20: bq_score += 10
        elif result['OPM_Latest'] >= 12: bq_score += 5
        elif result['OPM_Latest'] < 5: bq_score -= 10
    
    # Penalize if ROE is leverage-driven
    if result['Leverage_Driven'] == "Yes":
        bq_score -= 15
    
    # Trend adjustments
    if result['ROE_Trend'] is not None and result['ROE_Trend'] < -2:
        bq_score -= 10
    if result['ROCE_Trend'] is not None and result['ROCE_Trend'] < -2:
        bq_score -= 10
    
    # CRITIQUE FIX: Standalone Transparency Discount
    # If only standalone data available (no consolidated), we can't see subsidiaries
    # Losses, debt, and "bad" business may be parked in unlisted subsidiaries
    if result['Data_Type'] == 'standalone':
        bq_score -= 10  # Assume the worst if we can't see the whole group
        result['Transparency_Discount'] = "Yes"
    else:
        result['Transparency_Discount'] = "No"
    
    result['Business_Quality_Score'] = max(0, min(100, bq_score))
    
    # --- B) FINANCIAL STRENGTH SCORE (0-100) ---
    # Based on: debt levels, interest coverage, liquidity, cash flow
    fs_score = 50
    
    if result['Debt_Equity'] is not None:
        if result['Debt_Equity'] <= 0.1: fs_score += 20
        elif result['Debt_Equity'] <= 0.3: fs_score += 10
        elif result['Debt_Equity'] <= 0.5: fs_score += 5
        elif result['Debt_Equity'] > 1: fs_score -= 15
    
    if result['Interest_Coverage'] is not None:
        if result['Interest_Coverage'] >= 10: fs_score += 15
        elif result['Interest_Coverage'] >= 5: fs_score += 10
        elif result['Interest_Coverage'] >= 2: fs_score += 5
        elif result['Interest_Coverage'] < 1.5: fs_score -= 15
    
    if result['Current_Ratio'] is not None:
        if result['Current_Ratio'] >= 2: fs_score += 10
        elif result['Current_Ratio'] >= 1.5: fs_score += 5
        elif result['Current_Ratio'] < 1: fs_score -= 15
    
    if result['CFO_Latest'] is not None:
        if result['CFO_Latest'] > 0: fs_score += 10
        else: fs_score -= 20
    
    # Rising debt is a concern
    if result['Debt_Trend'] is not None and result['Debt_Trend'] > 0.05:
        fs_score -= 10
    
    result['Financial_Strength_Score'] = max(0, min(100, fs_score))
    
    # --- C) GROWTH DURABILITY SCORE (0-100) ---
    # Based on: consistency, volatility, dilution
    gd_score = 50
    
    if result['Revenue_Volatility'] is not None:
        if result['Revenue_Volatility'] < 15: gd_score += 15  # Very stable
        elif result['Revenue_Volatility'] < 25: gd_score += 10  # Stable
        elif result['Revenue_Volatility'] < 40: gd_score += 5  # Moderate
        elif result['Revenue_Volatility'] > 60: gd_score -= 15  # Very volatile
    
    if result['Profit_Consistency'] is not None:
        if result['Profit_Consistency'] >= 100: gd_score += 15  # Always profitable
        elif result['Profit_Consistency'] >= 80: gd_score += 10
        elif result['Profit_Consistency'] >= 60: gd_score += 5
        elif result['Profit_Consistency'] < 40: gd_score -= 15
    
    if result['Profit_Growth_Consistency'] is not None:
        if result['Profit_Growth_Consistency'] >= 80: gd_score += 10
        elif result['Profit_Growth_Consistency'] >= 60: gd_score += 5
        elif result['Profit_Growth_Consistency'] < 40: gd_score -= 10
    
    # Penalize heavy dilution
    if result['Share_Dilution_Pct'] is not None:
        if result['Share_Dilution_Pct'] > 50: gd_score -= 15  # Heavy dilution
        elif result['Share_Dilution_Pct'] > 20: gd_score -= 10
        elif result['Share_Dilution_Pct'] < 5: gd_score += 5  # Minimal dilution
    
    result['Growth_Durability_Score'] = max(0, min(100, gd_score))
    
    # --- D) VALUATION COMFORT SCORE (0-100) ---
    # Based on: PE, EV/EBITDA, pricing flags
    vc_score = 50
    
    if result['Valuation_Band'] == "Cheap":
        vc_score += 25
    elif result['Valuation_Band'] == "Fair":
        vc_score += 10
    elif result['Valuation_Band'] == "Expensive":
        vc_score -= 10
    elif result['Valuation_Band'] == "Very_Expensive":
        vc_score -= 25
    
    # Additional granular adjustments
    if pe is not None and pe > 0:
        if pe < 12: vc_score += 10
        elif pe > 50: vc_score -= 15
    
    if ev_ebitda is not None:
        if ev_ebitda < 8: vc_score += 10
        elif ev_ebitda > 25: vc_score -= 15
    
    # Negative PE (loss-making) is bad for valuation comfort
    if pe is not None and pe < 0:
        vc_score -= 30
    
    result['Valuation_Comfort_Score'] = max(0, min(100, vc_score))
    
    # ========================================
    # 16. COMPOSITE SCORE (Weighted average of sub-scores)
    # ========================================
    # Weights: Business Quality 35%, Financial Strength 25%, Growth Durability 20%, Valuation 20%
    composite = (
        result['Business_Quality_Score'] * 0.35 +
        result['Financial_Strength_Score'] * 0.25 +
        result['Growth_Durability_Score'] * 0.20 +
        result['Valuation_Comfort_Score'] * 0.20
    )
    
    # --- STRUCTURAL FLAG HARD PENALTIES ---
    # These override the weighted average
    structural_penalty = len(quality_flags) * 10
    composite -= structural_penalty
    
    if len(quality_flags) >= 2:
        composite -= 15  # Extra penalty
    if len(quality_flags) >= 3:
        composite -= 15  # Even more
    
    # --- DATA CONFIDENCE HARD CAP ---
    if result['Data_Confidence'] < 40:
        composite = min(composite, 40)
    elif result['Data_Confidence'] < 60:
        composite = min(composite, 60)
    
    # --- FINANCIAL SECTOR ANALYSIS GAP ---
    # For financial sector stocks, we removed flags (D/E, EBITDA, CFO) without
    # replacing them with banking-specific metrics (NPA, NIM, CAR).
    # The score is structurally inflated by the absence of applicable checks.
    # Apply a confidence discount to reflect this analysis gap.
    if result.get('Is_Financial_Sector') == "Yes":
        bypassed = result.get('Financial_Sector_Bypass', '')
        if 'Bypassed' in str(bypassed):
            # Count bypassed flags: each one is a check we couldn't do
            bypass_count = str(bypassed).count(',') + 1
            financial_penalty = min(bypass_count * 5, 20)  # Max 20-point penalty
            composite -= financial_penalty
            result['Financial_Analysis_Gap'] = f"-{financial_penalty}pts ({bypass_count} checks inapplicable)"
        else:
            result['Financial_Analysis_Gap'] = "None (no flags were bypassed)"
    else:
        result['Financial_Analysis_Gap'] = "N/A"
    
    # --- STRUCTURAL FLAG HARD CAP ---
    if len(quality_flags) >= 2:
        composite = min(composite, 50)
    
    result['Composite_Score'] = max(0, min(100, round(composite)))
    
    # ========================================
    # 17. QUALITY RISK LEVEL (Sort by this, not score)
    # ========================================
    if len(quality_flags) == 0:
        result['Quality_Risk'] = "LOW"
    elif len(quality_flags) == 1:
        result['Quality_Risk'] = "MEDIUM"
    else:
        result['Quality_Risk'] = "HIGH"
    
    # ========================================
    # 17b-PREREQ. LOAD SHAREHOLDING EARLY (needed for Pledge Gate in SCREEN_ELIGIBLE)
    # ========================================
    # FIX: Shareholding must be loaded BEFORE SCREEN_ELIGIBLE check,
    # otherwise Pledge_Gate_Failed is always None and the pledge kill-switch never fires.
    shareholding = load_shareholding_data(filepath, shareholding_dir, resolved_path=_resolved_sh)
    if shareholding:
        result['Promoter_Holding'] = shareholding.get('Promoter_Holding')
        result['Promoter_Pledge'] = shareholding.get('Promoter_Pledge')
        result['FII_Holding'] = shareholding.get('FII_Holding')
        result['MF_Holding'] = shareholding.get('MF_Holding')
        result['Public_Holding'] = shareholding.get('Public_Holding')
        result['Others_Holding'] = shareholding.get('Others_Holding')
        result['Promoter_Change_1Yr'] = shareholding.get('Promoter_Change_1Yr')
        result['FII_Change_1Yr'] = shareholding.get('FII_Change_1Yr')
        result['MF_Change_1Yr'] = shareholding.get('MF_Change_1Yr')
        
        # Governance flags based on shareholding
        governance_flags = []
        
        # Pledge thresholds (Fix 3.6 - India-Specific Kill Switches)
        pledge_pct = result['Promoter_Pledge']
        if pledge_pct is not None:
            if pledge_pct > 50:
                governance_flags.append(f"CRITICAL_PLEDGE ({pledge_pct:.1f}%)")
                result['Pledge_Risk'] = "CRITICAL"
                result['Pledge_Gate_Failed'] = True
            elif pledge_pct > 25:
                governance_flags.append(f"SEVERE_PLEDGE ({pledge_pct:.1f}%)")
                result['Pledge_Risk'] = "SEVERE"
                result['Pledge_Gate_Failed'] = False
            elif pledge_pct > 10:
                governance_flags.append(f"ELEVATED_PLEDGE ({pledge_pct:.1f}%)")
                result['Pledge_Risk'] = "ELEVATED"
                result['Pledge_Gate_Failed'] = False
            else:
                result['Pledge_Risk'] = "LOW"
                result['Pledge_Gate_Failed'] = False
        else:
            result['Pledge_Risk'] = None
            result['Pledge_Gate_Failed'] = False
        
        if result['Promoter_Change_1Yr'] is not None and result['Promoter_Change_1Yr'] < -5:
            governance_flags.append(f"PROMOTER_SELLING ({result['Promoter_Change_1Yr']:.1f}%)")
        if result['FII_Change_1Yr'] is not None and result['FII_Change_1Yr'] < -5:
            governance_flags.append(f"FII_EXIT ({result['FII_Change_1Yr']:.1f}%)")
        
        result['Governance_Flags'] = "; ".join(governance_flags) if governance_flags else "None"
        result['Shareholding_Available'] = "Yes"
    else:
        result['Promoter_Holding'] = None
        result['Promoter_Pledge'] = None
        result['FII_Holding'] = None
        result['MF_Holding'] = None
        result['Public_Holding'] = None
        result['Others_Holding'] = None
        result['Promoter_Change_1Yr'] = None
        result['FII_Change_1Yr'] = None
        result['MF_Change_1Yr'] = None
        result['Governance_Flags'] = "Data not available"
        result['Shareholding_Available'] = "No"
        result['Pledge_Risk'] = None
        result['Pledge_Gate_Failed'] = False
    
    # ========================================
    # 17b-PREREQ2. LOAD INSIDER TRADING DATA (needed for Conviction Override context)
    # ========================================
    # Insider trading data provides CRITICAL context for interpreting promoter
    # shareholding changes. Without it, a promoter family consolidation (inter-se
    # transfer) is indistinguishable from genuine insider exit selling.
    #
    # Example: TINNARUBR promoter dropped from 73.55% to 67.58% (-5.97pp).
    #   Shareholding alone → looks like massive promoter exit → VALUE_TRAP
    #   With insider trading → reveals inter-se family consolidation + MF entry + 
    #   promoter trust buying on dips → NOT a value trap at all
    insider_trading = load_insider_trading_data(
        filepath, insider_trading_dir, resolved_path=_resolved_it,
        market_cap_cr=result.get('Market_Cap')
    )
    if insider_trading:
        result['Insider_Trading_Available'] = "Yes"
        result['Insider_Action'] = insider_trading.get('Insider_Action', 'UNKNOWN')
        result['Insider_Sentiment'] = insider_trading.get('Insider_Sentiment', 'UNKNOWN')
        result['Inter_Se_Pct_Of_Disposals'] = insider_trading.get('Inter_Se_Pct_Of_Disposals', 0)
        result['Market_Sell_Pct_MCap'] = insider_trading.get('Market_Sell_Pct_MCap')
        result['Recent_Market_Buy_6m'] = insider_trading.get('Recent_Market_Buy_6m', False)
        result['Market_Buy_Value_1Yr'] = insider_trading.get('Market_Buy_Value_1Yr', 0)
        result['Market_Sell_Value_1Yr'] = insider_trading.get('Market_Sell_Value_1Yr', 0)
        result['Insider_Context'] = insider_trading.get('Insider_Context', '')
    else:
        result['Insider_Trading_Available'] = "No"
        result['Insider_Action'] = None
        result['Insider_Sentiment'] = None
        result['Inter_Se_Pct_Of_Disposals'] = None
        result['Market_Sell_Pct_MCap'] = None
        result['Recent_Market_Buy_6m'] = None
        result['Market_Buy_Value_1Yr'] = None
        result['Market_Sell_Value_1Yr'] = None
        result['Insider_Context'] = None
    
    # ========================================
    # 17b. SCREEN_ELIGIBLE (Hard mechanical rule - NO DISCRETION)
    # ========================================
    # This is a binary gate, not a score. Either eligible or not.
    # UPDATED: Uses severity scores instead of flag counts
    screen_blockers = []
    
    if result['Quality_Severity'] >= 2.0:
        if critical_flags:
            screen_blockers.append(f"Critical flag: {critical_flags[0]}")
        else:
            screen_blockers.append(f"High severity ({result['Quality_Severity']})")
    if result['Data_Confidence'] < 50:
        screen_blockers.append("Low data confidence")
    if result['Earnings_Quality'] == "Aggressive":
        screen_blockers.append("Aggressive accounting")
    
    # Fix 3.6: Critical pledge threshold (>50%) is a hard blocker
    # NOTE: Pledge_Gate_Failed is now guaranteed to be set above
    if result.get('Pledge_Gate_Failed'):
        pledge_pct = result.get('Promoter_Pledge', 0)
        screen_blockers.append(f"Critical pledge ({pledge_pct:.1f}%) - survival risk")
    
    # Fix 3.1: Financial sector requires manual review (not auto-fail, but flag it)
    if result.get('Is_Financial_Sector') == "Yes" and result.get('Manual_Review_Reason'):
        # Don't block, but note it
        result['Requires_Manual_Review'] = result['Manual_Review_Reason']
    else:
        result['Requires_Manual_Review'] = None
    
    if screen_blockers:
        result['SCREEN_ELIGIBLE'] = "NO"
        result['Screen_Blocker'] = screen_blockers[0]  # Primary reason
    else:
        result['SCREEN_ELIGIBLE'] = "YES"
        result['Screen_Blocker'] = "None"
    
    # ========================================
    # 17c. CYCLIC PEAK RISK DETECTION
    # ========================================
    # Detects if company might be at earnings peak (dangerous for entry)
    # Signals: margin at multi-year high + revenue growth slowing
    cyclic_risk_signals = 0
    
    # Check if OPM is at/near peak (latest > 3yr avg by significant margin)
    if result['OPM_Latest'] is not None and result['OPM_Trend'] is not None:
        opm_3yr = [safe_float(safe_get(yearly_data.get(y, {}), 'OPMPCT_A')) for y in available_years[:3]]
        valid_opm = [o for o in opm_3yr if o is not None]
        if len(valid_opm) >= 2:
            opm_avg = mean(valid_opm)
            if result['OPM_Latest'] > opm_avg * 1.2:  # 20% above average
                cyclic_risk_signals += 1
    
    # Check if revenue growth is decelerating
    if result['Revenue_Growth_1Yr'] is not None and result['Revenue_Growth_3Yr'] is not None:
        if result['Revenue_Growth_3Yr'] is not None and result['Revenue_Growth_3Yr'] > 0:
            if result['Revenue_Growth_1Yr'] < result['Revenue_Growth_3Yr'] * 0.5:  # 1Y growth < half of 3Y
                cyclic_risk_signals += 1
    
    # Check if ROCE is very high but growth is low (reinvestment saturation)
    if result['ROCE_Latest'] is not None and result['ROCE_Latest'] > 25:
        if result['Revenue_Growth_1Yr'] is not None and result['Revenue_Growth_1Yr'] < 10:
            cyclic_risk_signals += 1
    
    if cyclic_risk_signals >= 2:
        result['Cyclic_Peak_Risk'] = "HIGH"
    elif cyclic_risk_signals == 1:
        result['Cyclic_Peak_Risk'] = "MODERATE"
    else:
        result['Cyclic_Peak_Risk'] = "LOW"
    
    # ========================================
    # 16. SCORE BAND (Use this for decisions, not raw score)
    # ========================================
    # The difference between 88 and 92 is noise. Use bands.
    if result['Composite_Score'] >= 80:
        result['Score_Band'] = "A"  # Strong candidate
    elif result['Composite_Score'] >= 65:
        result['Score_Band'] = "B"  # Decent, worth analysis
    elif result['Composite_Score'] >= 50:
        result['Score_Band'] = "C"  # Marginal
    elif result['Composite_Score'] >= 30:
        result['Score_Band'] = "D"  # Weak
    else:
        result['Score_Band'] = "F"  # Avoid
    
    # ========================================
    # 17. DECISION BUCKET (Severity-Based, Neutral Language)
    # ========================================
    # UPDATED: Use severity scores instead of flag counts
    # - CRITICAL flags = 2.0 weight, MAJOR = 1.0, MINOR = 0.5
    # - Severity >= 2.0 = SCREEN_FAILED (equivalent to old "2 flags" but weighted)
    # 
    # NEUTRAL LANGUAGE (doesn't imply recommendation):
    #   GATES_CLEARED           → Passed quality gates (was CORE_CANDIDATE)
    #   SCREEN_PASSED_VERIFY    → Passed but needs verification (standalone data)
    #   SCREEN_PASSED_EXPENSIVE → Passed quality, expensive (was GOOD_WAIT_PRICE)
    #   SCREEN_PASSED_FLAGS     → Passed with concerns (was GOOD_BUT_RISKY)
    #   DATA_INCOMPLETE         → Insufficient data (was NEEDS_MORE_DATA)
    #   SCREEN_MARGINAL         → Borderline metrics (was MARGINAL)
    #   SCREEN_FAILED           → Failed quality gates (was REJECT)
    
    if result['Quality_Severity'] >= 2.0:
        # HARD RULE: Severity >= 2.0 = failed (one CRITICAL or two MAJOR flags)
        result['Decision_Bucket'] = "SCREEN_FAILED"
        if critical_flags:
            result['Reject_Reason'] = f"Critical flag: {critical_flags[0]}"
        else:
            result['Reject_Reason'] = f"Multiple major flags (severity {result['Quality_Severity']})"
    elif result['Data_Confidence'] < 40:
        # Very low data confidence = can't analyze, but NOT a bad business
        result['Decision_Bucket'] = "DATA_INCOMPLETE"
        result['Reject_Reason'] = "Critical data gap (confidence < 40)"
    elif result['Data_Confidence'] < 60:
        # Low confidence = can't be core, regardless of score
        result['Decision_Bucket'] = "DATA_INCOMPLETE"
        result['Reject_Reason'] = "Low data confidence"
    elif result['Cyclic_Peak_Risk'] == 'HIGH' and len(pricing_flags) >= 1:
        # FIX 2: Cyclic peak + expensive valuation = TIMING TRAP
        # Business at peak earnings + priced for perfection = maximum downside risk
        result['Decision_Bucket'] = "SCREEN_FAILED"
        result['Reject_Reason'] = f"Cyclic peak with expensive valuation ({pricing_flags[0]})—timing trap"
    elif result['Quality_Severity'] >= 1.0:
        # Severity 1.0-1.9 = passed with concerns (one MAJOR flag or two MINOR)
        result['Decision_Bucket'] = "SCREEN_PASSED_FLAGS"
        result['Reject_Reason'] = f"Quality concerns (severity {result['Quality_Severity']})"
    elif result['Quality_Severity'] >= 0.5:
        # Severity 0.5-0.9 = minor concern only
        result['Decision_Bucket'] = "SCREEN_PASSED_FLAGS"
        result['Reject_Reason'] = f"Minor flag: {minor_flags[0] if minor_flags else 'N/A'}"
    elif len(pricing_flags) >= 1 and result['Score_Band'] in ['A', 'B']:
        # Good business but expensive = wait for better price
        result['Decision_Bucket'] = "SCREEN_PASSED_EXPENSIVE"
        result['Reject_Reason'] = f"Valuation concern: {pricing_flags[0]}"
    elif result['Earnings_Quality'] == "Aggressive":
        # Aggressive accounting = risky
        result['Decision_Bucket'] = "SCREEN_PASSED_FLAGS"
        result['Reject_Reason'] = "Aggressive earnings quality"
    elif result['Score_Band'] in ['A', 'B'] and result['Earnings_Quality'] != "Aggressive":
        # Clean slate, good score = screen passed
        result['Decision_Bucket'] = "GATES_CLEARED"
        result['Reject_Reason'] = "None"
    elif result['Score_Band'] == 'C':
        # Marginal score = needs more analysis
        result['Decision_Bucket'] = "SCREEN_MARGINAL"
        result['Reject_Reason'] = "Marginal score"
    else:
        # Everything else = failed
        result['Decision_Bucket'] = "SCREEN_FAILED"
        result['Reject_Reason'] = "Low score"
    
    # FIX 1: Standalone data = requires verification (upgrade to SCREEN_PASSED_VERIFY)
    # If we passed but only have standalone data, we can't see subsidiaries
    if result['Data_Type'] == 'standalone' and result['Decision_Bucket'] in ['GATES_CLEARED', 'SCREEN_PASSED_EXPENSIVE']:
        result['Decision_Bucket'] = "SCREEN_PASSED_VERIFY"
        result['Reject_Reason'] = "Standalone data only—verify group financials before investing"
    
    # ========================================
    # 17a-POST. CONVICTION OVERRIDES (Stochastic Inversions)
    # ========================================
    # These detect "contradictory signals" — situations where the score says one
    # thing but insider behavior (promoter buying/selling) says the opposite.
    # Such contradictions are more informative than the score itself.
    #
    # CRITICAL ENHANCEMENT: Now cross-references insider trading data to avoid
    # false VALUE_TRAP signals from inter-se transfers (family restructuring).
    #
    # Classification hierarchy for promoter shareholding drops:
    #   1. If insider trading shows >50% of disposals are inter-se → NOT real selling
    #   2. If promoter has been buying in open market recently → conviction signal
    #   3. If open market sell value < 3% of MCap → minor liquidity, not exit
    #   4. Only if genuine open-market selling dominates → VALUE_TRAP
    #
    # CONTRARIAN_BET: Low score + significant promoter buying = insiders disagree
    #   with the market. Don't trash it — flag it for special attention.
    # VALUE_TRAP: High score + significant promoter selling = insiders are exiting
    #   despite good-looking numbers. The numbers may be deceiving.
    
    promoter_change = result.get('Promoter_Change_1Yr')
    original_bucket = result['Decision_Bucket']
    result['Conviction_Override'] = "None"
    
    # Extract insider trading context (may be None if data unavailable)
    it_available = result.get('Insider_Trading_Available') == "Yes"
    inter_se_pct = result.get('Inter_Se_Pct_Of_Disposals', 0) or 0
    insider_action = result.get('Insider_Action')
    recent_buy = result.get('Recent_Market_Buy_6m', False)
    market_sell_pct_mcap = result.get('Market_Sell_Pct_MCap')
    
    if promoter_change is not None:
        # CONTRARIAN_BET: Failed or marginal stock + promoter accumulating > 3%
        if (result['Decision_Bucket'] in ['SCREEN_FAILED', 'SCREEN_MARGINAL'] 
                and promoter_change >= 3.0
                and result.get('Quality_Severity', 99) < 4.0  # Not catastrophically broken
                and result.get('Data_Confidence', 0) >= 50):  # Decent data quality
            
            # Enhancement: Strengthen CONTRARIAN_BET if insider trading confirms buying
            contrarian_note = f"Promoter buying +{promoter_change:.1f}% despite low score"
            if it_available and insider_action in ('ONLY_BUYING', 'NET_BUYER'):
                contrarian_note += " (confirmed by open-market purchases)"
            
            result['Decision_Bucket'] = "CONTRARIAN_BET"
            result['Conviction_Override'] = contrarian_note
            result['Reject_Reason'] = (
                f"Score={result['Composite_Score']} (was {original_bucket}) BUT "
                f"promoter increased stake by {promoter_change:.1f}%—special situation"
            )
        
        # VALUE_TRAP: Good score + promoter dumping > 3%
        # BUT NOW: Cross-reference with insider trading to avoid false positives
        elif (result['Decision_Bucket'] in ['GATES_CLEARED', 'SCREEN_PASSED_EXPENSIVE']
                and promoter_change <= -3.0):
            
            # ============================================================
            # INSIDER TRADING CONTEXT CHECK (new logic)
            # ============================================================
            # Before flagging VALUE_TRAP, check if the promoter shareholding drop
            # is actually benign (inter-se transfer, minor HUF selling, etc.)
            
            value_trap_mitigated = False
            mitigation_reason = ""
            
            if it_available:
                # Check 1: Inter-se transfers dominate (>50% of disposals)
                # This means the shareholding "drop" is family restructuring, not selling
                if inter_se_pct > 50:
                    value_trap_mitigated = True
                    mitigation_reason = (
                        f"Shareholding drop is {inter_se_pct:.0f}% inter-se transfers "
                        f"(family restructuring), not genuine selling"
                    )
                
                # Check 2: Insider action is buying or inter-se only
                elif insider_action in ('ONLY_BUYING', 'NET_BUYER', 'INTER_SE_ONLY', 'NO_ACTIVITY'):
                    value_trap_mitigated = True
                    mitigation_reason = (
                        f"Insider trading shows {insider_action.replace('_', ' ').lower()}"
                        f"—no genuine open-market exit signal"
                    )
                
                # Check 3: Recent open-market buying by promoter
                elif recent_buy:
                    value_trap_mitigated = True
                    mitigation_reason = (
                        f"Promoter bought in open market in last 6 months"
                        f"—contradicts exit narrative"
                    )
                
                # Check 4: Market sell value is tiny relative to market cap (<3%)
                elif market_sell_pct_mcap is not None and market_sell_pct_mcap < 3.0:
                    value_trap_mitigated = True
                    mitigation_reason = (
                        f"Open-market selling is only {market_sell_pct_mcap:.1f}% of MCap"
                        f"—minor liquidity, not conviction exit"
                    )
            
            if value_trap_mitigated:
                # DON'T downgrade to VALUE_TRAP — keep original bucket
                # But add a context note about the promoter change
                result['Decision_Bucket'] = original_bucket  # Restore original
                result['Conviction_Override'] = (
                    f"Promoter holding {promoter_change:.1f}% BUT "
                    f"mitigated: {mitigation_reason}"
                )
                # Append insider context to reject reason if it exists
                existing_reason = result.get('Reject_Reason', '')
                if existing_reason and existing_reason != "None":
                    result['Reject_Reason'] = (
                        f"{existing_reason} | "
                        f"Promoter change {promoter_change:.1f}% mitigated by insider trading context"
                    )
            else:
                # Genuine VALUE_TRAP: either no insider trading data to check,
                # or insider trading confirms the selling signal
                result['Decision_Bucket'] = "VALUE_TRAP"
                
                # Build enhanced override message
                override_msg = f"Promoter selling {promoter_change:.1f}% despite high score"
                if it_available and insider_action == 'NET_SELLER':
                    override_msg += " (confirmed: net open-market seller)"
                elif not it_available:
                    override_msg += " (no insider trading data to verify)"
                
                result['Conviction_Override'] = override_msg
                result['Reject_Reason'] = (
                    f"Score={result['Composite_Score']} (was {original_bucket}) BUT "
                    f"promoter reduced stake by {abs(promoter_change):.1f}%—insider exit signal"
                )
    
    # ========================================
    # 17b. CONTEXT NOTES (Important caveats)
    # ========================================
    # Add notes that help analyst understand what to verify
    context_notes = []
    
    if consolidated_rejected:
        context_notes.append("Consolidated data existed but had < 2 years; used standalone for trend analysis")
    # ROE below 15% - verify competitive advantage
    if result['ROE_Latest'] is not None and result['ROE_Latest'] < 15:
        context_notes.append(f"ROE {result['ROE_Latest']:.1f}% below 15%—verify competitive advantage")
    
    # High other income - may be business model
    if "HIGH_OTHER_INCOME" in quality_flags:
        context_notes.append("High other income—verify if holding company/financial")
    
    # Sector adjustment needed
    if result['Sector_Adjustment_Needed'] == "Yes":
        context_notes.append(f"Sector context needed: {result['Sector_Warnings']}")
    
    # Standalone data only
    if result['Data_Type'] == 'standalone':
        context_notes.append("Standalone data only—subsidiary risks may be hidden")
    
    # Declining but still good ROE
    if result['ROE_Trend'] is not None and result['ROE_Trend'] < -2:
        if result['ROE_Latest'] is not None and result['ROE_Latest'] >= 15:
            context_notes.append(f"ROE declining but still {result['ROE_Latest']:.1f}%—may be mean reversion")
    
    result['Context_Notes'] = " | ".join(context_notes) if context_notes else "None"
    
    # ========================================
    # 17. PRIMARY CONCERN
    # ========================================
    # What's the main issue (if any)?
    if "NEGATIVE_CFO" in quality_flags or "INCONSISTENT_CFO" in quality_flags or "POOR_CASH_CONVERSION" in quality_flags:
        result['Primary_Concern'] = "Cash_Conversion"
    elif "RISING_DEBT" in quality_flags:
        result['Primary_Concern'] = "Leverage"
    elif "HIGH_PE" in pricing_flags or "HIGH_EV_EBITDA" in pricing_flags:
        result['Primary_Concern'] = "Valuation"
    elif "FREQUENT_EXCEPTIONALS" in quality_flags or "HIGH_OTHER_INCOME" in quality_flags:
        result['Primary_Concern'] = "Earnings_Quality"
    elif "DECLINING_ROE" in quality_flags or "DECLINING_ROCE" in quality_flags or "MARGIN_COMPRESSION" in quality_flags:
        result['Primary_Concern'] = "Deteriorating_Fundamentals"
    elif "LOW_ROE" in quality_flags or "LOW_ROCE" in quality_flags:
        result['Primary_Concern'] = "Weak_Returns"
    elif "RISING_RECEIVABLES" in quality_flags:
        result['Primary_Concern'] = "Working_Capital"
    elif "NEGATIVE_PE" in pricing_flags:
        result['Primary_Concern'] = "Unprofitable"
    elif result['Data_Confidence'] < 60:
        result['Primary_Concern'] = "Data_Quality"
    else:
        result['Primary_Concern'] = "None"
    
    # ========================================
    # 19. INVESTMENT THESIS (One-liner - NEUTRAL LANGUAGE)
    # ========================================
    if result['Decision_Bucket'] == "SCREEN_FAILED":
        result['Investment_Thesis'] = f"FAILED SCREEN: {result['Reject_Reason']}"
    elif result['Decision_Bucket'] == "SCREEN_PASSED_VERIFY":
        result['Investment_Thesis'] = f"PASSED—VERIFY GROUP: Standalone data only, check consolidated financials"
    elif result['Decision_Bucket'] == "SCREEN_PASSED_EXPENSIVE":
        result['Investment_Thesis'] = f"PASSED BUT EXPENSIVE: Verify if premium is justified ({pricing_flags[0] if pricing_flags else 'valuation'})"
    elif result['Decision_Bucket'] == "SCREEN_PASSED_FLAGS":
        result['Investment_Thesis'] = f"PASSED WITH FLAGS: Verify {result['Reject_Reason']}"
    elif result['Decision_Bucket'] == "DATA_INCOMPLETE":
        result['Investment_Thesis'] = f"VERIFY DATA: Insufficient data quality"
    elif result['Decision_Bucket'] == "SCREEN_MARGINAL":
        result['Investment_Thesis'] = f"BORDERLINE: Marginal metrics, needs deeper analysis"
    else:
        # Screen passed - highlight what metrics support it (but NOT a recommendation)
        notes = []
        if result['ROE_Latest'] and result['ROE_Latest'] >= 18:
            notes.append(f"ROE {result['ROE_Latest']:.0f}%")
        elif result['ROE_Latest'] and result['ROE_Latest'] < 15:
            notes.append(f"ROE only {result['ROE_Latest']:.0f}%—verify moat")
        if result['ROCE_Latest'] and result['ROCE_Latest'] >= 18:
            notes.append(f"ROCE {result['ROCE_Latest']:.0f}%")
        if result['CFO_PAT_3Yr_Avg'] and result['CFO_PAT_3Yr_Avg'] >= 1:
            notes.append("Strong cash conversion")
        if result['Earnings_Quality'] == "Clean":
            notes.append("Clean earnings")
        
        if notes:
            result['Investment_Thesis'] = f"PASSED SCREEN: {', '.join(notes[:3])}"
        else:
            result['Investment_Thesis'] = "PASSED SCREEN: Meets quality thresholds"
    
    # ========================================
    # 19. SHAREHOLDING DATA — Already loaded before SCREEN_ELIGIBLE (section 17b-PREREQ)
    # ========================================
    # Shareholding was loaded early so Pledge_Gate_Failed is available for SCREEN_ELIGIBLE.
    # No need to reload here.
    
    # ========================================
    # 20. GENERIC STOCK ANALYSIS (Neglected Firm Strategy)
    # ========================================
    # Based on the concept that neglected stocks with solid fundamentals
    # can deliver superior risk-adjusted returns when they "spring to life"
    
    # 20a. Institutional Holding (combined FII + MF)
    # Lower institutional ownership = more neglected
    if result.get('FII_Holding') is not None and result.get('MF_Holding') is not None:
        result['Institutional_Holding'] = round(result['FII_Holding'] + result['MF_Holding'], 2)
    elif result.get('FII_Holding') is not None:
        result['Institutional_Holding'] = result['FII_Holding']
    elif result.get('MF_Holding') is not None:
        result['Institutional_Holding'] = result['MF_Holding']
    else:
        result['Institutional_Holding'] = None
    
    # 20b. CFO to Debt Ratio (Bankruptcy Risk Filter)
    # Article recommends: Operating Cash Flow / Total Debt
    # Higher is better - captures ability to repay debt without external financing
    # Rule of thumb: > 0.2 is acceptable, > 0.4 is good
    total_debt = None
    if result.get('Debt_Equity') is not None and result.get('Debt_Equity') > 0:
        # Estimate total debt from D/E ratio and book value
        book_value = safe_float(safe_get(latest, 'BVSH_A'))
        shares = safe_float(safe_get(latest, 'EQCAP_A'))  # Equity capital
        if book_value and shares:
            # This is approximate - shareholders funds * D/E ratio
            shareholders_funds = safe_float(safe_get(latest, 'TotalShareHoldersFunds_A'))
            if shareholders_funds:
                total_debt = result['Debt_Equity'] * shareholders_funds
    
    # Also try direct borrowings
    if total_debt is None:
        borrowings = safe_float(safe_get(latest, 'Borrowings_A'))
        if borrowings and borrowings > 0:
            total_debt = borrowings
    
    if result.get('CFO_Latest') is not None and total_debt is not None and total_debt > 0:
        result['CFO_Debt_Ratio'] = round(result['CFO_Latest'] / total_debt, 2)
    else:
        result['CFO_Debt_Ratio'] = None
    
    # Store total debt for reference
    result['Total_Debt'] = round(total_debt, 2) if total_debt else None
    
    # 20c. Revenue Growth Consistency
    # Count years with positive revenue growth out of available years
    # Article recommends: "Consistent growth in revenues during the last 5 years"
    rev_growth_years = []
    for i, y in enumerate(available_years[:-1]):  # Compare each year to previous
        curr_rev = safe_float(safe_get(yearly_data.get(y, {}), 'SR_A'))
        prev_rev = safe_float(safe_get(yearly_data.get(available_years[i+1], {}), 'SR_A'))
        if curr_rev is not None and prev_rev is not None and prev_rev > 0:
            if curr_rev > prev_rev:
                rev_growth_years.append(1)
            else:
                rev_growth_years.append(0)
    
    if rev_growth_years:
        result['Revenue_Growth_Years'] = sum(rev_growth_years)
        result['Revenue_Growth_Consistency'] = round((sum(rev_growth_years) / len(rev_growth_years)) * 100, 1)
    else:
        result['Revenue_Growth_Years'] = None
        result['Revenue_Growth_Consistency'] = None
    
    # 20d. Neglect Score
    # Based on: institutional ownership + market cap (as proxy for coverage)
    # LOW neglect = well-covered, HIGH neglect = potential generic stock
    neglect_score = 0
    neglect_reasons = []
    
    # Factor 1: Institutional ownership < 10% (article's threshold)
    if result.get('Institutional_Holding') is not None:
        if result['Institutional_Holding'] < 5:
            neglect_score += 3
            neglect_reasons.append(f"Very low inst. ownership ({result['Institutional_Holding']:.1f}%)")
        elif result['Institutional_Holding'] < 10:
            neglect_score += 2
            neglect_reasons.append(f"Low inst. ownership ({result['Institutional_Holding']:.1f}%)")
        elif result['Institutional_Holding'] < 20:
            neglect_score += 1
            neglect_reasons.append(f"Moderate inst. ownership ({result['Institutional_Holding']:.1f}%)")
    
    # Factor 2: Small market cap (proxy for less analyst coverage)
    if result.get('Market_Cap') is not None:
        if result['Market_Cap'] < 500:  # < 500 Cr = micro cap
            neglect_score += 2
            neglect_reasons.append("Micro cap (<500 Cr)")
        elif result['Market_Cap'] < 2000:  # < 2000 Cr = small cap
            neglect_score += 1
            neglect_reasons.append("Small cap (<2000 Cr)")
    
    # Factor 3: No/low promoter holding can indicate less "story" appeal
    if result.get('Promoter_Holding') is not None:
        if result['Promoter_Holding'] < 25:
            neglect_score += 1
            neglect_reasons.append(f"Low promoter ({result['Promoter_Holding']:.1f}%)")
    
    # Factor 4: AGE DIMENSION (Fix 3.5)
    # Low institutional ownership means DIFFERENT things for old vs new companies:
    # - Old (10+ years data) + low institutional = LIKELY INSPECTED AND REJECTED
    # - New (< 5 years data) + low institutional = POTENTIALLY UNDISCOVERED
    #
    # Years_Available is a proxy for listing age (Trendlyne data goes back to listing)
    years_listed = result.get('Years_Available', 0)
    is_low_institutional = result.get('Institutional_Holding') is not None and result['Institutional_Holding'] < 10
    
    if is_low_institutional and years_listed is not None:
        if years_listed >= 10:
            # OLD company with low institutional = probably rejected by smart money
            # This is a WARNING, not a bonus
            neglect_score -= 1  # Reduce score (less attractive)
            neglect_reasons.append(f"⚠️ Old ({years_listed}yr) + low inst. = likely rejected by institutions")
            result['Neglect_Age_Warning'] = "LIKELY_REJECTED"
        elif years_listed < 5:
            # NEW company with low institutional = potentially undiscovered
            neglect_score += 1  # Bonus
            neglect_reasons.append(f"✓ New ({years_listed}yr) + low inst. = potentially undiscovered")
            result['Neglect_Age_Warning'] = "POTENTIALLY_UNDISCOVERED"
        else:
            # Middle-aged (5-10 years) - neutral
            result['Neglect_Age_Warning'] = "MODERATE_HISTORY"
    else:
        result['Neglect_Age_Warning'] = None
    
    # Assign neglect category
    if neglect_score >= 4:
        result['Neglect_Score'] = "HIGH"
    elif neglect_score >= 2:
        result['Neglect_Score'] = "MEDIUM"
    else:
        result['Neglect_Score'] = "LOW"
    
    result['Neglect_Reasons'] = "; ".join(neglect_reasons) if neglect_reasons else "Well-covered stock"
    
    # 20e. Generic Stock Candidate Flag
    # A stock qualifies if it meets ALL criteria from the article:
    # 1. Shows signs of neglect (institutional ownership < 10% OR small cap)
    # 2. Debt/Equity < 0.5 (financial leverage < 50%)
    # 3. ROE 5Yr Avg > 18% (quality threshold from article)
    # 4. Revenue growth consistent (≥ 60% of years positive)
    # 5. CFO/Debt > 0.2 OR debt-free (not bankruptcy risk)
    # 6. No CRITICAL red flags (not a "deserved lemon")
    
    generic_criteria = []
    generic_pass = True
    
    # Criterion 1: Neglect indicator
    is_neglected = (
        (result.get('Institutional_Holding') is not None and result['Institutional_Holding'] < 10) or
        (result.get('Market_Cap') is not None and result['Market_Cap'] < 2000)
    )
    if is_neglected:
        generic_criteria.append("✓ Neglected (low inst. or small cap)")
    else:
        generic_criteria.append("✗ Well-covered")
        generic_pass = False
    
    # Criterion 2: Low leverage
    if result.get('Debt_Equity') is not None:
        if result['Debt_Equity'] < 0.5:
            generic_criteria.append(f"✓ Low leverage (D/E={result['Debt_Equity']:.2f})")
        else:
            generic_criteria.append(f"✗ High leverage (D/E={result['Debt_Equity']:.2f})")
            generic_pass = False
    else:
        generic_criteria.append("? Leverage unknown")
    
    # Criterion 3: Quality - ROE > 18%
    if result.get('ROE_5Yr_Avg') is not None:
        if result['ROE_5Yr_Avg'] > 18:
            generic_criteria.append(f"✓ Quality ROE ({result['ROE_5Yr_Avg']:.1f}%)")
        else:
            generic_criteria.append(f"✗ Low ROE ({result['ROE_5Yr_Avg']:.1f}%)")
            generic_pass = False
    elif result.get('ROE_3Yr_Avg') is not None:
        if result['ROE_3Yr_Avg'] > 18:
            generic_criteria.append(f"✓ Quality ROE 3Y ({result['ROE_3Yr_Avg']:.1f}%)")
        else:
            generic_criteria.append(f"✗ Low ROE 3Y ({result['ROE_3Yr_Avg']:.1f}%)")
            generic_pass = False
    else:
        generic_criteria.append("? ROE unknown")
        generic_pass = False
    
    # Criterion 4: Revenue consistency (≥60% years positive)
    if result.get('Revenue_Growth_Consistency') is not None:
        if result['Revenue_Growth_Consistency'] >= 60:
            generic_criteria.append(f"✓ Consistent revenue ({result['Revenue_Growth_Consistency']:.0f}%)")
        else:
            generic_criteria.append(f"✗ Inconsistent revenue ({result['Revenue_Growth_Consistency']:.0f}%)")
            generic_pass = False
    else:
        generic_criteria.append("? Revenue consistency unknown")
    
    # Criterion 5: Not bankruptcy risk
    if result.get('CFO_Debt_Ratio') is not None:
        if result['CFO_Debt_Ratio'] > 0.2:
            generic_criteria.append(f"✓ Solvent (CFO/Debt={result['CFO_Debt_Ratio']:.2f})")
        else:
            generic_criteria.append(f"✗ Bankruptcy risk (CFO/Debt={result['CFO_Debt_Ratio']:.2f})")
            generic_pass = False
    elif result.get('Debt_Equity') is not None and result['Debt_Equity'] < 0.1:
        generic_criteria.append("✓ Nearly debt-free")
    else:
        generic_criteria.append("? Solvency unknown")
    
    # Criterion 6: No critical red flags (not a "deserved lemon")
    # FIX: Critical_Flags is set to "NONE" (all-caps), not "None"
    if result.get('Critical_Flags') and result['Critical_Flags'] != "NONE":
        generic_criteria.append(f"✗ Critical flags: {result['Critical_Flags']}")
        generic_pass = False
    else:
        generic_criteria.append("✓ No critical flags")
    
    result['Generic_Stock_Candidate'] = "YES" if generic_pass else "NO"
    result['Generic_Criteria_Detail'] = " | ".join(generic_criteria)
    
    # ========================================
    # 21. CORPORATE ACTIONS DATA (from companion JSON)
    # ========================================
    corp_actions = load_corporate_actions_data(filepath, corporate_actions_dir, resolved_path=_resolved_ca)
    if corp_actions:
        # Dividend data
        result['Dividend_Count'] = corp_actions.get('Dividend_Count')
        result['Latest_Dividend_Date'] = corp_actions.get('Latest_Dividend_Date')
        result['Latest_Dividend_Amount'] = corp_actions.get('Latest_Dividend_Amount')
        result['Latest_Dividend_Type'] = corp_actions.get('Latest_Dividend_Type')
        result['Years_Since_Dividend'] = corp_actions.get('Years_Since_Dividend')
        result['Dividend_5Yr_Consistency'] = corp_actions.get('Dividend_5Yr_Consistency')
        result['Dividend_Trend'] = corp_actions.get('Dividend_Trend')
        result['Total_Dividends_Paid'] = corp_actions.get('Total_Dividends_Paid')
        
        # Bonus data
        result['Bonus_Count'] = corp_actions.get('Bonus_Count')
        result['Latest_Bonus_Date'] = corp_actions.get('Latest_Bonus_Date')
        result['Latest_Bonus_Ratio'] = corp_actions.get('Latest_Bonus_Ratio')
        
        # Split data
        result['Split_Count'] = corp_actions.get('Split_Count')
        result['Latest_Split_Date'] = corp_actions.get('Latest_Split_Date')
        result['Latest_Split_Ratio'] = corp_actions.get('Latest_Split_Ratio')
        
        # Rights issue data
        result['Rights_Count'] = corp_actions.get('Rights_Count')
        result['Latest_Rights_Date'] = corp_actions.get('Latest_Rights_Date')
        result['Latest_Rights_Ratio'] = corp_actions.get('Latest_Rights_Ratio')
        result['Latest_Rights_Premium'] = corp_actions.get('Latest_Rights_Premium')
        
        # Board meetings
        result['Next_Board_Meeting'] = corp_actions.get('Next_Board_Meeting')
        result['Next_Meeting_Purpose'] = corp_actions.get('Next_Meeting_Purpose')
        
        # Analyst coverage
        result['Analyst_Reports_Count'] = corp_actions.get('Analyst_Reports_Count')
        result['Latest_Target_Price'] = corp_actions.get('Latest_Target_Price')
        result['Latest_Reco_Type'] = corp_actions.get('Latest_Reco_Type')
        result['Latest_Upside'] = corp_actions.get('Latest_Upside')
        
        # Capital allocation assessment
        result['Capital_Allocation_Score'] = corp_actions.get('Capital_Allocation_Score')
        result['Capital_Allocation_Notes'] = corp_actions.get('Capital_Allocation_Notes')
        
        # Company info
        result['Executive_Count'] = corp_actions.get('Executive_Count')
        result['Director_Count'] = corp_actions.get('Director_Count')
        result['Website'] = corp_actions.get('Website')
        
        result['Corporate_Actions_Available'] = "Yes"
    else:
        # Set defaults when corporate actions data not available
        result['Dividend_Count'] = None
        result['Latest_Dividend_Date'] = None
        result['Latest_Dividend_Amount'] = None
        result['Latest_Dividend_Type'] = None
        result['Years_Since_Dividend'] = None
        result['Dividend_5Yr_Consistency'] = None
        result['Dividend_Trend'] = None
        result['Total_Dividends_Paid'] = None
        result['Bonus_Count'] = None
        result['Latest_Bonus_Date'] = None
        result['Latest_Bonus_Ratio'] = None
        result['Split_Count'] = None
        result['Latest_Split_Date'] = None
        result['Latest_Split_Ratio'] = None
        result['Rights_Count'] = None
        result['Latest_Rights_Date'] = None
        result['Latest_Rights_Ratio'] = None
        result['Latest_Rights_Premium'] = None
        result['Next_Board_Meeting'] = None
        result['Next_Meeting_Purpose'] = None
        result['Analyst_Reports_Count'] = None
        result['Latest_Target_Price'] = None
        result['Latest_Reco_Type'] = None
        result['Latest_Upside'] = None
        result['Capital_Allocation_Score'] = None
        result['Capital_Allocation_Notes'] = None
        result['Executive_Count'] = None
        result['Director_Count'] = None
        result['Website'] = None
        result['Corporate_Actions_Available'] = "No"
    
    # ========================================
    # 22. OVERVIEW DATA - REMAINING FIELDS
    # (Overview was loaded early for PE_TTM, PEG_TTM, Market_Cap - now store other fields)
    # ========================================
    if overview_available:
        # Piotroski Score (key quality metric) - already extracted as piotroski_score
        result['Piotroski_Score'] = piotroski_score
        result['Piotroski_Assessment'] = overview.get('Piotroski_Assessment')
        
        # TTM Growth metrics
        result['Revenue_Growth_TTM'] = overview.get('Revenue_Growth_TTM')
        result['NP_Growth_TTM'] = overview.get('NP_Growth_TTM')
        result['Revenue_Growth_Qtr_YoY'] = overview.get('Revenue_Growth_Qtr_YoY')
        result['NP_Growth_Qtr_YoY'] = overview.get('NP_Growth_Qtr_YoY')
        
        # TTM Profitability - already extracted as opm_ttm
        result['OPM_TTM'] = opm_ttm
        
        # 52-Week High/Low
        result['Week52_High'] = overview.get('Week52_High')
        result['Week52_Low'] = overview.get('Week52_Low')
        result['LTP'] = overview.get('LTP')
        result['Return_1Yr'] = overview.get('Return_1Yr')
        result['Price_Position_52W'] = overview.get('Price_Position_52W')
        result['Distance_From_52W_High'] = overview.get('Distance_From_52W_High')
        
        # Relative performance
        result['Returns_vs_Nifty50_Qtr'] = overview.get('Returns_vs_Nifty50_Qtr')
        result['Returns_vs_Sector_Qtr'] = overview.get('Returns_vs_Sector_Qtr')
        result['Return_vs_Nifty_1Yr'] = overview.get('Return_vs_Nifty_1Yr')
        
        # Institutional holding from overview (may be more current)
        if overview.get('Institutional_Holding_Overview') is not None:
            result['Institutional_Holding_Overview'] = overview.get('Institutional_Holding_Overview')
        
        result['Overview_Available'] = "Yes"
    else:
        # Set defaults when overview data not available
        result['Piotroski_Score'] = None
        result['Piotroski_Assessment'] = None
        result['Revenue_Growth_TTM'] = None
        result['NP_Growth_TTM'] = None
        result['Revenue_Growth_Qtr_YoY'] = None
        result['NP_Growth_Qtr_YoY'] = None
        result['OPM_TTM'] = None
        result['Week52_High'] = None
        result['Week52_Low'] = None
        result['LTP'] = None
        result['Return_1Yr'] = None
        result['Price_Position_52W'] = None
        result['Distance_From_52W_High'] = None
        result['Returns_vs_Nifty50_Qtr'] = None
        result['Returns_vs_Sector_Qtr'] = None
        result['Return_vs_Nifty_1Yr'] = None
        result['Institutional_Holding_Overview'] = None
        result['Overview_Available'] = "No"
    
    return result


def apply_sector_relative_scoring(results: List[Dict]) -> List[Dict]:
    """
    Post-processing: adjust scores based on sector-relative position.
    
    Problem: Absolute thresholds (ROE >= 20 → +15) are sector-blind.
    A 15% ROE is mediocre for FMCG (Nestle ~80%) but excellent for Infra (L&T ~15%).
    A PE of 15 may be cyclical peak for Steel but a bargain for Consumer.
    
    Fix: Compute sector medians, then adjust scores based on how each stock
    compares to its sector peers. This preserves the absolute scoring as a
    baseline but adds a ±15 point sector-relative overlay.
    
    Requires all stocks to be analyzed first (can't do per-stock).
    """
    from collections import defaultdict
    
    # Metrics to compute sector medians for
    SECTOR_METRICS = ['ROE_Latest', 'ROCE_Latest', 'OPM_Latest', 'PE', 'Debt_Equity', 'Revenue_Growth_1Yr']
    MIN_SECTOR_SIZE = 10  # Need enough stocks for a meaningful median
    
    # ── Step 1: Group by sector and compute medians ──
    sector_data = defaultdict(lambda: defaultdict(list))
    for r in results:
        sector = r.get('Sector', 'Unknown')
        if not sector or sector == 'Unknown':
            continue
        for metric in SECTOR_METRICS:
            val = r.get(metric)
            if val is not None and isinstance(val, (int, float)):
                sector_data[sector][metric].append(val)
    
    # Compute percentile thresholds per sector
    sector_stats = {}
    for sector, metrics in sector_data.items():
        sector_stats[sector] = {}
        for metric, values in metrics.items():
            if len(values) >= MIN_SECTOR_SIZE:
                sorted_vals = sorted(values)
                n = len(sorted_vals)
                sector_stats[sector][metric] = {
                    'median': sorted_vals[n // 2],
                    'p25': sorted_vals[n // 4],
                    'p75': sorted_vals[3 * n // 4],
                    'count': n,
                }
    
    # ── Step 2: Adjust each stock's scores ──
    adjusted_count = 0
    for r in results:
        sector = r.get('Sector', 'Unknown')
        if sector not in sector_stats:
            r['Sector_Relative_Adj'] = "N/A (sector too small or unknown)"
            continue
        
        s_stats = sector_stats[sector]
        adj_details = []
        bq_adj = 0  # Business Quality adjustment
        vc_adj = 0  # Valuation Comfort adjustment
        
        # ROE relative to sector
        if 'ROE_Latest' in s_stats:
            roe = r.get('ROE_Latest')
            s = s_stats['ROE_Latest']
            if roe is not None:
                if roe >= s['p75']:
                    bq_adj += 8
                    adj_details.append(f"ROE top-quartile in {sector} (+8)")
                elif roe <= s['p25']:
                    bq_adj -= 8
                    adj_details.append(f"ROE bottom-quartile in {sector} (-8)")
        
        # ROCE relative to sector
        if 'ROCE_Latest' in s_stats:
            roce = r.get('ROCE_Latest')
            s = s_stats['ROCE_Latest']
            if roce is not None:
                if roce >= s['p75']:
                    bq_adj += 7
                    adj_details.append(f"ROCE top-quartile (+7)")
                elif roce <= s['p25']:
                    bq_adj -= 7
                    adj_details.append(f"ROCE bottom-quartile (-7)")
        
        # OPM relative to sector (caps at ±5)
        if 'OPM_Latest' in s_stats:
            opm = r.get('OPM_Latest')
            s = s_stats['OPM_Latest']
            if opm is not None:
                if opm >= s['p75']:
                    bq_adj += 5
                elif opm <= s['p25']:
                    bq_adj -= 5
        
        # PE relative to sector (INVERTED: lower PE = better value)
        if 'PE' in s_stats:
            pe = r.get('PE')
            s = s_stats['PE']
            if pe is not None and pe > 0:  # Only positive PE is meaningful
                if pe <= s['p25']:
                    vc_adj += 10
                    adj_details.append(f"PE below sector P25 ({pe:.1f} vs {s['p25']:.1f}) (+10)")
                elif pe >= s['p75']:
                    vc_adj -= 10
                    adj_details.append(f"PE above sector P75 ({pe:.1f} vs {s['p75']:.1f}) (-10)")
                # Key insight: stock PE vs 0.8 × sector median
                if pe < s['median'] * 0.8:
                    vc_adj += 5
                    adj_details.append(f"PE < 80% of sector median (+5)")
        
        # D/E relative to sector
        if 'Debt_Equity' in s_stats:
            de = r.get('Debt_Equity')
            s = s_stats['Debt_Equity']
            if de is not None and de >= 0:
                if de >= s['p75']:
                    bq_adj -= 5
                    adj_details.append(f"D/E top-quartile in sector (-5)")
        
        # Apply adjustments (capped at ±15 per sub-score)
        bq_adj = max(-15, min(15, bq_adj))
        vc_adj = max(-15, min(15, vc_adj))
        
        if bq_adj != 0 or vc_adj != 0:
            # Adjust sub-scores
            old_bq = r.get('Business_Quality_Score', 50)
            old_vc = r.get('Valuation_Comfort_Score', 50)
            
            r['Business_Quality_Score'] = max(0, min(100, old_bq + bq_adj))
            r['Valuation_Comfort_Score'] = max(0, min(100, old_vc + vc_adj))
            
            # Recompute composite with same weights
            composite = (
                r['Business_Quality_Score'] * 0.35 +
                r['Financial_Strength_Score'] * 0.25 +
                r['Growth_Durability_Score'] * 0.20 +
                r['Valuation_Comfort_Score'] * 0.20
            )
            
            # Re-apply structural penalty (flag count × 10 + extras)
            quality_flag_count = r.get('Quality_Flag_Count', 0)
            composite -= quality_flag_count * 10
            if quality_flag_count >= 2:
                composite -= 15
            if quality_flag_count >= 3:
                composite -= 15
            
            # Re-apply caps
            if r.get('Data_Confidence', 100) < 40:
                composite = min(composite, 40)
            elif r.get('Data_Confidence', 100) < 60:
                composite = min(composite, 60)
            if quality_flag_count >= 2:
                composite = min(composite, 50)
            
            old_composite = r.get('Composite_Score', 0)
            r['Composite_Score'] = max(0, min(100, round(composite)))
            
            adjusted_count += 1
            r['Sector_Relative_Adj'] = f"BQ {bq_adj:+d}, VC {vc_adj:+d} → Composite {old_composite}→{r['Composite_Score']}"
            if adj_details:
                r['Sector_Relative_Adj'] += f" [{'; '.join(adj_details)}]"
        else:
            r['Sector_Relative_Adj'] = "No adjustment (mid-sector on all metrics)"
    
    print(f"[INFO] Sector-relative scoring applied to {adjusted_count} stocks across {len(sector_stats)} sectors")
    return results


def compute_summary_stats(results: List[Dict]) -> Dict[str, List]:
    """
    Single-pass computation of all summary statistics from results.
    Eliminates duplicate list comprehensions in Excel summary and console output.
    
    At 4,500 stocks, the old code did 7+2+3+3 = 15 separate list scans (×2 for duplication = 30).
    This does 1 pass, bucketing each stock into all relevant categories.
    """
    stats = {
        'screen_passed': [], 'screen_verify': [], 'screen_expensive': [],
        'screen_flags': [], 'screen_marginal': [], 'data_incomplete': [],
        'screen_failed': [],
        'contrarian_bet': [], 'value_trap': [],
        'eligible': [], 'blocked': [],
        'low_risk': [], 'medium_risk': [], 'high_risk': [],
        'generic_candidates': [], 'high_neglect': [], 'medium_neglect': [],
        # Insider trading & companion data coverage
        'has_shareholding': [], 'has_overview': [], 'has_insider_trading': [],
        'has_dividends': [],
        # Insider trading action breakdown
        'it_only_buying': [], 'it_net_buyer': [], 'it_inter_se_only': [],
        'it_net_seller': [], 'it_net_seller_with_buys': [], 'it_no_activity': [],
        # Conviction override tracking
        'value_trap_mitigated': [],  # Would have been VALUE_TRAP but insider trading saved it
        'contrarian_confirmed': [],  # CONTRARIAN_BET confirmed by insider trading
    }
    
    bucket_map = {
        'GATES_CLEARED': 'screen_passed', 'SCREEN_PASSED_VERIFY': 'screen_verify',
        'SCREEN_PASSED_EXPENSIVE': 'screen_expensive', 'SCREEN_PASSED_FLAGS': 'screen_flags',
        'SCREEN_MARGINAL': 'screen_marginal', 'DATA_INCOMPLETE': 'data_incomplete',
        'SCREEN_FAILED': 'screen_failed',
        'CONTRARIAN_BET': 'contrarian_bet', 'VALUE_TRAP': 'value_trap',
    }
    risk_map = {'LOW': 'low_risk', 'MEDIUM': 'medium_risk', 'HIGH': 'high_risk'}
    neglect_map = {'HIGH': 'high_neglect', 'MEDIUM': 'medium_neglect'}
    
    for s in results:
        # Decision bucket
        bucket = s.get('Decision_Bucket', '')
        if bucket in bucket_map:
            stats[bucket_map[bucket]].append(s)
        
        # Eligibility
        elig = s.get('SCREEN_ELIGIBLE', '')
        if elig == 'YES':
            stats['eligible'].append(s)
        elif elig == 'NO':
            stats['blocked'].append(s)
        
        # Risk
        risk = s.get('Quality_Risk', '')
        if risk in risk_map:
            stats[risk_map[risk]].append(s)
        
        # Generic candidate
        if s.get('Generic_Stock_Candidate') == 'YES':
            stats['generic_candidates'].append(s)
        
        # Neglect score
        neglect = s.get('Neglect_Score', '')
        if neglect in neglect_map:
            stats[neglect_map[neglect]].append(s)
        
        # Companion data coverage
        if s.get('Shareholding_Available') == 'Yes':
            stats['has_shareholding'].append(s)
        if s.get('Insider_Trading_Available') == 'Yes':
            stats['has_insider_trading'].append(s)
        if s.get('Overview_Available') == 'Yes':
            stats['has_overview'].append(s)
        if s.get('Dividend_Count', 0) is not None and s.get('Dividend_Count', 0) > 0:
            stats['has_dividends'].append(s)
        
        # Insider trading action breakdown
        it_action = s.get('Insider_Action')
        it_action_map = {
            'ONLY_BUYING': 'it_only_buying', 'NET_BUYER': 'it_net_buyer',
            'INTER_SE_ONLY': 'it_inter_se_only', 'NO_ACTIVITY': 'it_no_activity',
            'NET_SELLER': 'it_net_seller', 'NET_SELLER_WITH_BUYS': 'it_net_seller_with_buys',
        }
        if it_action in it_action_map:
            stats[it_action_map[it_action]].append(s)
        
        # Conviction override tracking
        override = s.get('Conviction_Override', 'None')
        if override != 'None' and 'mitigated' in str(override).lower():
            stats['value_trap_mitigated'].append(s)
        if override != 'None' and 'confirmed by open-market purchases' in str(override).lower():
            stats['contrarian_confirmed'].append(s)
    
    return stats


def _analyze_stock_worker(args):
    """
    Module-level wrapper for parallel stock processing.
    Must be at module level (not nested) for ProcessPoolExecutor pickle serialization.
    Accepts a single tuple of (filepath, shareholding_dir, corporate_actions_dir, overview_dir, insider_trading_dir, file_index).
    """
    filepath, sh_dir, ca_dir, ov_dir, it_dir, file_idx = args
    result = analyze_stock(str(filepath), sh_dir, ca_dir, ov_dir, it_dir, file_index=file_idx)
    if result:
        return round_dict_values(result, decimals=2)
    return None


def main():
    """Main driver - process all JSON files and output CSV."""
    global DATA_DIR, SHAREHOLDING_DIR, CORPORATE_ACTIONS_DIR, OVERVIEW_DIR, INSIDER_TRADING_DIR, OUTPUT_FILE
    
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='Analyze stock data from Trendlyne JSON files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python Create_stock_analysis.py
      Process JSONs in current directory, shareholding in same directory
      
  python Create_stock_analysis.py --data-dir ./financials
      Process JSONs in ./financials, shareholding in same directory
      
  python Create_stock_analysis.py --data-dir ./financials --shareholding-dir ./shareholding
      Financial JSONs in ./financials, shareholding JSONs in ./shareholding
      
  python Create_stock_analysis.py --data-dir ./financials --overview-dir ./overview
      Financial JSONs in ./financials, overview (TTM, Piotroski) in ./overview
      
  python Create_stock_analysis.py --data-dir ./financials --insider-trading-dir ./insider_trading
      Financial JSONs in ./financials, insider trading data in ./insider_trading
      
  python Create_stock_analysis.py --output ./results/analysis.xlsx
      Specify output file location
        """
    )
    parser.add_argument('--data-dir', '-d', default=DATA_DIR,
                        help='Directory containing financial JSON files (default: current directory)')
    parser.add_argument('--shareholding-dir', '-s', default=SHAREHOLDING_DIR,
                        help='Directory containing shareholding JSON files (default: ./Shareholding)')
    parser.add_argument('--corporate-actions-dir', '-c', default=CORPORATE_ACTIONS_DIR,
                        help='Directory containing corporate actions JSON files (default: ./CorporateActions)')
    parser.add_argument('--overview-dir', '-v', default=OVERVIEW_DIR,
                        help='Directory containing overview JSON files (default: ./Overview)')
    parser.add_argument('--insider-trading-dir', '-i', default=INSIDER_TRADING_DIR,
                        help='Directory containing insider trading JSON files (default: ./InsiderTrading)')
    parser.add_argument('--output', '-o', default=None,
                        help='Output Excel file path (default: stock_analysis.xlsx in data directory)')
    parser.add_argument('--workers', '-w', type=int, default=None,
                        help='Number of parallel workers (default: CPU count; use 1 for sequential)')
    
    args = parser.parse_args()
    
    # Update global configuration from arguments
    DATA_DIR = args.data_dir
    SHAREHOLDING_DIR = args.shareholding_dir
    CORPORATE_ACTIONS_DIR = args.corporate_actions_dir
    OVERVIEW_DIR = args.overview_dir
    INSIDER_TRADING_DIR = args.insider_trading_dir
    
    if args.output:
        OUTPUT_FILE = args.output
    else:
        OUTPUT_FILE = os.path.join(DATA_DIR, "stock_analysis.xlsx")
    
    # Check if data directory exists
    if not os.path.exists(DATA_DIR):
        print(f"[ERROR] Data directory not found: {DATA_DIR}")
        print("Please ensure JSON files are in the specified directory.")
        sys.exit(1)
    
    # Check shareholding directory if specified
    if SHAREHOLDING_DIR and not os.path.exists(SHAREHOLDING_DIR):
        print(f"[WARN] Shareholding directory not found: {SHAREHOLDING_DIR}")
        print("       Continuing without shareholding data...")
    
    # Check corporate actions directory if specified
    if CORPORATE_ACTIONS_DIR and not os.path.exists(CORPORATE_ACTIONS_DIR):
        print(f"[WARN] Corporate actions directory not found: {CORPORATE_ACTIONS_DIR}")
        print("       Continuing without corporate actions data...")
    
    # Check overview directory if specified
    if OVERVIEW_DIR and not os.path.exists(OVERVIEW_DIR):
        print(f"[WARN] Overview directory not found: {OVERVIEW_DIR}")
        print("       Continuing without overview data (TTM metrics, Piotroski)...")
    
    # Check insider trading directory if specified
    if INSIDER_TRADING_DIR and not os.path.exists(INSIDER_TRADING_DIR):
        print(f"[WARN] Insider trading directory not found: {INSIDER_TRADING_DIR}")
        print("       Continuing without insider trading data (conviction context)...")
    
    # Get all JSON files (exclude shareholding, dividend, and overview files which have different structure)
    all_json_files = list(Path(DATA_DIR).glob("*.json"))
    json_files = [f for f in all_json_files if '_shareholding' not in f.name and '_dividend' not in f.name and '_overview' not in f.name and '_insider_trading' not in f.name]
    #json_files = list(Path(DATA_DIR).glob("CHEMOPH.json"))
    
    if not json_files:
        print(f"[ERROR] No JSON files found in {DATA_DIR}")
        sys.exit(1)
    
    # =====================================================
    # PERFORMANCE: Pre-index companion files ONCE at startup
    # =====================================================
    # Replaces ~3 filesystem stat() calls per stock per companion type.
    # At 4,500 stocks: ~13,500 stat calls → 3 directory scans.
    t_index_start = time.time()
    file_index = build_companion_file_index(DATA_DIR, SHAREHOLDING_DIR, CORPORATE_ACTIONS_DIR, OVERVIEW_DIR, INSIDER_TRADING_DIR)
    
    # Count companion files from the index (no redundant glob needed)
    shareholding_count = sum(1 for v in file_index.values() if 'shareholding' in v)
    corp_actions_count = sum(1 for v in file_index.values() if 'dividend' in v)
    overview_count = sum(1 for v in file_index.values() if 'overview' in v)
    insider_trading_count = sum(1 for v in file_index.values() if 'insider_trading' in v)
    
    print(f"[INFO] Financial JSONs directory: {DATA_DIR}")
    if SHAREHOLDING_DIR and os.path.exists(SHAREHOLDING_DIR):
        print(f"[INFO] Shareholding JSONs directory: {SHAREHOLDING_DIR}")
    if CORPORATE_ACTIONS_DIR and os.path.exists(CORPORATE_ACTIONS_DIR):
        print(f"[INFO] Corporate Actions JSONs directory: {CORPORATE_ACTIONS_DIR}")
    if OVERVIEW_DIR and os.path.exists(OVERVIEW_DIR):
        print(f"[INFO] Overview JSONs directory: {OVERVIEW_DIR}")
    if INSIDER_TRADING_DIR and os.path.exists(INSIDER_TRADING_DIR):
        print(f"[INFO] Insider Trading JSONs directory: {INSIDER_TRADING_DIR}")
    print(f"[INFO] Found {len(json_files)} financial JSON files to process")
    if shareholding_count > 0:
        print(f"       ({shareholding_count} shareholding files indexed)")
    if corp_actions_count > 0:
        print(f"       ({corp_actions_count} corporate actions files indexed)")
    if overview_count > 0:
        print(f"       ({overview_count} overview files indexed)")
    if insider_trading_count > 0:
        print(f"       ({insider_trading_count} insider trading files indexed)")
    print(f"[INFO] File index built in {time.time() - t_index_start:.2f}s")
    
    # =====================================================
    # PERFORMANCE: Parallel stock processing
    # =====================================================
    # Each analyze_stock() is independent — no shared mutable state.
    # ProcessPoolExecutor gives near-linear speedup on multi-core machines.
    num_workers = args.workers if args.workers else min(os.cpu_count() or 1, len(json_files))
    
    t_process_start = time.time()
    results = []
    
    if num_workers == 1 or len(json_files) <= 10:
        # Sequential mode — cleaner output for small batches or debugging
        for filepath in json_files:
            result = analyze_stock(str(filepath), SHAREHOLDING_DIR, CORPORATE_ACTIONS_DIR, OVERVIEW_DIR,
                                   INSIDER_TRADING_DIR, file_index=file_index)
            if result:
                result = round_dict_values(result, decimals=2)
                results.append(result)
    else:
        # Parallel mode using module-level worker (required for pickle serialization)
        print(f"[INFO] Processing {len(json_files)} stocks with {num_workers} parallel workers...")
        
        worker_args = [(fp, SHAREHOLDING_DIR, CORPORATE_ACTIONS_DIR, OVERVIEW_DIR,
                        INSIDER_TRADING_DIR, file_index)
                       for fp in json_files]
        
        with ProcessPoolExecutor(max_workers=num_workers) as pool:
            chunksize = max(1, len(json_files) // (num_workers * 4))
            for result in pool.map(_analyze_stock_worker, worker_args, chunksize=chunksize):
                if result:
                    results.append(result)
    
    t_process_elapsed = time.time() - t_process_start
    stocks_per_sec = len(results) / t_process_elapsed if t_process_elapsed > 0 else 0
    print(f"[INFO] Processed {len(results)} stocks in {t_process_elapsed:.1f}s ({stocks_per_sec:.0f} stocks/sec)")
    
    if not results:
        print("[ERROR] No valid stock data extracted")
        sys.exit(1)
    
    # =====================================================
    # DEDUPLICATION: Ensure no duplicate stocks in output
    # =====================================================
    # Two dedup passes:
    # 1. By NSE_Code (filename-based) — catches same file under different names
    # 2. By ISIN (data-based) — catches same company under different tickers/filenames
    # Keep the one with better data quality in each case.
    
    def pick_better(existing, new):
        """Return the stock record with better data quality."""
        if (new.get('Data_Confidence', 0) > existing.get('Data_Confidence', 0) or
            (new.get('Data_Confidence', 0) == existing.get('Data_Confidence', 0) and
             new.get('Years_Available', 0) > existing.get('Years_Available', 0))):
            return new
        return existing
    
    # Pass 1: Dedup by NSE_Code
    seen_codes = {}
    for r in results:
        code = r['NSE_Code']
        if code in seen_codes:
            seen_codes[code] = pick_better(seen_codes[code], r)
        else:
            seen_codes[code] = r
    
    dedup_nse = len(results) - len(seen_codes)
    results = list(seen_codes.values())
    
    # Pass 2: Dedup by ISIN (if available and non-empty)
    seen_isin = {}
    deduped = []
    for r in results:
        isin = r.get('ISIN')
        if isin and isin.strip():
            if isin in seen_isin:
                # Keep better one, update in-place
                idx = seen_isin[isin]
                deduped[idx] = pick_better(deduped[idx], r)
            else:
                seen_isin[isin] = len(deduped)
                deduped.append(r)
        else:
            deduped.append(r)
    
    dedup_isin = len(results) - len(deduped)
    results = deduped
    
    total_dedup = dedup_nse + dedup_isin
    if total_dedup > 0:
        print(f"[INFO] Removed {total_dedup} duplicate(s) ({dedup_nse} by NSE_Code, {dedup_isin} by ISIN)")
    
    print(f"[INFO] Successfully analyzed {len(results)} stocks")
    
    # =====================================================
    # COMPANION DATA COVERAGE (quick validation at a glance)
    # =====================================================
    sh_count = sum(1 for s in results if s.get('Shareholding_Available') == 'Yes')
    it_count = sum(1 for s in results if s.get('Insider_Trading_Available') == 'Yes')
    ov_count = sum(1 for s in results if s.get('Overview_Available') == 'Yes')
    dv_count = sum(1 for s in results if s.get('Dividend_Count', 0) is not None and s.get('Dividend_Count', 0) > 0)
    total = len(results)
    
    print(f"[INFO] Companion data coverage ({total} stocks):")
    print(f"       Shareholding:    {sh_count:>4} / {total}  ({100*sh_count/total:.0f}%)")
    print(f"       Overview (TTM):  {ov_count:>4} / {total}  ({100*ov_count/total:.0f}%)")
    print(f"       Insider Trading: {it_count:>4} / {total}  ({100*it_count/total:.0f}%)")
    print(f"       Dividends:       {dv_count:>4} / {total}  ({100*dv_count/total:.0f}%)")
    
    # =====================================================
    # SECTOR-RELATIVE SCORING (Post-processing)
    # =====================================================
    # Adjusts Business_Quality and Valuation_Comfort scores based on how each
    # stock compares to its sector peers. An ROE of 15% means very different
    # things in FMCG vs Infrastructure — this corrects for that.
    t_sector = time.time()
    results = apply_sector_relative_scoring(results)
    print(f"[INFO] Sector-relative adjustment in {time.time() - t_sector:.2f}s")
    
    # Sort by: Quality_Risk first (LOW before HIGH), then Decision_Bucket, then Score_Band
    # This ensures LOW risk stocks appear first, regardless of score
    risk_priority = {'LOW': 0, 'MEDIUM': 1, 'HIGH': 2}
    bucket_priority = {
        'GATES_CLEARED': 0,
        'SCREEN_PASSED_VERIFY': 1,      # Passed but needs group verification (standalone)
        'SCREEN_PASSED_EXPENSIVE': 2,
        'CONTRARIAN_BET': 3,            # Failed score but insider accumulation — special situation
        'SCREEN_PASSED_FLAGS': 4,
        'VALUE_TRAP': 5,                # Good score but insider exit — investigate
        'SCREEN_MARGINAL': 6,
        'DATA_INCOMPLETE': 7,
        'SCREEN_FAILED': 8
    }
    band_priority = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'F': 4}
    
    results.sort(key=lambda x: (
        risk_priority.get(x.get('Quality_Risk', 'HIGH'), 2),
        bucket_priority.get(x.get('Decision_Bucket', 'SCREEN_FAILED'), 6),
        band_priority.get(x.get('Score_Band', 'F'), 4),
        -x.get('Composite_Score', 0)
    ))
    
    # Compute summary statistics ONCE (used by both Excel Summary sheet and console output)
    summary_stats = compute_summary_stats(results)
    
    # =====================================================
    # MULTI-SHEET EXCEL OUTPUT - Organized by Category
    # =====================================================
    
    # SHEET 1: ANALYSIS (~15 columns) - Decision-focused only
    analysis_columns = [
        'NSE_Code',
        'Decision_Bucket',
        'Conviction_Override',
        'SCREEN_ELIGIBLE',
        'Investment_Thesis',
        'Reject_Reason',
        'Composite_Score',
        'Sector_Relative_Adj',
        'Score_Band',
        'Quality_Risk',
        'Critical_Flags',
        'Major_Flags',
        'Primary_Concern',
        'Market_Cap',
        'PE',
        'Piotroski_Score',
        'Generic_Stock_Candidate',
        'Financial_Analysis_Gap',
    ]
    
    # SHEET 2: MASTER - Identity and metadata
    master_columns = [
        'NSE_Code',
        'Stock_Name',
        'Sector',
        'Industry',
        'BSE_Code',
        'ISIN',
        'Market_Cap',
        'Data_Type',
        'Years_Available',
        'Latest_Year',
        'Fiscal_Year_End',
        'Overview_Available',
        'Shareholding_Available',
        'Corporate_Actions_Available',
        # Financial sector handling (Fix 3.1)
        'Is_Financial_Sector',
        'Financial_Sector_Bypass',
        'Requires_Manual_Review',
    ]
    
    # SHEET 3: VALUATION - All valuation metrics
    valuation_columns = [
        'NSE_Code',
        # Primary Valuation (TTM preferred)
        'PE', 'PE_TTM', 'PE_Annual',
        'PBV',
        'EV_EBITDA',
        'PEG', 'PEG_TTM', 'PEG_Computed',
        'Price_To_Sales',
        'Earnings_Yield',
        'Valuation_Band',
        'Valuation_Comfort_Score',
        # Market Data
        'Market_Cap', 'Market_Cap_Annual',
        'Enterprise_Value',
        'LTP',
        # 52-Week Range
        'Week52_High', 'Week52_Low',
        'Price_Position_52W',
        'Distance_From_52W_High',
        # Returns
        'Return_1Yr',
        'Returns_vs_Nifty50_Qtr',
        'Returns_vs_Sector_Qtr',
        'Return_vs_Nifty_1Yr',
    ]
    
    # SHEET 4: QUALITY - Profitability and quality scores
    quality_columns = [
        'NSE_Code',
        # Quality Scores
        'Business_Quality_Score',
        'Earnings_Quality',
        'Piotroski_Score', 'Piotroski_Assessment',
        # ROE
        'ROE_Latest', 'ROE_3Yr_Avg', 'ROE_5Yr_Avg', 'ROE_Trend',
        # ROCE
        'ROCE_Latest', 'ROCE_3Yr_Avg', 'ROCE_5Yr_Avg', 'ROCE_Trend',
        # Other Profitability
        'ROA_Latest',
        'OPM_Latest', 'OPM_TTM', 'OPM_Trend',
        'NPM_Latest', 'NPM_Trend',
        'EBITDA', 'EBITDA_Margin',
        # Leverage Analysis
        'ROE_ROA_Gap', 'Leverage_Driven',
        # Earnings Quality Details
        'Exceptional_Items_Count', 'Exceptional_Items_Latest',
        'Other_Income_Pct_PAT', 'High_Other_Income_Years',
    ]
    
    # SHEET 5: CASH_FLOW - Cash flow metrics
    cashflow_columns = [
        'NSE_Code',
        # Primary Cash Flow
        'CFO_Latest',
        'PAT_Latest',
        'CFO_PAT_Latest',
        'CFO_PAT_3Yr_Avg',
        # Cash Flow Quality
        'Positive_CFO_Years',
        'CFO_Trend',
        'CFROA',
        'Accruals',
        # Other Cash Flow
        'CFI',
        'CEPS',
        # Working Capital Impact
        'WC_Growth_Pct',
        'Rev_Growth_Pct',
        'WC_Rev_Growth_Ratio',
    ]
    
    # SHEET 6: LEVERAGE - Debt and balance sheet
    leverage_columns = [
        'NSE_Code',
        # Debt Ratios
        'Debt_Equity',
        'LT_Debt_Equity',
        'Net_Debt_EBITDA',
        'Interest_Coverage',
        'Debt_Trend',
        'Financial_Strength_Score',
        # Liquidity
        'Current_Ratio',
        'Quick_Ratio',
        # Balance Sheet
        'Total_Debt',
        'Total_Equity',
        'Equity_Capital',
        'Total_Assets',
        # Working Capital
        'Working_Capital',
        'WC_Turnover',
        'Asset_Turnover',
        'Capital_Intensity',
        # Efficiency
        'Receivable_Days', 'Recv_Days_Trend',
        'Inventory_Days', 'Inv_Days_Trend',
        'Payables_Days',
        'Cash_Conversion_Cycle',
        'Trade_Payables',
    ]
    
    # SHEET 7: GROWTH - Growth metrics
    growth_columns = [
        'NSE_Code',
        # TTM Growth (most current)
        'Revenue_Growth_TTM',
        'NP_Growth_TTM',
        'Revenue_Growth_Qtr_YoY',
        'NP_Growth_Qtr_YoY',
        # Annual Growth
        'Revenue_Growth_1Yr',
        'Revenue_Growth_2Yr',
        'Revenue_Growth_3Yr',
        'Revenue_CAGR_3Yr',
        'NP_Growth',
        'EPS_Growth_3Yr',
        # Growth Quality
        'Revenue_Volatility',
        'Profit_Consistency',
        'Profit_Growth_Consistency',
        'Growth_Durability_Score',
        # Size
        'Total_Revenue',
        'Book_Value_Per_Share',
        'EPS',
        # Dilution
        'Share_Dilution_Pct',
        # Capital Allocation
        'Dividend_Per_Share',
        'Dividend_Payout_NP',
        'Retention_Ratio',
        'ROIC',
    ]
    
    # SHEET 8: SHAREHOLDING - Holdings, changes, and insider trading context
    shareholding_columns = [
        'NSE_Code',
        'Shareholding_Available',
        # Promoter
        'Promoter_Holding',
        'Promoter_Pledge',
        'Pledge_Risk',  # Fix 3.6: Pledge severity assessment
        'Promoter_Change_1Yr',
        # Institutional
        'FII_Holding',
        'FII_Change_1Yr',
        'MF_Holding',
        'MF_Change_1Yr',
        'Institutional_Holding',
        'Institutional_Holding_Overview',
        # Others
        'Public_Holding',
        'Others_Holding',
        # Governance Assessment
        'Governance_Flags',
        # Insider Trading Context (cross-reference for promoter changes)
        'Insider_Trading_Available',
        'Insider_Action',
        'Insider_Sentiment',
        'Inter_Se_Pct_Of_Disposals',
        'Market_Sell_Pct_MCap',
        'Recent_Market_Buy_6m',
        'Insider_Context',
    ]
    
    # SHEET 9: NEGLECTED_FIRM - Generic stock strategy
    neglected_firm_columns = [
        'NSE_Code',
        'Generic_Stock_Candidate',
        'Neglect_Score',
        'Neglect_Age_Warning',  # Fix 3.5: Age dimension for neglect
        'Neglect_Reasons',
        # Key Criteria
        'Institutional_Holding',
        'Years_Available',  # For age context
        'CFO_Debt_Ratio',
        'Total_Debt',
        'Revenue_Growth_Years',
        'Revenue_Growth_Consistency',
        # Supporting Data
        'ROE_5Yr_Avg',
        'Debt_Equity',
        'Market_Cap',
        'Generic_Criteria_Detail',
    ]
    
    # SHEET 10: DIVIDENDS - Dividend history (slimmed from Corporate_Actions)
    dividends_columns = [
        'NSE_Code',
        'Corporate_Actions_Available',
        # Dividend Summary
        'Dividend_Count',
        'Latest_Dividend_Date',
        'Latest_Dividend_Amount',
        'Latest_Dividend_Type',
        'Years_Since_Dividend',
        'Dividend_5Yr_Consistency',
        'Dividend_Trend',
        'Total_Dividends_Paid',
        'Dividend_Paying_Years',
        # Bonus Summary
        'Bonus_Count',
        'Latest_Bonus_Date',
        'Latest_Bonus_Ratio',
        # Split Summary
        'Split_Count',
        'Latest_Split_Date',
        'Latest_Split_Ratio',
        # Rights Issue Summary (important for dilution)
        'Rights_Count',
        'Latest_Rights_Date',
        'Latest_Rights_Ratio',
        'Latest_Rights_Premium',
    ]
    
    # SHEET 11: RED_FLAGS - All flag details
    red_flags_columns = [
        'NSE_Code',
        # Flag Summary
        'Quality_Risk',
        'Quality_Severity',
        'Quality_Flag_Count',
        'Pricing_Flag_Count',
        'Red_Flag_Count',
        # Flag Details
        'Quality_Flags',
        'Pricing_Flags',
        'Red_Flags',
        'Red_Flags_Explained',
        'Sector_Warnings',
        'Sector_Adjustment_Needed',
        'Sector_Adjustments_Made',
        # Individual Structural Flags
        'FLAG_LOW_ROE',
        'FLAG_DECLINING_ROE',
        'FLAG_LOW_ROCE',
        'FLAG_DECLINING_ROCE',
        'FLAG_POOR_CASH_CONVERSION',
        'FLAG_NEGATIVE_CFO',
        'FLAG_INCONSISTENT_CFO',
        'FLAG_FREQUENT_EXCEPTIONALS',
        'FLAG_HIGH_OTHER_INCOME',
        'FLAG_RISING_RECEIVABLES',
        'FLAG_RISING_INVENTORY',
        'FLAG_MARGIN_COMPRESSION',
        'FLAG_RISING_DEBT',
        'FLAG_WC_DIVERGENCE',
        'FLAG_NPM_OPM_DIVERGENCE',
        # Individual Pricing Flags
        'FLAG_HIGH_PE',
        'FLAG_NEGATIVE_PE',
        'FLAG_HIGH_EV_EBITDA',
        'FLAG_NEGATIVE_EBITDA',
        'FLAG_HIGH_PBV_ROE',
    ]
    
    # Write Excel file with multiple sheets using openpyxl
    t_excel_start = time.time()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Define color schemes
        bucket_colors = {
            'GATES_CLEARED': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            'SCREEN_PASSED_VERIFY': PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
            'SCREEN_PASSED_EXPENSIVE': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'CONTRARIAN_BET': PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid"),  # Purple — special situation
            'SCREEN_PASSED_FLAGS': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'VALUE_TRAP': PatternFill(start_color="E06666", end_color="E06666", fill_type="solid"),  # Dark red — danger
            'SCREEN_MARGINAL': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            'DATA_INCOMPLETE': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
            'SCREEN_FAILED': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        }
        
        header_font_white = Font(bold=True, color="FFFFFF")
        header_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_fill_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        header_fill_purple = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_fill_teal = PatternFill(start_color="00B0B0", end_color="00B0B0", fill_type="solid")
        header_fill_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        def write_sheet(wb, title, columns, data, header_fill, is_first=False, color_rules=None):
            """Helper to write a sheet with standard formatting.
            
            color_rules: dict mapping column_name -> list of (condition_fn, fill) tuples.
                        Applied during a SINGLE pass over all rows (not per-column separate passes).
            """
            if is_first:
                ws = wb.active
                ws.title = title
            else:
                ws = wb.create_sheet(title=title)
            
            # Write header
            ws.append(columns)
            for col_num, cell in enumerate(ws[1], 1):
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Pre-resolve color rule column indices for O(1) lookup during row iteration
            resolved_rules = {}
            if color_rules:
                for col_name, rules in color_rules.items():
                    if col_name in columns:
                        resolved_rules[columns.index(col_name)] = rules  # 0-based index
            
            # Write data + apply color rules in SINGLE PASS (no separate row iterations)
            for stock in data:
                row = [stock.get(col, '') for col in columns]
                ws.append(row)
                
                if resolved_rules:
                    row_num = ws.max_row
                    for col_0idx, rules in resolved_rules.items():
                        cell = ws.cell(row=row_num, column=col_0idx + 1)
                        value = cell.value
                        for condition, fill in rules:
                            if condition(value):
                                cell.fill = fill
                                break
            
            # Auto-fit column widths
            for col_num, col_name in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = max(12, min(25, len(str(col_name)) + 2))
            
            return ws
        
        # Color rule helpers
        def is_bucket(val):
            return val in bucket_colors
        
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        # Sheet 1: Analysis (Decision-focused) — color rules applied during write (single pass)
        def _safe_int(v, default=0):
            try: return int(v) if v else default
            except (ValueError, TypeError): return default
        
        def _safe_float(v, default=0.0):
            try: return float(v) if v else default
            except (ValueError, TypeError): return default
        
        analysis_color_rules = {
            'Piotroski_Score': [
                (lambda v: _safe_int(v) >= 7, green_fill),
                (lambda v: _safe_int(v) >= 5, yellow_fill),
                (lambda v: _safe_int(v) >= 3, orange_fill),
                (lambda v: True, red_fill),
            ],
        }
        # Decision_Bucket needs per-bucket color — bind k as default arg to avoid late-binding
        analysis_color_rules['Decision_Bucket'] = [
            (lambda v, k=k: v == k, f) for k, f in bucket_colors.items()
        ]
        
        ws_analysis = write_sheet(wb, "Analysis", analysis_columns, results, header_fill_blue,
                                  is_first=True, color_rules=analysis_color_rules)
        
        # Sheet 2: Master
        write_sheet(wb, "Master", master_columns, results, header_fill_green)
        
        # Sheet 3: Valuation
        write_sheet(wb, "Valuation", valuation_columns, results, header_fill_blue)
        
        # Sheet 4: Quality
        write_sheet(wb, "Quality", quality_columns, results, header_fill_green)
        
        # Sheet 5: Cash_Flow
        write_sheet(wb, "Cash_Flow", cashflow_columns, results, header_fill_teal)
        
        # Sheet 6: Leverage
        write_sheet(wb, "Leverage", leverage_columns, results, header_fill_orange)
        
        # Sheet 7: Growth
        write_sheet(wb, "Growth", growth_columns, results, header_fill_green)
        
        # Sheet 8: Shareholding — color rules applied during write
        shareholding_color_rules = {
            'Promoter_Pledge': [
                (lambda v: _safe_float(v) > 50, red_fill),
                (lambda v: _safe_float(v) > 20, orange_fill),
                (lambda v: _safe_float(v) > 0, yellow_fill),
            ],
        }
        ws_sh = write_sheet(wb, "Shareholding", shareholding_columns, results, header_fill_purple,
                            color_rules=shareholding_color_rules)
        
        # Sheet 9: Neglected_Firm — color rules applied during write
        neglected_color_rules = {
            'Generic_Stock_Candidate': [
                (lambda v: v == 'YES', green_fill),
            ],
        }
        ws_nf = write_sheet(wb, "Neglected_Firm", neglected_firm_columns, results, header_fill_teal,
                            color_rules=neglected_color_rules)
        
        # Sheet 10: Dividends — color rules applied during write
        dividends_color_rules = {
            'Dividend_5Yr_Consistency': [
                (lambda v: _safe_int(v) >= 4, green_fill),
                (lambda v: _safe_int(v) >= 2, yellow_fill),
                (lambda v: _safe_int(v) == 0, red_fill),
            ],
        }
        ws_div = write_sheet(wb, "Dividends", dividends_columns, results, header_fill_purple,
                             color_rules=dividends_color_rules)
        
        # Sheet 11: Red_Flags — color rules applied during write
        red_flags_color_rules = {
            'Quality_Risk': [
                (lambda v: v == 'LOW', green_fill),
                (lambda v: v == 'MEDIUM', orange_fill),
                (lambda v: v == 'HIGH', red_fill),
            ],
        }
        ws_rf = write_sheet(wb, "Red_Flags", red_flags_columns, results, header_fill_red,
                            color_rules=red_flags_color_rules)
        
        # =====================================================
        # Sheet 12: SUMMARY (Decision Summary Report)
        # =====================================================
        ws_summary = wb.create_sheet(title="Summary")
        
        # Generate summary statistics (use pre-computed stats - single pass over results)
        screen_passed = summary_stats['screen_passed']
        screen_verify = summary_stats['screen_verify']
        screen_expensive = summary_stats['screen_expensive']
        screen_flags = summary_stats['screen_flags']
        screen_marginal = summary_stats['screen_marginal']
        data_incomplete = summary_stats['data_incomplete']
        screen_failed = summary_stats['screen_failed']
        
        eligible = summary_stats['eligible']
        blocked = summary_stats['blocked']
        
        low_risk = summary_stats['low_risk']
        medium_risk = summary_stats['medium_risk']
        high_risk = summary_stats['high_risk']
        
        generic_candidates = summary_stats['generic_candidates']
        high_neglect = summary_stats['high_neglect']
        medium_neglect = summary_stats['medium_neglect']
        
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        header_font_summary = Font(bold=True, size=11)
        
        ws_summary['A1'] = "STOCK SCREENING DECISION SUMMARY"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:F1')
        ws_summary['A2'] = "(This is a SCREEN, not a Recommendation)"
        ws_summary['A2'].font = Font(italic=True, size=10)
        
        row = 4
        
        # Section 1: Screen Eligibility Gate
        ws_summary.cell(row=row, column=1, value="SCREEN ELIGIBILITY GATE").font = section_font
        row += 1
        ws_summary.cell(row=row, column=1, value="Category").font = header_font_summary
        ws_summary.cell(row=row, column=2, value="Count").font = header_font_summary
        ws_summary.cell(row=row, column=3, value="Description").font = header_font_summary
        row += 1
        
        ws_summary.cell(row=row, column=1, value="✅ ELIGIBLE")
        ws_summary.cell(row=row, column=2, value=len(eligible))
        ws_summary.cell(row=row, column=3, value="Passed quality gates")
        ws_summary.cell(row=row, column=2).fill = green_fill
        row += 1
        
        ws_summary.cell(row=row, column=1, value="🚫 BLOCKED")
        ws_summary.cell(row=row, column=2, value=len(blocked))
        ws_summary.cell(row=row, column=3, value="Failed hard rule")
        ws_summary.cell(row=row, column=2).fill = red_fill
        row += 2
        
        # Section 2: Decision Buckets
        ws_summary.cell(row=row, column=1, value="DECISION BUCKETS").font = section_font
        row += 1
        ws_summary.cell(row=row, column=1, value="Bucket").font = header_font_summary
        ws_summary.cell(row=row, column=2, value="Count").font = header_font_summary
        ws_summary.cell(row=row, column=3, value="Description").font = header_font_summary
        row += 1
        
        bucket_data = [
            ("GATES_CLEARED", len(screen_passed), "Passed quality gates"),
            ("SCREEN_PASSED_VERIFY", len(screen_verify), "Passed—verify group financials"),
            ("SCREEN_PASSED_EXPENSIVE", len(screen_expensive), "Passed but expensive"),
            ("CONTRARIAN_BET", len(summary_stats['contrarian_bet']), "Low score + promoter buying—special situation"),
            ("SCREEN_PASSED_FLAGS", len(screen_flags), "Passed with concerns"),
            ("VALUE_TRAP", len(summary_stats['value_trap']), "High score + promoter selling—insider exit"),
            ("SCREEN_MARGINAL", len(screen_marginal), "Borderline"),
            ("DATA_INCOMPLETE", len(data_incomplete), "Verify manually"),
            ("SCREEN_FAILED", len(screen_failed), "Failed quality gates"),
        ]
        
        for bucket_name, count, desc in bucket_data:
            ws_summary.cell(row=row, column=1, value=bucket_name)
            ws_summary.cell(row=row, column=2, value=count)
            ws_summary.cell(row=row, column=3, value=desc)
            if bucket_name in bucket_colors:
                ws_summary.cell(row=row, column=1).fill = bucket_colors[bucket_name]
                ws_summary.cell(row=row, column=2).fill = bucket_colors[bucket_name]
            row += 1
        row += 1
        
        # Section 3: Risk Distribution
        ws_summary.cell(row=row, column=1, value="QUALITY RISK").font = section_font
        row += 1
        ws_summary.cell(row=row, column=1, value="🟢 LOW"); ws_summary.cell(row=row, column=2, value=len(low_risk))
        ws_summary.cell(row=row, column=2).fill = green_fill
        row += 1
        ws_summary.cell(row=row, column=1, value="🟡 MEDIUM"); ws_summary.cell(row=row, column=2, value=len(medium_risk))
        ws_summary.cell(row=row, column=2).fill = orange_fill
        row += 1
        ws_summary.cell(row=row, column=1, value="🔴 HIGH"); ws_summary.cell(row=row, column=2, value=len(high_risk))
        ws_summary.cell(row=row, column=2).fill = red_fill
        row += 2
        
        # Section 4: Generic Stock Candidates
        ws_summary.cell(row=row, column=1, value="GENERIC STOCK CANDIDATES").font = section_font
        row += 1
        ws_summary.cell(row=row, column=1, value="HIGH Neglect"); ws_summary.cell(row=row, column=2, value=len(high_neglect))
        row += 1
        ws_summary.cell(row=row, column=1, value="MEDIUM Neglect"); ws_summary.cell(row=row, column=2, value=len(medium_neglect))
        row += 1
        ws_summary.cell(row=row, column=1, value="Candidates (✅)"); ws_summary.cell(row=row, column=2, value=len(generic_candidates))
        ws_summary.cell(row=row, column=2).fill = green_fill
        row += 2
        
        # Section 5: GATES_CLEARED List
        if screen_passed:
            ws_summary.cell(row=row, column=1, value="GATES_CLEARED STOCKS").font = section_font
            row += 1
            headers = ['Rank', 'NSE_Code', 'Market_Cap', 'Risk', 'Score', 'Band', 'Thesis']
            for col_idx, header in enumerate(headers, 1):
                ws_summary.cell(row=row, column=col_idx, value=header).font = header_font_summary
                ws_summary.cell(row=row, column=col_idx).fill = green_fill
            row += 1
            
            risk_priority = {'LOW': 0, 'MEDIUM': 1, 'HIGH': 2}
            screen_passed.sort(key=lambda x: (risk_priority.get(x.get('Quality_Risk', 'HIGH'), 2), -x.get('Composite_Score', 0)))
            
            for rank, stock in enumerate(screen_passed[:30], 1):
                ws_summary.cell(row=row, column=1, value=rank)
                ws_summary.cell(row=row, column=2, value=stock.get('NSE_Code', ''))
                ws_summary.cell(row=row, column=3, value=stock.get('Market_Cap', ''))
                ws_summary.cell(row=row, column=4, value=stock.get('Quality_Risk', ''))
                ws_summary.cell(row=row, column=5, value=stock.get('Composite_Score', ''))
                ws_summary.cell(row=row, column=6, value=stock.get('Score_Band', ''))
                thesis = stock.get('Investment_Thesis', '') or ''
                ws_summary.cell(row=row, column=7, value=thesis[:60] + '...' if len(thesis) > 60 else thesis)
                row += 1
            row += 1
        
        # Section 6: Generic Stock Candidates List
        if generic_candidates:
            ws_summary.cell(row=row, column=1, value="GENERIC STOCK CANDIDATES").font = section_font
            row += 1
            headers = ['Rank', 'NSE_Code', 'Market_Cap', 'Neglect', 'Inst%', 'ROE_5Y', 'D/E', 'CFO/Debt']
            for col_idx, header in enumerate(headers, 1):
                ws_summary.cell(row=row, column=col_idx, value=header).font = header_font_summary
                ws_summary.cell(row=row, column=col_idx).fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            row += 1
            
            for rank, stock in enumerate(generic_candidates[:20], 1):
                ws_summary.cell(row=row, column=1, value=rank)
                ws_summary.cell(row=row, column=2, value=stock.get('NSE_Code', ''))
                ws_summary.cell(row=row, column=3, value=stock.get('Market_Cap', ''))
                ws_summary.cell(row=row, column=4, value=stock.get('Neglect_Score', ''))
                ws_summary.cell(row=row, column=5, value=stock.get('Institutional_Holding', ''))
                ws_summary.cell(row=row, column=6, value=stock.get('ROE_5Yr_Avg') or stock.get('ROE_3Yr_Avg', ''))
                ws_summary.cell(row=row, column=7, value=stock.get('Debt_Equity', ''))
                ws_summary.cell(row=row, column=8, value=stock.get('CFO_Debt_Ratio', ''))
                row += 1
        
        # Adjust column widths for Summary
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 12
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 10
        ws_summary.column_dimensions['E'].width = 10
        ws_summary.column_dimensions['F'].width = 10
        ws_summary.column_dimensions['G'].width = 60
        ws_summary.column_dimensions['H'].width = 12
        
        # Save Excel file
        wb.save(OUTPUT_FILE)
        
        # Print summary of sheets
        sheet_info = [
            ('Analysis', len(analysis_columns), 'Decision-focused'),
            ('Master', len(master_columns), 'Identity & metadata'),
            ('Valuation', len(valuation_columns), 'PE, PBV, 52W range'),
            ('Quality', len(quality_columns), 'ROE, ROCE, Piotroski'),
            ('Cash_Flow', len(cashflow_columns), 'CFO, accruals'),
            ('Leverage', len(leverage_columns), 'Debt, working capital'),
            ('Growth', len(growth_columns), 'Revenue, EPS growth'),
            ('Shareholding', len(shareholding_columns), 'Promoter, FII, MF'),
            ('Neglected_Firm', len(neglected_firm_columns), 'Generic stock strategy'),
            ('Dividends', len(dividends_columns), 'Dividend history'),
            ('Red_Flags', len(red_flags_columns), 'All flags'),
            ('Summary', '-', 'Decision report'),
        ]
        
        print(f"[INFO] Excel output written to: {OUTPUT_FILE} ({time.time() - t_excel_start:.1f}s)")
        print(f"       12 Sheets:")
        for name, cols, desc in sheet_info:
            col_str = f"{cols} cols" if cols != '-' else "Report"
            print(f"         • {name:<15} ({col_str:<10}) - {desc}")
        
    except ImportError:
        # Fallback to CSV if openpyxl not available
        print("[WARN] openpyxl not installed, falling back to CSV output")
        OUTPUT_FILE_CSV = OUTPUT_FILE.replace('.xlsx', '.csv')
        # Use all known columns from the sheet definitions
        all_columns = list(dict.fromkeys(
            analysis_columns + master_columns + valuation_columns + quality_columns +
            cashflow_columns + leverage_columns + growth_columns + shareholding_columns +
            neglected_firm_columns + dividends_columns + red_flags_columns
        ))
        with open(OUTPUT_FILE_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=all_columns, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(results)
        print(f"[INFO] CSV output written to: {OUTPUT_FILE_CSV}")
    
    # Print decision summary
    print("\n" + "="*100)
    print("DECISION SUMMARY (Neutral Language - This is a SCREEN, not a Recommendation)")
    print("="*100)
    
    # SCREEN ELIGIBILITY GATE (most important - hard mechanical rule)
    screen_eligible = summary_stats['eligible']
    screen_blocked = summary_stats['blocked']
    
    print(f"\n🚦 SCREEN ELIGIBILITY GATE (Hard mechanical rule - NO DISCRETION)")
    print(f"   ✅ ELIGIBLE:   {len(screen_eligible):>3}  (Passed quality gates)")
    print(f"   🚫 BLOCKED:    {len(screen_blocked):>3}  (Failed hard rule)")
    
    # Blocker reasons breakdown
    if screen_blocked:
        blocker_reasons = {}
        for s in screen_blocked:
            reason = s['Screen_Blocker']
            blocker_reasons[reason] = blocker_reasons.get(reason, 0) + 1
        print(f"\n   Blocked by:")
        for reason, count in sorted(blocker_reasons.items(), key=lambda x: -x[1]):
            print(f"     • {reason}: {count}")
    
    # Decision buckets with NEUTRAL LANGUAGE (use pre-computed stats)
    screen_passed = summary_stats['screen_passed']
    screen_verify = summary_stats['screen_verify']
    screen_expensive = summary_stats['screen_expensive']
    screen_flags = summary_stats['screen_flags']
    data_incomplete = summary_stats['data_incomplete']
    screen_marginal = summary_stats['screen_marginal']
    screen_failed = summary_stats['screen_failed']
    contrarian_bets = summary_stats['contrarian_bet']
    value_traps = summary_stats['value_trap']
    
    print(f"\n📊 DECISION BUCKETS (Severity-Based)")
    print(f"   ✅ GATES_CLEARED:          {len(screen_passed):>3}  (Passed quality gates)")
    print(f"   🔍 SCREEN_PASSED_VERIFY:   {len(screen_verify):>3}  (Passed—verify group financials)")
    print(f"   💰 SCREEN_PASSED_EXPENSIVE:{len(screen_expensive):>3}  (Passed but expensive)")
    print(f"   🟣 CONTRARIAN_BET:         {len(contrarian_bets):>3}  (Low score + promoter buying—special situation)")
    print(f"   ⚠️  SCREEN_PASSED_FLAGS:    {len(screen_flags):>3}  (Passed with concerns)")
    print(f"   🪤 VALUE_TRAP:             {len(value_traps):>3}  (High score + promoter selling—insider exit)")
    print(f"   📊 DATA_INCOMPLETE:        {len(data_incomplete):>3}  (Verify manually)")
    print(f"   📉 SCREEN_MARGINAL:        {len(screen_marginal):>3}  (Borderline)")
    print(f"   ❌ SCREEN_FAILED:          {len(screen_failed):>3}  (Failed quality gates)")
    
    # Insider Trading Impact (shows how insider trading context affected decisions)
    vt_mitigated = summary_stats['value_trap_mitigated']
    cb_confirmed = summary_stats['contrarian_confirmed']
    it_stocks = summary_stats['has_insider_trading']
    
    if it_stocks or vt_mitigated or value_traps:
        print(f"\n🔍 INSIDER TRADING CONTEXT (Conviction Override Impact)")
        print(f"   📊 Stocks with insider trading data: {len(it_stocks):>3} / {len(results)}")
        if vt_mitigated:
            print(f"   ✅ VALUE_TRAP mitigated:    {len(vt_mitigated):>3}  (Promoter drop was inter-se/benign)")
            for s in vt_mitigated[:5]:
                override = s.get('Conviction_Override', '')[:80]
                print(f"      • {s['NSE_Code']:<12} {override}")
        if value_traps:
            print(f"   🪤 VALUE_TRAP confirmed:    {len(value_traps):>3}  (Genuine promoter exit)")
            for s in value_traps[:5]:
                override = s.get('Conviction_Override', '')[:80]
                print(f"      • {s['NSE_Code']:<12} {override}")
        if cb_confirmed:
            print(f"   🟣 CONTRARIAN confirmed:    {len(cb_confirmed):>3}  (Open-market buying verified)")
        
        # Insider action distribution (only if meaningful count)
        if len(it_stocks) >= 5:
            print(f"\n   Insider Action Distribution ({len(it_stocks)} stocks):")
            action_labels = [
                ('it_only_buying', 'ONLY_BUYING', '🟢'),
                ('it_net_buyer', 'NET_BUYER', '🟢'),
                ('it_inter_se_only', 'INTER_SE_ONLY', '🔵'),
                ('it_no_activity', 'NO_ACTIVITY', '⚪'),
                ('it_net_seller_with_buys', 'NET_SELLER+BUYS', '🟡'),
                ('it_net_seller', 'NET_SELLER', '🔴'),
            ]
            for key, label, emoji in action_labels:
                count = len(summary_stats[key])
                if count > 0:
                    print(f"   {emoji} {label:<20}: {count:>3}")
    
    # Structural Risk distribution - THIS IS THE KEY METRIC
    print("\n" + "-"*50)
    print("QUALITY RISK DISTRIBUTION (Sort by this!)")
    print("-"*50)
    for risk in ['LOW', 'MEDIUM', 'HIGH']:
        count = sum(1 for s in results if s['Quality_Risk'] == risk)
        bar = '█' * (count // 2) if count > 0 else ''
        emoji = '🟢' if risk == 'LOW' else ('🟡' if risk == 'MEDIUM' else '🔴')
        print(f"{emoji} {risk:<8}: {count:>3}  {bar}")
    
    # Score band distribution (secondary)
    print("\n" + "-"*50)
    print("SCORE BAND DISTRIBUTION (Secondary, don't over-index)")
    print("-"*50)
    for band in ['A', 'B', 'C', 'D', 'F']:
        count = sum(1 for s in results if s['Score_Band'] == band)
        bar = '▪' * (count // 2) if count > 0 else ''
        print(f"Band {band}: {count:>3}  {bar}")
    
    # Print top candidates - show sub-scores
    print("\n" + "-"*120)
    print("GATES_CLEARED (sorted by Structural Risk, then score band)")
    print("-"*120)
    print(f"{'Rank':<4} {'NSE Code':<12} {'MCap':<7} {'Risk':<6} {'BizQual':<7} {'FinStr':<7} {'Growth':<7} {'Valuation':<9} {'Comp':<5} {'Thesis'}")
    print("-"*120)
    
    for i, stock in enumerate(screen_passed[:25], 1):
        mcap_str = f"{stock['Market_Cap']:.0f}" if stock['Market_Cap'] else "N/A"
        bq = f"{stock['Business_Quality_Score']:.0f}" if stock['Business_Quality_Score'] else "N/A"
        fs = f"{stock['Financial_Strength_Score']:.0f}" if stock['Financial_Strength_Score'] else "N/A"
        gd = f"{stock['Growth_Durability_Score']:.0f}" if stock['Growth_Durability_Score'] else "N/A"
        vc = f"{stock['Valuation_Comfort_Score']:.0f}" if stock['Valuation_Comfort_Score'] else "N/A"
        thesis = stock['Investment_Thesis'][:35] + "..." if len(stock['Investment_Thesis']) > 35 else stock['Investment_Thesis']
        print(f"{i:<4} {stock['NSE_Code']:<12} {mcap_str:<7} {stock['Quality_Risk']:<6} {bq:<7} {fs:<7} {gd:<7} {vc:<9} {stock['Composite_Score']:<5} {thesis}")
    
    # Print SCREEN_PASSED_EXPENSIVE
    if screen_expensive:
        print("\n" + "-"*100)
        print("SCREEN_PASSED_EXPENSIVE (Passed quality, verify if premium justified)")
        print("-"*100)
        print(f"{'NSE Code':<15} {'MCap':<8} {'Risk':<8} {'PE':<10} {'EV/EBITDA':<12} {'Pricing Flag'}")
        print("-"*100)
        for stock in screen_expensive[:10]:
            mcap_str = f"{stock['Market_Cap']:.0f}" if stock['Market_Cap'] else "N/A"
            pe_str = f"{stock['PE']:.1f}" if stock['PE'] else "N/A"
            ev_str = f"{stock['EV_EBITDA']:.1f}" if stock['EV_EBITDA'] else "N/A"
            print(f"{stock['NSE_Code']:<15} {mcap_str:<8} {stock['Quality_Risk']:<8} {pe_str:<10} {ev_str:<12} {stock['Pricing_Flags']}")
    
    # Print SCREEN_PASSED_FLAGS
    if screen_flags:
        print("\n" + "-"*100)
        print("SCREEN_PASSED_FLAGS (Passed with concerns - verify these issues)")
        print("-"*100)
        print(f"{'NSE Code':<15} {'MCap':<8} {'Severity':<8} {'Structural Flag':<30} {'Sector Adj?'}")
        print("-"*100)
        for stock in screen_flags[:10]:
            mcap_str = f"{stock['Market_Cap']:.0f}" if stock['Market_Cap'] else "N/A"
            severity = f"{stock['Quality_Severity']:.1f}"
            sector_adj = stock.get('Sector_Adjustment_Needed', 'No')
            print(f"{stock['NSE_Code']:<15} {mcap_str:<8} {severity:<8} {stock['Quality_Flags']:<30} {sector_adj}")
    
    # Print failures summary (count by reason)
    if screen_failed:
        print("\n" + "-"*100)
        print("SCREEN_FAILED SUMMARY (by reason)")
        print("-"*100)
        fail_reasons = {}
        for s in screen_failed:
            reason = s['Reject_Reason']
            fail_reasons[reason] = fail_reasons.get(reason, 0) + 1
        for reason, count in sorted(fail_reasons.items(), key=lambda x: -x[1]):
            print(f"  {reason:<40} {count:>3} stocks")
    
    # ==========================================
    # GENERIC STOCK CANDIDATES (Neglected Firm Strategy)
    # ==========================================
    generic_candidates = [s for s in results if s.get('Generic_Stock_Candidate') == 'YES']
    high_neglect = [s for s in results if s.get('Neglect_Score') == 'HIGH']
    medium_neglect = [s for s in results if s.get('Neglect_Score') == 'MEDIUM']
    
    print("\n" + "="*100)
    print("GENERIC STOCK CANDIDATES (Neglected Firm Strategy)")
    print("="*100)
    print("""
    Based on the 'Neglected Firm' effect: stocks with low institutional coverage
    but solid fundamentals can deliver superior risk-adjusted returns when they
    "spring to life" due to increased popularity.
    
    Criteria: Low inst. ownership (<10%) OR small cap + D/E<0.5 + ROE>18% + 
              Consistent revenue growth + No bankruptcy risk + No critical flags
    """)
    
    print(f"\n📊 NEGLECT DISTRIBUTION")
    print(f"   🔴 HIGH Neglect:   {len(high_neglect):>3}  (Very low inst. ownership / micro cap)")
    print(f"   🟡 MEDIUM Neglect: {len(medium_neglect):>3}  (Low inst. ownership / small cap)")
    print(f"   🟢 LOW Neglect:    {len(results) - len(high_neglect) - len(medium_neglect):>3}  (Well-covered)")
    
    print(f"\n✅ GENERIC STOCK CANDIDATES: {len(generic_candidates)}")
    print("   (Neglected stocks that pass ALL quality filters - potential 'diamonds in the rough')")
    
    if generic_candidates:
        print("\n" + "-"*120)
        print(f"{'Rank':<5} {'NSE Code':<12} {'MCap':>8} {'Neglect':>8} {'Inst%':>7} {'ROE_5Y':>7} {'D/E':>6} {'CFO/Debt':>9} {'Rev Cons':>9}")
        print("-"*120)
        
        # Sort by neglect score (HIGH first), then by ROE
        neglect_priority = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
        generic_candidates.sort(key=lambda x: (
            neglect_priority.get(x.get('Neglect_Score', 'LOW'), 2),
            -(x.get('ROE_5Yr_Avg') or x.get('ROE_3Yr_Avg') or 0)
        ))
        
        for i, s in enumerate(generic_candidates[:20], 1):  # Show top 20
            mcap = f"{s.get('Market_Cap', 0):,.0f}" if s.get('Market_Cap') else "N/A"
            neglect = s.get('Neglect_Score', 'N/A')
            inst = f"{s.get('Institutional_Holding', 0):.1f}%" if s.get('Institutional_Holding') is not None else "N/A"
            roe = f"{s.get('ROE_5Yr_Avg') or s.get('ROE_3Yr_Avg', 0):.1f}%" if (s.get('ROE_5Yr_Avg') or s.get('ROE_3Yr_Avg')) else "N/A"
            de = f"{s.get('Debt_Equity', 0):.2f}" if s.get('Debt_Equity') is not None else "N/A"
            cfo_debt = f"{s.get('CFO_Debt_Ratio', 0):.2f}" if s.get('CFO_Debt_Ratio') is not None else "N/A"
            rev_cons = f"{s.get('Revenue_Growth_Consistency', 0):.0f}%" if s.get('Revenue_Growth_Consistency') is not None else "N/A"
            
            print(f"{i:<5} {s['NSE_Code']:<12} {mcap:>8} {neglect:>8} {inst:>7} {roe:>7} {de:>6} {cfo_debt:>9} {rev_cons:>9}")
        
        if len(generic_candidates) > 20:
            print(f"\n   ... and {len(generic_candidates) - 20} more (see Excel for full list)")
    else:
        print("\n   No stocks meet all Generic Stock criteria in current dataset.")
        print("   This could mean: (a) stringent filters, (b) dataset skewed to large caps,")
        print("   or (c) shareholding data not available for most stocks.")
    
    # Print red flag legend with severity tiers
    print("\n" + "="*100)
    print("SEVERITY WEIGHTS (How flags are scored)")
    print("="*100)
    print("""
    CRITICAL = 2.0  (Survival risk - one flag = SCREEN_FAILED)
                    NEGATIVE_CFO, POOR_CASH_CONVERSION, INCONSISTENT_CFO, RISING_DEBT, NEGATIVE_EBITDA
    
    MAJOR    = 1.0  (Structural concern - two flags = SCREEN_FAILED)
                    LOW_ROE, DECLINING_ROE, LOW_ROCE, DECLINING_ROCE, FREQUENT_EXCEPTIONALS,
                    HIGH_OTHER_INCOME, NPM_OPM_DIVERGENCE, NEGATIVE_PE
    
    MINOR    = 0.5  (May be cyclical/sector-normal - check Sector_Warnings)
                    RISING_RECEIVABLES, RISING_INVENTORY, MARGIN_COMPRESSION, WC_DIVERGENCE,
                    HIGH_PE, HIGH_EV_EBITDA, HIGH_PBV_ROE
    
    SCREEN_FAILED triggers: Severity >= 2.0 OR (Cyclic_Peak_Risk=HIGH + Any pricing flag)
    """)
    
    print("=" * 100)
    print("SECTOR-AWARE ADJUSTMENTS (Automatic severity downgrades)")
    print("=" * 100)
    print("""
    These adjustments prevent mechanical rejection of valid business models:
    
    POOR_CASH_CONVERSION  + Asset_Turnover < 0.8  →  CRITICAL (2.0) downgraded to MINOR (0.5)
                            Reason: Capital-intensive businesses have lumpy CFO timing
    
    NEGATIVE_CFO          + Asset_Turnover < 0.5  →  CRITICAL (2.0) downgraded to MAJOR (1.0)
                            Reason: NBFCs, infra, heavy industry may have negative CFO during growth
    
    INCONSISTENT_CFO      + Asset_Turnover < 0.8  →  CRITICAL (2.0) downgraded to MAJOR (1.0)
                            Reason: Project-based businesses have lumpy cash flows
    
    Check 'Sector_Adjustments_Made' column to see which flags were auto-downgraded.
    """)
    
    print("\n" + "="*100)
    print("RED FLAG REFERENCE")
    print("="*100)
    print("\n📛 QUALITY FLAGS (Business Quality Issues):")
    print(f"{'Flag Code':<25} {'Severity':<10} {'Trigger':<35} {'Concern'}")
    print("-"*100)
    for code, defn in STRUCTURAL_RED_FLAGS.items():
        severity = defn.get('severity', 'MAJOR')
        print(f"{code:<25} {severity:<10} {defn['trigger']:<35} {defn['concern'][:45]}")
    
    print("\n💰 PRICING FLAGS (Valuation Concerns):")
    print(f"{'Flag Code':<25} {'Severity':<10} {'Trigger':<35} {'Concern'}")
    print("-"*100)
    for code, defn in PRICING_RED_FLAGS.items():
        severity = defn.get('severity', 'MINOR')
        print(f"{code:<25} {severity:<10} {defn['trigger']:<35} {defn['concern'][:45]}")
    
    # Write red flag reference to separate CSV (with severity)
    with open(RED_FLAGS_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Flag_Code', 'Type', 'Severity', 'Weight', 'Name', 'Trigger_Condition', 'Meaning', 'Concern'])
        for code, defn in STRUCTURAL_RED_FLAGS.items():
            writer.writerow([code, 'STRUCTURAL', defn.get('severity', 'MAJOR'), defn.get('weight', 1.0), 
                           defn['name'], defn['trigger'], defn['meaning'], defn['concern']])
        for code, defn in PRICING_RED_FLAGS.items():
            writer.writerow([code, 'PRICING', defn.get('severity', 'MINOR'), defn.get('weight', 0.5),
                           defn['name'], defn['trigger'], defn['meaning'], defn['concern']])
    
    print(f"\n[INFO] Red flag reference saved to: {RED_FLAGS_FILE}")


if __name__ == "__main__":
    main()
